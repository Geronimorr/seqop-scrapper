"""
Scraper SEQOP v2 – Extração de Aprovações MPD (Tabela Direta)
=============================================================
Acessa https://csdpocos.petrobras.com.br/seqop/#/arquivadas2
e extrai dados de aprovação diretamente dos atributos 'title' das
<span class="aprovador"> na tabela, sem navegar para páginas de detalhe.

Isto elimina o problema de stale elements que acontecia ao abrir/voltar
de páginas de detalhe.

Uso:
    C:\\SharedPython\\venv\\Scripts\\python.exe scraper_seqop.py

Pré-requisitos:
    - Microsoft Edge instalado
    - selenium, openpyxl, pandas no venv compartilhado
    - Arquivo 'pocos_entrada.xlsx' com coluna "poco" (ou editar POCOS_DEFAULT)
"""

# ── Re-launch com o venv compartilhado ──────────────────────────────────────
import subprocess
import sys
from pathlib import Path

_SHARED_PYTHON = Path(r"C:\SharedPython\venv\Scripts\python.exe")

if _SHARED_PYTHON.exists() and Path(sys.executable).resolve() != _SHARED_PYTHON.resolve():
    print(f"[LAUNCH] Relançando com o Python do venv compartilhado: {_SHARED_PYTHON}")
    result = subprocess.run([str(_SHARED_PYTHON)] + sys.argv, check=False)
    sys.exit(result.returncode)

# ── Imports ─────────────────────────────────────────────────────────────────
import logging
import os
import re
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    WebDriverException,
)

# ── Configuração ────────────────────────────────────────────────────────────

BASE_URL = "https://csdpocos.petrobras.com.br/seqop"
ARQUIVADAS_URL = f"{BASE_URL}/#/arquivadas2"

SCRIPT_DIR = Path(__file__).resolve().parent
ENTRADA_XLSX = SCRIPT_DIR / "pocos_entrada.xlsx"
SAIDA_XLSX = SCRIPT_DIR / "resultado_aprovacoes_mpd.xlsx"

DEBUGGING_PORT = 9222
WAIT_TIMEOUT = 25
PAUSA = 1.5

POCOS_DEFAULT = [
    "1-APS-57", "1-RJS-763DA", "3-SPS-114", "3-SPS-114A", "4-RJS-764",
    "7-BR-86DB-RJS", "7-BUZ-100DA-RJS", "7-BUZ-94D-RJS", "7-BUZ-95-RJS",
    "8-BUZ-89D-RJS", "7-OATP-3B-RJS", "8-BUZ-96D-RJS", "8-MRO-36-RJS",
    "3-SPS-111D", "4-SPS-112", "3-RJS-762", "8-ATP-8D-RJS",
    "7-STUP-10DA-RJS", "7-BUZ-90D-RJS",
]

# Regex para extrair aprovação do atributo title das spans
# Formato: "Aprovado em 21/02/2026 17:56 por Renato Marques Lamounier Fernandes"
RE_APROVACAO = re.compile(
    r"Aprovado\s+em\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})\s+por\s+(.+)",
    re.IGNORECASE,
)

# ── Logging ─────────────────────────────────────────────────────────────────

LOG_FILE = SCRIPT_DIR / "scraper_seqop.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("seqop_scraper")


# ── Dataclass para resultados ───────────────────────────────────────────────

@dataclass
class AprovacaoMPD:
    poco: str
    seq_id: str
    titulo: str
    publicado_em: str
    versao: str
    aprovado_por: str
    aprovado_em: str
    papel_aprovador: str   # badge: TAB, EQSB, SIP, SF, MPD …
    fonte_url: str
    carimbo_extracao_utc: str = field(
        default_factory=lambda: datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    )


# ── Índices das colunas (0-based, = aria-colindex − 1) ─────────────────────
# Sonda(0) | Polo(1) | Poço(2) | Interv(3) | Título(4) |
# Versão(5) | Publicado(6) | Aprovadores(7) | Ações(8)
COL_POCO = 2
COL_TITULO = 4
COL_VERSAO = 5
COL_PUBLICADO = 6
COL_APROVADORES = 7


# ── Funções auxiliares ──────────────────────────────────────────────────────

def carregar_pocos() -> list[str]:
    """Carrega lista de poços do Excel ou usa a lista padrão."""
    if ENTRADA_XLSX.exists():
        log.info(f"Carregando poços de {ENTRADA_XLSX}")
        df = pd.read_excel(ENTRADA_XLSX, engine="openpyxl")
        col = None
        for c in df.columns:
            if c.strip().lower().replace("ç", "c") in ("poco", "poço"):
                col = c
                break
        if col is None:
            log.warning("Coluna 'poco' não encontrada. Usando lista padrão.")
            return POCOS_DEFAULT
        pocos = df[col].dropna().astype(str).str.strip().tolist()
        log.info(f"  → {len(pocos)} poços carregados")
        return pocos
    log.info(f"Usando lista padrão ({len(POCOS_DEFAULT)} poços).")
    return POCOS_DEFAULT


def _edge_profile_dir() -> Path:
    """Diretório de perfil dedicado para o scraper."""
    scraper_profile = Path(os.environ.get("LOCALAPPDATA", "")) / "SeqopScraper_Edge"
    scraper_profile.mkdir(parents=True, exist_ok=True)
    return scraper_profile


def criar_driver() -> webdriver.Edge:
    """Cria instância do Edge.

    1. Tenta conectar a um Edge já rodando com --remote-debugging-port.
    2. Se falhar, abre um Edge novo com perfil dedicado.
    """
    # Tentativa 1: conectar via remote debugging
    try:
        log.info(f"Conectando ao Edge na porta {DEBUGGING_PORT}...")
        opts = EdgeOptions()
        opts.debugger_address = f"127.0.0.1:{DEBUGGING_PORT}"
        driver = webdriver.Edge(options=opts)
        log.info("Conectado ao Edge existente via remote debugging.")
        return driver
    except WebDriverException:
        log.info("Edge com debugging não encontrado. Abrindo nova instância...")

    # Tentativa 2: Edge novo com perfil dedicado
    opts = EdgeOptions()
    profile = _edge_profile_dir()
    log.info(f"Perfil dedicado: {profile}")
    opts.add_argument(f"--user-data-dir={profile}")
    opts.add_argument("--profile-directory=Default")
    opts.add_argument("--start-maximized")
    opts.add_argument(f"--remote-debugging-port={DEBUGGING_PORT}")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    driver = webdriver.Edge(options=opts)
    log.info("Edge aberto (perfil dedicado).")
    return driver


# ── Login e navegação ───────────────────────────────────────────────────────

def _esta_logado(driver: webdriver.Edge) -> bool:
    """Detecta se o usuário está logado (aside.menu visível, sem botão 'Entrar')."""
    try:
        menus = driver.find_elements(By.CSS_SELECTOR, "aside.menu")
        if menus and menus[0].is_displayed():
            btns_entrar = driver.find_elements(
                By.XPATH,
                "//button[contains(text(),'Entrar')] | //a[contains(text(),'Entrar')]",
            )
            for b in btns_entrar:
                if b.is_displayed():
                    return False
            return True
        # Fallback: tabela visível
        for t in driver.find_elements(By.TAG_NAME, "table"):
            try:
                if t.is_displayed():
                    return True
            except StaleElementReferenceException:
                continue
    except WebDriverException:
        pass
    return False


def aguardar_login(driver: webdriver.Edge):
    """Abre o SEQOP e aguarda login manual se necessário."""
    driver.get(BASE_URL)
    log.info("Página SEQOP aberta.")
    time.sleep(4)

    if _esta_logado(driver):
        log.info("Já logado!")
        return

    log.info("=" * 50)
    log.info("  AGUARDANDO LOGIN MANUAL")
    log.info("  Faça login no navegador que abriu.")
    log.info("=" * 50)

    for t in range(120):  # até 10 min
        time.sleep(5)
        if _esta_logado(driver):
            log.info("Login detectado!")
            time.sleep(2)
            return
        if t > 0 and t % 12 == 0:
            log.info(f"  ... aguardando ({t * 5 // 60} min)")

    raise TimeoutError("Timeout aguardando login (10 min).")


def navegar_arquivadas(driver: webdriver.Edge):
    """Garante que estamos na página de Sequências Arquivadas."""
    if "arquivadas" in driver.current_url.lower():
        return

    log.info("Navegando para Arquivadas...")

    # Estratégia 1: clicar no link exato da sidebar
    try:
        link = driver.find_element(
            By.CSS_SELECTOR, 'a[href="/seqop/#/arquivadas2"]'
        )
        driver.execute_script("arguments[0].click();", link)
        time.sleep(3)
        if "arquivadas" in driver.current_url.lower():
            log.info(f"  Sidebar OK → {driver.current_url}")
            return
    except (NoSuchElementException, WebDriverException):
        pass

    # Estratégia 2: alterar hash via JS
    try:
        driver.execute_script("window.location.hash = '#/arquivadas2';")
        time.sleep(3)
        if "arquivadas" in driver.current_url.lower():
            log.info(f"  Hash OK → {driver.current_url}")
            return
    except WebDriverException:
        pass

    # Estratégia 3: URL direta
    driver.get(ARQUIVADAS_URL)
    time.sleep(5)
    if not _esta_logado(driver):
        log.warning("  Sessão perdida após reload. Re-login necessário.")
        aguardar_login(driver)
        navegar_arquivadas(driver)
    log.info(f"  URL direta → {driver.current_url}")


# ── Filtros e pesquisa ──────────────────────────────────────────────────────

def _limpar_filtros(driver: webdriver.Edge):
    """Clica no botão 'Limpar' (btn-warning) para resetar filtros."""
    try:
        btn = driver.find_element(By.CSS_SELECTOR, "button.btn.btn-warning")
        driver.execute_script("arguments[0].click();", btn)
        time.sleep(1)
        log.info("  Filtros limpos.")
    except (NoSuchElementException, WebDriverException) as e:
        log.warning(f"  Botão Limpar não encontrado: {e}")


def pesquisar_poco(driver: webdriver.Edge, wait: WebDriverWait, poco: str):
    """Preenche input#poco, aguarda Pesquisar habilitar, clica."""
    log.info(f"Pesquisando poço: {poco}")
    _limpar_filtros(driver)
    time.sleep(0.5)

    # Preencher campo Poço (input#poco, placeholder YYY-XX)
    campo = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#poco"))
    )
    campo.click()
    campo.send_keys(Keys.CONTROL, "a")
    campo.send_keys(Keys.DELETE)
    time.sleep(0.3)
    campo.send_keys(poco)
    time.sleep(1)  # tempo para Vue validar e habilitar o Pesquisar

    # Aguardar o botão Pesquisar ficar habilitado (perde o atributo disabled)
    btn_pesquisar = None
    try:
        btn_pesquisar = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button.btn.btn-success"))
        )
        log.info("  Botão Pesquisar habilitado.")
    except TimeoutException:
        log.warning("  Pesquisar ainda desabilitado. Forçando via JS...")
        btn_pesquisar = driver.find_element(
            By.CSS_SELECTOR, "button.btn.btn-success"
        )
        driver.execute_script(
            "arguments[0].removeAttribute('disabled');"
            "arguments[0].classList.remove('disabled');",
            btn_pesquisar,
        )
        time.sleep(0.3)

    driver.execute_script("arguments[0].click();", btn_pesquisar)
    log.info("  Pesquisar clicado.")

    # Aguardar tabela carregar
    time.sleep(2)
    try:
        wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr td"))
        )
        log.info("  Tabela carregada.")
    except TimeoutException:
        log.warning("  Tabela vazia ou demorou para carregar.")
        time.sleep(3)


# ── Extração de dados da tabela ─────────────────────────────────────────────

def eh_sequencia_mpd(titulo: str, badges_texto: str) -> bool:
    """True se a sequência é MPD (por título ou badge de aprovador)."""
    t = titulo.upper()
    b = badges_texto.upper()
    return ("MPD" in t or "MPD" in b
            or "CSD-MPD" in t or "CSD\u2011MPD" in t)


def extrair_aprovacoes_pagina(driver: webdriver.Edge, poco_pesquisado: str) -> list[dict]:
    """Extrai aprovações MPD das linhas visíveis da tabela.

    Cada <span class="aprovador"> na coluna Aprovadores tem um atributo
    title no formato:
      "Aprovado em DD/MM/YYYY HH:MM por Nome Completo"

    Retorna uma lista de dicts, um por aprovação.
    """
    dados = []
    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")

    for row in rows:
        try:
            tds = row.find_elements(By.TAG_NAME, "td")
            if len(tds) < 8:
                continue

            # ── Extrair texto das colunas ────────────────────────────
            poco_tabela = tds[COL_POCO].text.strip()

            # Título: atributo title da span interna (texto completo)
            titulo_spans = tds[COL_TITULO].find_elements(By.TAG_NAME, "span")
            if titulo_spans:
                titulo = (
                    titulo_spans[0].get_attribute("title")
                    or titulo_spans[0].text.strip()
                )
            else:
                titulo = tds[COL_TITULO].text.strip()

            # Versão: texto da primeira span (ignora ícone Word)
            versao_spans = tds[COL_VERSAO].find_elements(By.TAG_NAME, "span")
            versao = (
                versao_spans[0].text.strip()
                if versao_spans
                else tds[COL_VERSAO].text.strip()
            )

            publicado = tds[COL_PUBLICADO].text.strip()

            # ── Aprovadores ──────────────────────────────────────────
            aprov_spans = tds[COL_APROVADORES].find_elements(
                By.CSS_SELECTOR, "span.aprovador"
            )
            badges = []
            for sp in aprov_spans:
                badges.append({
                    "badge": sp.text.strip(),
                    "title": sp.get_attribute("title") or "",
                })

            # ── Filtrar: só interessa o badge "MPD" ────────────────
            # Procurar o span cujo texto do badge é "MPD" (case-insensitive,
            # strip whitespace). Se não houver badge MPD, pular a linha.
            mpd_badge = None
            for b in badges:
                if b["badge"].strip().upper() == "MPD":
                    mpd_badge = b
                    break

            if mpd_badge is None:
                # Sem badge MPD → não é sequência MPD, pular
                continue

            log.info(f"  → MPD: {titulo[:70]}")

            # seq_id: tentar extrair do título (ex: "S02 - …")
            m_seq = re.match(r"(S\d+)", titulo)
            seq_id = m_seq.group(1) if m_seq else ""

            # ── Extrair apenas o aprovador MPD ───────────────────────
            match = RE_APROVACAO.search(mpd_badge["title"])
            if match:
                dados.append({
                    "poco": poco_pesquisado,
                    "seq_id": seq_id,
                    "titulo": titulo,
                    "publicado_em": publicado,
                    "versao": versao,
                    "aprovado_por": match.group(3).strip(),
                    "aprovado_em": f"{match.group(1)} {match.group(2)}",
                    "papel_aprovador": mpd_badge["badge"].strip(),
                })
            else:
                # Badge MPD existe mas sem "Aprovado em …" (pendente?)
                dados.append({
                    "poco": poco_pesquisado,
                    "seq_id": seq_id,
                    "titulo": titulo,
                    "publicado_em": publicado,
                    "versao": versao,
                    "aprovado_por": mpd_badge["title"] or "(pendente)",
                    "aprovado_em": "",
                    "papel_aprovador": mpd_badge["badge"].strip(),
                })

        except StaleElementReferenceException:
            log.warning("  Stale element – pulando linha.")
            continue
        except Exception as e:
            log.warning(f"  Erro na linha: {e}")
            continue

    return dados


# ── Paginação ───────────────────────────────────────────────────────────────

def avancar_pagina(driver: webdriver.Edge) -> bool:
    """Clica no botão '›' (próxima página) se disponível.

    Na b-pagination do Bootstrap-Vue:
      - Habilitado: <li class="page-item"><button aria-label="Go to next page">
      - Desabilitado: <li class="page-item disabled"><span aria-label="Go to next page">

    Retorna True se avançou de página.
    """
    try:
        # Botão <button> existe somente quando habilitado
        btns = driver.find_elements(
            By.CSS_SELECTOR, 'button[aria-label="Go to next page"]'
        )
        if btns:
            btn = btns[0]
            # Checar se o <li> pai está desabilitado
            parent_li = btn.find_element(By.XPATH, "./ancestor::li[1]")
            if "disabled" in (parent_li.get_attribute("class") or ""):
                log.info("  Última página (botão › desabilitado).")
                return False

            driver.execute_script("arguments[0].click();", btn)
            time.sleep(PAUSA + 1)
            log.info("  Avançou para próxima página.")
            return True
        else:
            # Se não achou <button>, verificar se existe <span> (desabilitado)
            spans = driver.find_elements(
                By.CSS_SELECTOR, 'span[aria-label="Go to next page"]'
            )
            if spans:
                log.info("  Última página (botão › é <span> desabilitado).")
            else:
                log.info("  Sem paginação (página única ou vazia).")
            return False
    except (NoSuchElementException, StaleElementReferenceException, WebDriverException) as e:
        log.warning(f"  Erro na paginação: {e}")
        return False


# ── Excel ───────────────────────────────────────────────────────────────────

def salvar_excel(resultados: list[AprovacaoMPD]):
    """Salva os resultados em Excel formatado."""
    if not resultados:
        log.warning("Nenhum resultado para salvar.")
        return

    df = pd.DataFrame([asdict(r) for r in resultados])
    df.sort_values(
        by=["poco", "titulo", "versao", "papel_aprovador"],
        inplace=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "aprovacoes_mpd"

    colunas = [
        ("Poço", 22),
        ("Seq ID", 10),
        ("Título", 55),
        ("Publicado Em", 16),
        ("Versão", 10),
        ("Aprovado Por", 40),
        ("Aprovado Em", 18),
        ("Papel", 10),
        ("URL Fonte", 55),
        ("Extração (UTC)", 22),
    ]

    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ci, (nome, largura) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=ci, value=nome)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = largura

    for ri, row_data in enumerate(
        dataframe_to_rows(df, index=False, header=False), 2
    ):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(SAIDA_XLSX)
    log.info(f"Excel salvo: {SAIDA_XLSX}  ({len(resultados)} linhas)")


# ── Scraper principal ───────────────────────────────────────────────────────

def executar_scraper():
    log.info("=" * 60)
    log.info("SCRAPER SEQOP v2 – Extração direta da tabela")
    log.info(f"Data/hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    log.info("=" * 60)

    pocos = carregar_pocos()
    log.info(f"Poços a processar: {len(pocos)}")

    todos: list[AprovacaoMPD] = []
    driver = criar_driver()
    wait = WebDriverWait(
        driver, WAIT_TIMEOUT,
        ignored_exceptions=[StaleElementReferenceException],
    )

    try:
        aguardar_login(driver)
        navegar_arquivadas(driver)

        for idx, poco in enumerate(pocos, 1):
            log.info("-" * 50)
            log.info(f"[{idx}/{len(pocos)}] Poço: {poco}")

            try:
                navegar_arquivadas(driver)
                pesquisar_poco(driver, wait, poco)

                pagina = 1
                seq_mpd_total = 0

                while True:
                    log.info(f"  Página {pagina}")
                    dados = extrair_aprovacoes_pagina(driver, poco)
                    log.info(f"  → {len(dados)} aprovações MPD nesta página")
                    seq_mpd_total += len(dados)

                    url_atual = driver.current_url
                    for d in dados:
                        todos.append(AprovacaoMPD(
                            poco=d["poco"],
                            seq_id=d["seq_id"],
                            titulo=d["titulo"],
                            publicado_em=d["publicado_em"],
                            versao=d["versao"],
                            aprovado_por=d["aprovado_por"],
                            aprovado_em=d["aprovado_em"],
                            papel_aprovador=d["papel_aprovador"],
                            fonte_url=url_atual,
                        ))

                    if not avancar_pagina(driver):
                        break
                    pagina += 1

                log.info(
                    f"  Poço {poco}: {seq_mpd_total} aprovações MPD "
                    f"em {pagina} página(s)"
                )

            except Exception as e:
                log.error(f"Erro no poço {poco}: {e}", exc_info=True)
                try:
                    driver.get(ARQUIVADAS_URL)
                    time.sleep(3)
                except Exception:
                    log.error("Recriando driver...")
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    driver = criar_driver()
                    wait = WebDriverWait(
                        driver, WAIT_TIMEOUT,
                        ignored_exceptions=[StaleElementReferenceException],
                    )
                    aguardar_login(driver)

            # Salvar progresso parcial a cada poço
            if todos:
                salvar_excel(todos)
                log.info(f"  Progresso salvo: {len(todos)} aprovações até agora")

    except KeyboardInterrupt:
        log.info("Interrompido pelo usuário (Ctrl+C)")
    except Exception as e:
        log.error(f"Erro fatal: {e}", exc_info=True)
    finally:
        if todos:
            salvar_excel(todos)
        log.info("=" * 60)
        log.info(f"Total: {len(todos)} aprovações extraídas → {SAIDA_XLSX}")
        log.info("=" * 60)
        try:
            input("\nPressione Enter para fechar o navegador...")
        except EOFError:
            pass
        driver.quit()


# ── Ponto de entrada ────────────────────────────────────────────────────────

if __name__ == "__main__":
    executar_scraper()

"""
Scraper SEQOP – Extração de Sequências Operacionais MPD
========================================================
Acessa https://csdpocos.petrobras.com.br/seqop/#/arquivadas2
e coleta o histórico de aprovações de sequências MPD para uma lista de poços.

Uso:
    C:\\SharedPython\\venv\\Scripts\\python.exe scraper_seqop.py

Pré-requisitos:
    - Microsoft Edge instalado
    - selenium, openpyxl, pandas no venv compartilhado
    - Arquivo 'pocos_entrada.xlsx' com coluna "poco" (ou editar POCOS_DEFAULT)
"""

# ── Re-launch com o venv compartilhado ──────────────────────────────────────
import subprocess
import sys
from pathlib import Path

_SHARED_PYTHON = Path(r"C:\SharedPython\venv\Scripts\python.exe")

if _SHARED_PYTHON.exists() and Path(sys.executable).resolve() != _SHARED_PYTHON.resolve():
    print(f"[LAUNCH] Relançando com o Python do venv compartilhado: {_SHARED_PYTHON}")
    result = subprocess.run([str(_SHARED_PYTHON)] + sys.argv, check=False)
    sys.exit(result.returncode)

# ── Imports ─────────────────────────────────────────────────────────────────
import logging
import os
import re
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    WebDriverException,
)

# ── Configuração ────────────────────────────────────────────────────────────

BASE_URL = "https://csdpocos.petrobras.com.br/seqop"
ARQUIVADAS_URL = f"{BASE_URL}/#/arquivadas2"

# Diretório deste script
SCRIPT_DIR = Path(__file__).resolve().parent

# Arquivo Excel de entrada (coluna "poco")
ENTRADA_XLSX = SCRIPT_DIR / "pocos_entrada.xlsx"

# Arquivo Excel de saída
SAIDA_XLSX = SCRIPT_DIR / "resultado_aprovacoes_mpd.xlsx"

# Lista padrão de poços (usada se o Excel de entrada não existir)
POCOS_DEFAULT = [
    "1-APS-57",
    "1-RJS-763DA",
    "3-SPS-114",
    "3-SPS-114A",
    "4-RJS-764",
    "7-BR-86DB-RJS",
    "7-BUZ-100DA-RJS",
    "7-BUZ-94D-RJS",
    "7-BUZ-95-RJS",
    "8-BUZ-89D-RJS",
    "7-OATP-3B-RJS",
    "8-BUZ-96D-RJS",
    "8-MRO-36-RJS",
    "3-SPS-111D",
    "4-SPS-112",
    "3-RJS-762",
    "8-ATP-8D-RJS",
    "7-STUP-10DA-RJS",
    "7-BUZ-90D-RJS",
]

# Tempo máximo de espera (segundos) para elementos na página
WAIT_TIMEOUT = 20

# Pausa entre ações para não sobrecarregar o servidor (segundos)
BACKOFF_SECS = 1.5

# Regex para extrair histórico de aprovações
# Exemplos:  "v1: Aprovado por CHAVE em 01/01/2025"
#            "v 2 : Aprovado por ABC123 em 15/03/2025"
RE_HISTORICO = re.compile(
    r"v\s*(\d+)\s*:\s*Aprovado\s+por\s+([A-Z0-9]+)\s+em\s+(\d{2}/\d{2}/\d{4})",
    re.IGNORECASE,
)

# ── Logging ─────────────────────────────────────────────────────────────────

LOG_FILE = SCRIPT_DIR / "scraper_seqop.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("seqop_scraper")


# ── Dataclass para resultados ───────────────────────────────────────────────

@dataclass
class AprovacaoMPD:
    poco: str
    seq_id: str
    titulo: str
    publicado_em: str
    versao: str
    aprovado_por: str
    aprovado_em: str
    fonte_url: str
    carimbo_extracao_utc: str = field(
        default_factory=lambda: datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    )


# ── Funções auxiliares ──────────────────────────────────────────────────────

def carregar_pocos() -> list[str]:
    """Carrega lista de poços do Excel ou usa a lista padrão."""
    if ENTRADA_XLSX.exists():
        log.info(f"Carregando poços de {ENTRADA_XLSX}")
        df = pd.read_excel(ENTRADA_XLSX, engine="openpyxl")
        # Aceita coluna "poco", "poço", "Poco", "Poço", "POCO"
        col = None
        for c in df.columns:
            if c.strip().lower().replace("ç", "c") in ("poco", "poço"):
                col = c
                break
        if col is None:
            log.warning("Coluna 'poco' não encontrada no Excel. Usando lista padrão.")
            return POCOS_DEFAULT
        pocos = df[col].dropna().astype(str).str.strip().tolist()
        log.info(f"  → {len(pocos)} poços carregados do Excel")
        return pocos
    else:
        log.info(f"Arquivo {ENTRADA_XLSX} não encontrado. Usando lista padrão ({len(POCOS_DEFAULT)} poços).")
        return POCOS_DEFAULT


def _edge_profile_dir() -> Path:
    """Retorna um diretório de perfil dedicado para o scraper.

    Em vez de usar o User Data original (que fica travado pelo Edge aberto),
    cria uma cópia rasa em %LOCALAPPDATA%/SeqopScraper_Edge que o Selenium
    pode usar sem conflito com o Edge do dia-a-dia.
    Na primeira vez, o usuário terá que fazer login uma vez; depois o perfil
    é reutilizado automaticamente.
    """
    scraper_profile = Path(os.environ.get("LOCALAPPDATA", "")) / "SeqopScraper_Edge"
    scraper_profile.mkdir(parents=True, exist_ok=True)
    return scraper_profile


DEBUGGING_PORT = 9222  # porta para remote debugging


def criar_driver() -> webdriver.Edge:
    """Cria instância do Edge.

    Estratégia (em ordem):
      1. Tenta conectar a um Edge já rodando com --remote-debugging-port.
      2. Se falhar, abre um Edge novo usando um perfil dedicado para o scraper
         (evita conflito com o Edge pessoal/corporativo que já está aberto).
    """
    # ── Tentativa 1: conectar via remote debugging ──────────────────────
    try:
        log.info(f"Tentando conectar a um Edge já aberto na porta {DEBUGGING_PORT}...")
        options = EdgeOptions()
        options.debugger_address = f"127.0.0.1:{DEBUGGING_PORT}"
        driver = webdriver.Edge(options=options)
        log.info("Conectado ao Edge existente via remote debugging!")
        return driver
    except WebDriverException:
        log.info("Nenhum Edge com debugging encontrado. Abrindo nova instância...")

    # ── Tentativa 2: Edge novo com perfil dedicado ──────────────────────
    options = EdgeOptions()

    scraper_profile = _edge_profile_dir()
    log.info(f"Perfil dedicado do scraper: {scraper_profile}")
    options.add_argument(f"--user-data-dir={scraper_profile}")
    options.add_argument("--profile-directory=Default")

    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument(f"--remote-debugging-port={DEBUGGING_PORT}")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])

    # Selenium 4 gerencia o driver automaticamente (Selenium Manager)
    driver = webdriver.Edge(options=options)
    log.info("Edge aberto com sucesso (perfil dedicado).")
    return driver


# ── Índices das colunas da tabela (0-based) ────────────────────────────────
# Sonda(0) | Polo(1) | Poço(2) | Interv(3) | Título(4) | Versão(5) |
# Publicado(6) | Aprovadores(7) | Ações(8)
COL_SONDA = 0
COL_POLO = 1
COL_POCO = 2
COL_INTERV = 3
COL_TITULO = 4
COL_VERSAO = 5
COL_PUBLICADO = 6
COL_APROVADORES = 7
COL_ACOES = 8


def _esta_logado(driver: webdriver.Edge) -> bool:
    """Detecta se o usuário já está logado no SEQOP.

    Verifica se a sidebar com 'Sequências Ativas' / 'Sequências Arquivadas'
    está visível – isso só acontece após login bem-sucedido.
    """
    try:
        # Procurar texto que só existe pós-login
        indicadores = driver.find_elements(
            By.XPATH,
            "//*[contains(text(),'Ativas') or contains(text(),'Arquivadas') "
            "or contains(text(),'SEQOP') and contains(text(),'Sistema')]"
        )
        for el in indicadores:
            try:
                if el.is_displayed():
                    # Confirmar que NÃO é a tela de login
                    # (a tela de login também mostra 'SEQOP' mas mostra botão 'Entrar')
                    btns_entrar = driver.find_elements(
                        By.XPATH,
                        "//button[contains(text(),'Entrar')] | //a[contains(text(),'Entrar')]"
                    )
                    for btn in btns_entrar:
                        if btn.is_displayed():
                            return False  # tela de login
                    return True
            except StaleElementReferenceException:
                continue

        # Procurar tabela de sequências (indicador forte de estar logado)
        tabelas = driver.find_elements(By.CSS_SELECTOR, "table")
        for t in tabelas:
            try:
                if t.is_displayed() and len(t.text) > 50:
                    return True
            except StaleElementReferenceException:
                continue

    except WebDriverException:
        pass
    return False


def aguardar_login(driver: webdriver.Edge):
    """Aguarda o usuário fazer login (se necessário)."""
    driver.get(BASE_URL)
    log.info("Página inicial do SEQOP aberta.")
    time.sleep(4)

    # Verificar se já está logado
    if _esta_logado(driver):
        log.info("Já está logado!")
        return

    # Mostrar mensagem e aguardar login manual
    log.info("=" * 60)
    log.info("  AGUARDANDO LOGIN MANUAL")
    log.info("  Faça login no navegador que abriu.")
    log.info("  O script vai detectar automaticamente quando concluir.")
    log.info("=" * 60)

    tentativas = 0
    while tentativas < 120:  # Até 10 minutos
        time.sleep(5)
        tentativas += 1

        if _esta_logado(driver):
            log.info("Login detectado com sucesso!")
            time.sleep(2)
            return

        if tentativas % 12 == 0:
            minutos = tentativas * 5 // 60
            log.info(f"  ... ainda aguardando login ({minutos} min)")

    raise TimeoutError("Timeout aguardando login (10 min). Execute novamente.")


def _verificar_pagina_arquivadas(driver: webdriver.Edge) -> bool:
    """Retorna True se a página atual é a de sequências arquivadas.

    Verifica pela URL (deve conter 'arquivadas') OU pelo título da página
    'Sequências Arquivadas 2.1' (não confundir com o texto da sidebar).
    """
    url = driver.current_url.lower()
    # Critério principal: URL contém 'arquivadas'
    if "arquivadas" in url:
        return True
    # Critério secundário: heading específico (não a sidebar)
    # O heading da página é algo como <h1>Sequências Arquivadas 2.1</h1>
    try:
        headings = driver.find_elements(By.XPATH,
            "//h1[contains(text(),'Arquivadas')] | //h2[contains(text(),'Arquivadas')] "
            "| //h3[contains(text(),'Arquivadas')]"
        )
        for h in headings:
            if h.is_displayed():
                return True
    except WebDriverException:
        pass
    return False


def navegar_arquivadas(driver: webdriver.Edge, wait: WebDriverWait):
    """Navega até a página de sequências arquivadas clicando no menu lateral."""
    if _verificar_pagina_arquivadas(driver):
        log.info(f"Já na página de arquivadas: {driver.current_url}")
        return

    log.info(f"Navegando para Sequências Arquivadas... (URL atual: {driver.current_url})")

    # ── Estratégia 1: clicar no item da sidebar ─────────────────────────
    # A sidebar tem itens com quebra de linha: 'Sequências\nArquivadas'
    # Preferimos clicar no <a> pai ou no próprio item do menu
    try:
        # Buscar todos os elementos que contenham 'Arquivadas'
        candidatos = driver.find_elements(By.XPATH,
            "//*[contains(text(),'Arquivadas')]"
        )
        for el in candidatos:
            try:
                if not el.is_displayed():
                    continue
                tag = el.tag_name.lower()
                txt = el.text.strip().replace('\n', ' ')[:60]

                # Pular se for heading da página (h1/h2/h3) — não é menu
                if tag in ("h1", "h2", "h3"):
                    continue

                log.info(f"  Sidebar candidato: <{tag}> '{txt}'")

                # Tentar encontrar o <a> ancestral mais próximo (link da sidebar)
                elemento_clicavel = el
                try:
                    link_pai = el.find_element(By.XPATH, "./ancestor::a[1]")
                    if link_pai.is_displayed():
                        elemento_clicavel = link_pai
                        log.info(f"  Usando <a> pai: href='{link_pai.get_attribute('href')}'")
                except NoSuchElementException:
                    pass

                # Clicar via JS para evitar qualquer interception
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});",
                    elemento_clicavel
                )
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", elemento_clicavel)
                time.sleep(4)

                if _verificar_pagina_arquivadas(driver):
                    log.info(f"  Navegação sidebar OK → {driver.current_url}")
                    return
                else:
                    log.info(f"  Clique não navegou (URL: {driver.current_url}). Tentando próximo...")
            except (StaleElementReferenceException, WebDriverException) as e:
                log.warning(f"  Erro no candidato: {e}")
                continue
    except WebDriverException as e:
        log.warning(f"  Busca sidebar falhou: {e}")

    # ── Estratégia 2: mudar o hash diretamente (sem reload) ─────────────
    log.info("  Sidebar não funcionou. Tentando alterar hash via JS...")
    try:
        driver.execute_script("window.location.hash = '#/arquivadas2';")
        time.sleep(4)
        if _verificar_pagina_arquivadas(driver):
            log.info(f"  Hash change OK → {driver.current_url}")
            return
    except WebDriverException as e:
        log.warning(f"  Hash change falhou: {e}")

    # ── Estratégia 3: URL direta (último recurso) ───────────────────────
    log.info("  Tentando URL direta...")
    driver.get(ARQUIVADAS_URL)
    time.sleep(5)
    if not _esta_logado(driver):
        log.warning("  Sessão perdida após reload. Aguardando re-login...")
        aguardar_login(driver)
        navegar_arquivadas(driver, wait)
        return
    log.info(f"  URL direta → {driver.current_url}")


def _limpar_filtros(driver: webdriver.Edge):
    """Clica no botão 'Limpar' para resetar todos os filtros."""
    try:
        # Procurar por texto nos botões visíveis
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            try:
                if btn.is_displayed() and "limpar" in btn.text.strip().lower():
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(1)
                    log.info("  Filtros limpos")
                    return
            except StaleElementReferenceException:
                continue
    except WebDriverException:
        pass


def pesquisar_poco(driver: webdriver.Edge, wait: WebDriverWait, poco: str):
    """Preenche o campo Poço e clica em Pesquisar.

    A página 'Sequências Arquivadas 2.1' tem filtros específicos:
      - Sonda: placeholder 'NS-XX ou XX'
      - Poço:  placeholder 'YYY-XX'
      - Polo:  dropdown
      - Título: placeholder 'Palavras que devem estar contidas no título'
      - Botão 'Pesquisar' (azul)  e  'Limpar'
    """
    log.info(f"Pesquisando poço: {poco}")

    # Limpar filtros anteriores primeiro
    _limpar_filtros(driver)

    # Encontrar o campo Poço (placeholder 'YYY-XX')
    campo_poco = None
    seletores = [
        "input[placeholder*='YYY']",
        "input[placeholder*='yyy']",
        # Label 'Poço:' seguido de input
        "//label[contains(text(),'oço')]/following::input[1]",
        "//label[contains(text(),'oco')]/following::input[1]",
        "//span[contains(text(),'oço')]/following::input[1]",
        # Pelo texto antes do input
        "//td[contains(text(),'oço')]/following::input[1]",
        "//div[contains(text(),'oço')]/following::input[1]",
    ]

    for sel in seletores:
        try:
            if sel.startswith("//"):
                campo_poco = driver.find_element(By.XPATH, sel)
            else:
                campo_poco = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, sel))
                )
            if campo_poco and campo_poco.is_displayed():
                log.info(f"  Campo Poço encontrado: {sel}")
                break
            campo_poco = None
        except (TimeoutException, NoSuchElementException):
            campo_poco = None
            continue

    if campo_poco is None:
        # Fallback: o segundo input de texto (primeiro é Sonda)
        inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='text'], input:not([type])")
        inputs_visiveis = [i for i in inputs if i.is_displayed()]
        if len(inputs_visiveis) >= 2:
            campo_poco = inputs_visiveis[1]  # segundo campo = Poço
            log.info("  Campo Poço encontrado por posição (2º input)")
        elif inputs_visiveis:
            campo_poco = inputs_visiveis[0]
            log.info("  Campo Poço encontrado (1º input disponível)")
        else:
            raise NoSuchElementException("Campo Poço não encontrado.")

    # Limpar e preencher
    campo_poco.click()
    campo_poco.send_keys(Keys.CONTROL, "a")
    campo_poco.send_keys(Keys.DELETE)
    time.sleep(0.3)
    campo_poco.send_keys(poco)
    time.sleep(0.5)

    # Clicar no botão 'Pesquisar' (azul)
    btn_pesquisar = None
    seletores_btn = [
        "//button[contains(text(),'Pesquisar')]",
        "//button[contains(text(),'pesquisar')]",
        "//button[@type='submit']",
        "//input[@type='submit']",
    ]
    for sel in seletores_btn:
        try:
            btn_pesquisar = driver.find_element(By.XPATH, sel)
            if btn_pesquisar.is_displayed():
                log.info(f"  Botão Pesquisar encontrado: {sel}")
                break
            btn_pesquisar = None
        except NoSuchElementException:
            continue

    if btn_pesquisar is None:
        # Fallback: buscar todos os botões visueis
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            try:
                if btn.is_displayed() and "pesquisar" in btn.text.strip().lower():
                    btn_pesquisar = btn
                    break
            except StaleElementReferenceException:
                continue

    if btn_pesquisar:
        btn_pesquisar.click()
        log.info("  Botão Pesquisar clicado")
    else:
        log.warning("  Botão Pesquisar não encontrado – pressionando Enter")
        campo_poco.send_keys(Keys.RETURN)

    # Aguardar resultados (server-side, pode demorar)
    time.sleep(3)
    _aguardar_tabela(driver, wait)
    log.info(f"  Pesquisa concluída para '{poco}'")


def _aguardar_tabela(driver: webdriver.Edge, wait: WebDriverWait):
    """Aguarda a tabela de resultados renderizar."""
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr td")))
        log.info("  Tabela com dados carregada")
        return
    except TimeoutException:
        pass
    # Fallback: pelo menos a tag table
    try:
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
        return
    except TimeoutException:
        log.warning("  Tabela não encontrada após espera")
        time.sleep(3)


def eh_sequencia_mpd(titulo: str, aprovadores_texto: str) -> bool:
    """Verifica se uma sequência é MPD pelas regras definidas."""
    titulo_upper = titulo.upper()
    aprov_upper = aprovadores_texto.upper()

    # Regra 1: Badge/texto "MPD" na coluna Aprovadores
    if "MPD" in aprov_upper:
        return True

    # Regra 2: Título contém "MPD"
    if "MPD" in titulo_upper:
        return True

    # Regra 3: "CSD-MPD" ou "CSD‑MPD" (meia-risca unicode)
    if "CSD-MPD" in titulo_upper or "CSD\u2011MPD" in titulo_upper:
        return True
    if "CSD-MPD" in aprov_upper or "CSD\u2011MPD" in aprov_upper:
        return True

    return False


def extrair_linhas_tabela(driver: webdriver.Edge) -> list[dict]:
    """Lê todas as linhas da tabela de resultados na página atual.

    Colunas reais do SEQOP:
    Sonda(0) | Polo(1) | Poço(2) | Interv(3) | Título(4) |
    Versão(5) | Publicado(6) | Aprovadores(7) | Ações(8)
    """
    resultados = []

    # Encontrar linhas da tabela
    linhas = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
    if not linhas:
        linhas = driver.find_elements(By.CSS_SELECTOR, "table tr")

    for linha in linhas:
        try:
            colunas = linha.find_elements(By.TAG_NAME, "td")
            if len(colunas) < 7:  # precisa ter pelo menos até Aprovadores
                continue

            textos = [col.text.strip() for col in colunas]

            # Extrair dados pelas posições corretas
            sonda    = textos[COL_SONDA]    if len(textos) > COL_SONDA else ""
            polo     = textos[COL_POLO]     if len(textos) > COL_POLO else ""
            poco     = textos[COL_POCO]     if len(textos) > COL_POCO else ""
            interv   = textos[COL_INTERV]   if len(textos) > COL_INTERV else ""
            titulo   = textos[COL_TITULO]   if len(textos) > COL_TITULO else ""
            versao   = textos[COL_VERSAO]   if len(textos) > COL_VERSAO else ""
            publicado = textos[COL_PUBLICADO] if len(textos) > COL_PUBLICADO else ""
            aprovadores = textos[COL_APROVADORES] if len(textos) > COL_APROVADORES else ""

            # Botão de ação (coluna Ações – última coluna)
            btn_acao = None
            if len(colunas) > COL_ACOES:
                col_acoes = colunas[COL_ACOES]
                # Procurar botões, links ou ícones clicáveis
                clicaveis = (
                    col_acoes.find_elements(By.TAG_NAME, "button")
                    or col_acoes.find_elements(By.TAG_NAME, "a")
                    or col_acoes.find_elements(By.CSS_SELECTOR, "i, mat-icon, .fa, span[class*='icon']")
                )
                btn_acao = clicaveis[0] if clicaveis else None

            resultados.append({
                "sonda": sonda,
                "polo": polo,
                "poco": poco,
                "interv": interv,
                "titulo": titulo,
                "versao": versao,
                "publicado": publicado,
                "aprovadores": aprovadores,
                "btn_acao": btn_acao,
                "elemento_linha": linha,
            })
        except StaleElementReferenceException:
            continue

    return resultados


def tem_proxima_pagina(driver: webdriver.Edge) -> bool:
    """Verifica se existe botão de próxima página (›) e clica nele.

    Paginação do SEQOP: « ‹ [1] 2 3 4 › »
    O '›' avança uma página. Está dentro de <li> de uma paginação ngb-pagination.
    """
    seletores = [
        # Próxima página: ›
        "//a[normalize-space(text())='›']",
        "//a[normalize-space(text())='>']",
        "//li[contains(@class,'next')]/a",
        "//a[contains(@aria-label,'Next')]",
        "//a[contains(@aria-label,'next')]",
        "//a[contains(@aria-label,'Próxim')]",
        # ngb-pagination specific
        "//ngb-pagination//a[normalize-space(text())='›']",
    ]

    for sel in seletores:
        try:
            btn = driver.find_element(By.XPATH, sel)
            if not btn.is_displayed():
                continue
            # Verificar se o parent <li> está desabilitado
            try:
                parent_li = btn.find_element(By.XPATH, "./ancestor::li[1]")
                classes = (parent_li.get_attribute("class") or "").lower()
                if "disabled" in classes:
                    log.info("  Botão › desabilitado (última página)")
                    return False
            except NoSuchElementException:
                pass

            log.info("  Clicando próxima página (›)")
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(BACKOFF_SECS + 1)
            return True
        except (NoSuchElementException, StaleElementReferenceException):
            continue

    log.info("  Botão '›' não encontrado (página única ou última)")
    return False


def extrair_historico_detalhe(driver: webdriver.Edge, wait: WebDriverWait) -> list[dict]:
    """Extrai o histórico de aprovações da página de detalhe de uma sequência."""
    time.sleep(BACKOFF_SECS)

    # Pegar todo o texto da página de detalhe
    try:
        body = wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        page_text = body.text
    except TimeoutException:
        page_text = driver.page_source

    # Extrair ID da URL (#/view/<ID>)
    url_atual = driver.current_url
    match_id = re.search(r"#/view/(\d+)", url_atual)
    seq_id = match_id.group(1) if match_id else ""

    # Se não encontrou ID na URL, tentar outros padrões
    if not seq_id:
        match_id = re.search(r"/(\d+)\s*$", url_atual)
        seq_id = match_id.group(1) if match_id else "DESCONHECIDO"

    # Verificar se contém CSD-MPD no detalhe (regra 3)
    tem_csd_mpd = bool(re.search(r"CSD[\-\u2011]MPD", page_text, re.IGNORECASE))

    # Extrair todas as aprovações com regex
    aprovacoes = []
    for match in RE_HISTORICO.finditer(page_text):
        aprovacoes.append({
            "versao": match.group(1),
            "aprovado_por": match.group(2),
            "aprovado_em": match.group(3),
        })

    return aprovacoes, seq_id, tem_csd_mpd


def abrir_detalhe_sequencia(driver: webdriver.Edge, wait: WebDriverWait, btn_acao) -> bool:
    """Clica no botão de ação para abrir o detalhe da sequência."""
    try:
        # Scroll para o elemento se necessário
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn_acao)
        time.sleep(0.3)
        btn_acao.click()
        time.sleep(BACKOFF_SECS)
        return True
    except Exception as e:
        log.warning(f"  Erro ao clicar no botão de ação: {e}")
        # Tentar JavaScript click
        try:
            driver.execute_script("arguments[0].click();", btn_acao)
            time.sleep(BACKOFF_SECS)
            return True
        except Exception:
            return False


def voltar_para_lista(driver: webdriver.Edge, wait: WebDriverWait):
    """Volta para a página de arquivadas."""
    driver.back()
    time.sleep(BACKOFF_SECS)
    # Verificar se voltou corretamente
    if not _verificar_pagina_arquivadas(driver):
        log.info("  Back não funcionou, navegando via sidebar...")
        navegar_arquivadas(driver, wait)
        time.sleep(2)


def salvar_excel(resultados: list[AprovacaoMPD]):
    """Salva os resultados em um Excel formatado."""
    if not resultados:
        log.warning("Nenhum resultado para salvar.")
        return

    df = pd.DataFrame([asdict(r) for r in resultados])

    # Ordenar
    df.sort_values(by=["poco", "seq_id", "versao"], inplace=True)

    # Criar workbook formatado
    wb = Workbook()
    ws = wb.active
    ws.title = "aprovacoes_mpd"

    # Cabeçalhos
    colunas = [
        ("Poço", 20),
        ("Seq ID", 12),
        ("Título", 45),
        ("Publicado Em", 16),
        ("Versão", 10),
        ("Aprovado Por", 18),
        ("Aprovado Em", 16),
        ("URL Fonte", 50),
        ("Extração (UTC)", 22),
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, (nome, largura) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = largura

    # Dados
    for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Filtro automático
    ws.auto_filter.ref = ws.dimensions

    # Congelar primeira linha
    ws.freeze_panes = "A2"

    wb.save(SAIDA_XLSX)
    log.info(f"Excel salvo: {SAIDA_XLSX} ({len(resultados)} linhas)")


# ── Scraper principal ───────────────────────────────────────────────────────

def executar_scraper():
    """Função principal do scraper."""
    log.info("=" * 60)
    log.info("SCRAPER SEQOP – Início")
    log.info(f"Data/hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    log.info("=" * 60)

    # Carregar lista de poços
    pocos = carregar_pocos()
    log.info(f"Total de poços a processar: {len(pocos)}")

    # Resultados
    todos_resultados: list[AprovacaoMPD] = []

    # Criar driver
    log.info("Iniciando navegador Edge...")
    driver = criar_driver()
    wait = WebDriverWait(
        driver, WAIT_TIMEOUT,
        ignored_exceptions=[StaleElementReferenceException]
    )

    try:
        # Aguardar login
        aguardar_login(driver)

        # Navegar para arquivadas
        navegar_arquivadas(driver, wait)

        for idx_poco, poco in enumerate(pocos, 1):
            log.info("-" * 50)
            log.info(f"[{idx_poco}/{len(pocos)}] Processando poço: {poco}")

            try:
                # Garantir que estamos na página de arquivadas
                navegar_arquivadas(driver, wait)

                # Pesquisar
                pesquisar_poco(driver, wait, poco)

                # Processar todas as páginas
                pagina = 1
                while True:
                    log.info(f"  Página {pagina} de resultados")

                    linhas = extrair_linhas_tabela(driver)
                    log.info(f"  {len(linhas)} linhas encontradas na tabela")

                    if not linhas:
                        log.info(f"  Nenhum resultado para o poço {poco}")
                        break

                    # Filtrar sequências MPD
                    sequencias_mpd = []
                    for linha in linhas:
                        if eh_sequencia_mpd(linha["titulo"], linha["aprovadores"]):
                            sequencias_mpd.append(linha)
                            log.info(f"  → MPD encontrada: {linha['titulo'][:60]}")

                    log.info(f"  {len(sequencias_mpd)} sequências MPD nesta página")

                    # Para cada sequência MPD, abrir detalhe e extrair histórico
                    for seq_idx, seq in enumerate(sequencias_mpd):
                        log.info(f"    Abrindo detalhe ({seq_idx + 1}/{len(sequencias_mpd)}): {seq['titulo'][:50]}")

                        if seq["btn_acao"] is None:
                            log.warning(f"    Sem botão de ação para: {seq['titulo']}")
                            continue

                        # Abrir detalhe
                        if not abrir_detalhe_sequencia(driver, wait, seq["btn_acao"]):
                            log.warning(f"    Falha ao abrir detalhe: {seq['titulo']}")
                            continue

                        # Extrair histórico
                        aprovacoes, seq_id, tem_csd_mpd = extrair_historico_detalhe(driver, wait)

                        url_detalhe = driver.current_url

                        if aprovacoes:
                            for aprov in aprovacoes:
                                todos_resultados.append(AprovacaoMPD(
                                    poco=poco,
                                    seq_id=seq_id,
                                    titulo=seq["titulo"],
                                    publicado_em=seq["publicado"],
                                    versao=f"v{aprov['versao']}",
                                    aprovado_por=aprov["aprovado_por"],
                                    aprovado_em=aprov["aprovado_em"],
                                    fonte_url=url_detalhe,
                                ))
                            log.info(f"    → {len(aprovacoes)} aprovações extraídas")
                        else:
                            log.info(f"    → Nenhuma aprovação encontrada no histórico")
                            # Registrar mesmo sem aprovações
                            todos_resultados.append(AprovacaoMPD(
                                poco=poco,
                                seq_id=seq_id,
                                titulo=seq["titulo"],
                                publicado_em=seq["publicado"],
                                versao="",
                                aprovado_por="(sem histórico)",
                                aprovado_em="",
                                fonte_url=url_detalhe,
                            ))

                        # Voltar para a lista
                        voltar_para_lista(driver, wait)

                        # Re-pesquisar o poço para retomar a tabela
                        pesquisar_poco(driver, wait, poco)

                        # Navegar até a página correta novamente
                        for _ in range(pagina - 1):
                            if not tem_proxima_pagina(driver):
                                break

                        time.sleep(BACKOFF_SECS)

                    # Tentar próxima página
                    if not tem_proxima_pagina(driver):
                        log.info(f"  Última página processada para {poco}")
                        break
                    pagina += 1

            except Exception as e:
                log.error(f"Erro processando poço {poco}: {e}", exc_info=True)
                # Tentar recuperar a sessão
                try:
                    driver.get(ARQUIVADAS_URL)
                    time.sleep(3)
                except Exception:
                    log.error("Falha ao recuperar sessão. Recriando driver...")
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    driver = criar_driver()
                    wait = WebDriverWait(driver, WAIT_TIMEOUT,
                                        ignored_exceptions=[StaleElementReferenceException])
                    aguardar_login(driver)

            # Salvar progresso parcial a cada poço
            if todos_resultados:
                salvar_excel(todos_resultados)
                log.info(f"  Progresso salvo ({len(todos_resultados)} aprovações até agora)")

    except KeyboardInterrupt:
        log.info("Interrompido pelo usuário (Ctrl+C)")
    except Exception as e:
        log.error(f"Erro fatal: {e}", exc_info=True)
    finally:
        # Salvar resultados finais
        if todos_resultados:
            salvar_excel(todos_resultados)

        log.info("=" * 60)
        log.info(f"SCRAPER SEQOP – Fim")
        log.info(f"Total de aprovações extraídas: {len(todos_resultados)}")
        log.info(f"Resultados salvos em: {SAIDA_XLSX}")
        log.info("=" * 60)

        # Perguntar se quer fechar o navegador
        try:
            input("\nPressione Enter para fechar o navegador...")
        except EOFError:
            pass
        driver.quit()


# ── Ponto de entrada ────────────────────────────────────────────────────────

if __name__ == "__main__":
    executar_scraper()

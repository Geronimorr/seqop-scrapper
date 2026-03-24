import openpyxl

wb = openpyxl.load_workbook('resultado_aprovacoes_mpd.xlsx', read_only=True, data_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"Headers: {headers}")

# Find comment-related columns
for i, h in enumerate(headers):
    if h:
        print(f"  Col {i}: {h}")

# Check all rows for comment content
comment_col = None
for i, h in enumerate(headers):
    if h and 'coment' in str(h).lower():
        comment_col = i
        break

rows_with_comments = 0
total_rows = 0
comment_lengths = []
sample_comments = []

for row in ws.iter_rows(min_row=2, values_only=True):
    total_rows += 1
    if comment_col is not None:
        val = row[comment_col]
        if val and str(val).strip():
            rows_with_comments += 1
            comment_lengths.append(len(str(val)))
            if len(sample_comments) < 5:
                sample_comments.append(str(val)[:200])

print(f"\nTotal rows: {total_rows}")
if comment_col is not None:
    print(f"Comment column: {headers[comment_col]} (col {comment_col})")
    print(f"Rows with comments: {rows_with_comments}")
    if comment_lengths:
        print(f"Comment length: min={min(comment_lengths)}, max={max(comment_lengths)}, avg={sum(comment_lengths)//len(comment_lengths)}")
    print(f"\nSample comments:")
    for i, c in enumerate(sample_comments):
        print(f"  {i+1}. {c}")
else:
    print("No comment column found")

wb.close()

"""Gera o Excel de entrada com a lista de poços."""
import subprocess, sys
from pathlib import Path

_SHARED_PYTHON = Path(r"C:\SharedPython\venv\Scripts\python.exe")
if _SHARED_PYTHON.exists() and Path(sys.executable).resolve() != _SHARED_PYTHON.resolve():
    result = subprocess.run([str(_SHARED_PYTHON)] + sys.argv, check=False)
    sys.exit(result.returncode)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = Path(__file__).resolve().parent
SAIDA = SCRIPT_DIR / "pocos_entrada.xlsx"

POCOS = [
    "1-APS-57",
    "1-RJS-763DA",
    "3-SPS-114",
    "3-SPS-114A",
    "4-RJS-764",
    "7-BR-86DB-RJS",
    "7-BUZ-100DA-RJS",
    "7-BUZ-94D-RJS",
    "7-BUZ-95-RJS",
    "8-BUZ-89D-RJS",
    "7-OATP-3B-RJS",
    "8-BUZ-96D-RJS",
    "8-MRO-36-RJS",
    "3-SPS-111D",
    "4-SPS-112",
    "3-RJS-762",
    "8-ATP-8D-RJS",
    "7-STUP-10DA-RJS",
    "7-BUZ-90D-RJS",
]

wb = Workbook()
ws = wb.active
ws.title = "pocos"

# Cabeçalho
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

cell = ws.cell(row=1, column=1, value="poco")
cell.fill = header_fill
cell.font = header_font
cell.alignment = Alignment(horizontal="center")
cell.border = thin_border
ws.column_dimensions["A"].width = 25

for idx, poco in enumerate(POCOS, 2):
    c = ws.cell(row=idx, column=1, value=poco)
    c.border = thin_border

ws.auto_filter.ref = ws.dimensions
wb.save(SAIDA)
print(f"Arquivo criado: {SAIDA}")

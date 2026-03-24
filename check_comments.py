import json
import openpyxl

# Check comentarios_mpd.json
with open('comentarios_mpd.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

print(f"=== comentarios_mpd.json ===")
print(f"Total SEQOPs scraped: {len(data)}")
total_c = 0
total_m = 0
for s in data:
    nc = len(s.get('comentarios', []))
    nm = len(s.get('comentarios_mpd', []))
    total_c += nc
    total_m += nm
    sid = s.get('seq_id', '?')
    poco = s.get('poco', '?')
    titulo = s.get('titulo', '?')
    print(f"  {sid} | {poco} | {titulo}")
    print(f"    all comments: {nc}, mpd comments: {nm}")
print(f"Totals: {total_c} comments, {total_m} MPD")

# Check resultado_aprovacoes_mpd.xlsx
print(f"\n=== resultado_aprovacoes_mpd.xlsx ===")
wb = openpyxl.load_workbook('resultado_aprovacoes_mpd.xlsx', read_only=True, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
print(f"Headers: {rows[0]}")
all_rows = list(ws.iter_rows(min_row=2, values_only=True))
print(f"Data rows: {len(all_rows)}")

# Check for URL column
headers = rows[0]
url_col = None
for i, h in enumerate(headers):
    if h and 'url' in str(h).lower():
        url_col = i
        break

if url_col is not None:
    urls = set()
    for r in all_rows:
        if r[url_col]:
            urls.add(str(r[url_col]))
    print(f"Unique URLs: {len(urls)}")
    print(f"Sample URLs:")
    for u in list(urls)[:5]:
        print(f"  {u}")
else:
    print("No URL column found")
    # Check all columns for anything that looks like URL
    for i, h in enumerate(headers):
        sample_vals = [str(r[i]) for r in all_rows[:3] if r[i]]
        has_url = any('http' in v or 'seqop' in v.lower() for v in sample_vals)
        if has_url:
            print(f"  Col {i} ({h}) has URL-like values: {sample_vals[:2]}")

# Check unique pocos and sequences
poco_col = None
seq_col = None
for i, h in enumerate(headers):
    hs = str(h).lower() if h else ''
    if 'poco' in hs or 'poço' in hs:
        poco_col = i
    if 'seq' in hs:
        seq_col = i

if poco_col is not None:
    pocos = set(str(r[poco_col]) for r in all_rows if r[poco_col])
    print(f"Unique pocos: {len(pocos)}")

if seq_col is not None:
    seqs = set(str(r[seq_col]) for r in all_rows if r[seq_col])
    print(f"Unique sequences: {len(seqs)}")

wb.close()

# Check for other comment-related files
import os
for f in os.listdir('.'):
    if 'coment' in f.lower() or 'comment' in f.lower():
        size = os.path.getsize(f)
        print(f"\nFound: {f} ({size} bytes)")

# Check corpus folder
if os.path.exists('corpus'):
    print(f"\n=== corpus/ folder ===")
    for f in os.listdir('corpus'):
        print(f"  {f}")

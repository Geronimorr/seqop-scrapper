"""
Scraper SEQOP v3 – 2 Fases (Coleta URLs + Extração de Histórico)
=================================================================
Fase 1: Varre a tabela de Arquivadas, identifica sequências com badge MPD,
        abre "Visualizar" em nova aba para capturar a URL (e o ID).
Fase 2: Visita cada URL e extrai o histórico COMPLETO de aprovações MPD
        de todas as versões da sequência.

Uso:
    C:\\SharedPython\\venv\\Scripts\\python.exe scraper_seqop.py

Pré-requisitos:
    - Microsoft Edge instalado
    - selenium, openpyxl, pandas no venv compartilhado
    - Arquivo 'pocos_entrada.xlsx' com coluna "poco" (ou editar POCOS_DEFAULT)
"""

# ── Re-launch com o venv compartilhado ──────────────────────────────────────
import subprocess
import sys
from pathlib import Path

_SHARED_PYTHON = Path(r"C:\SharedPython\venv\Scripts\python.exe")

if _SHARED_PYTHON.exists() and Path(sys.executable).resolve() != _SHARED_PYTHON.resolve():
    print(f"[LAUNCH] Relançando com o Python do venv compartilhado: {_SHARED_PYTHON}")
    result = subprocess.run([str(_SHARED_PYTHON)] + sys.argv, check=False)
    sys.exit(result.returncode)

# ── Imports ─────────────────────────────────────────────────────────────────
import logging
import os
import re
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    WebDriverException,
)

# ── Configuração ────────────────────────────────────────────────────────────

BASE_URL = "https://csdpocos.petrobras.com.br/seqop"
ARQUIVADAS_URL = f"{BASE_URL}/#/arquivadas2"

SCRIPT_DIR = Path(__file__).resolve().parent
ENTRADA_XLSX = SCRIPT_DIR / "pocos_entrada.xlsx"
SAIDA_XLSX = SCRIPT_DIR / "resultado_aprovacoes_mpd.xlsx"
DIAG_HTML = SCRIPT_DIR / "diagnostico_detalhe.html"

DEBUGGING_PORT = 9222
WAIT_TIMEOUT = 25
PAUSA = 1.5

POCOS_DEFAULT = [
    "1-APS-57", "1-RJS-763DA", "3-SPS-114", "3-SPS-114A", "4-RJS-764",
    "7-BR-86DB-RJS", "7-BUZ-100DA-RJS", "7-BUZ-94D-RJS", "7-BUZ-95-RJS",
    "8-BUZ-89D-RJS", "7-OATP-3B-RJS", "8-BUZ-96D-RJS", "8-MRO-36-RJS",
    "3-SPS-111D", "4-SPS-112", "3-RJS-762", "8-ATP-8D-RJS",
    "7-STUP-10DA-RJS", "7-BUZ-90D-RJS",
]

# Regex para seção "Histórico para CSD-MPD:" na página de detalhe
# Formato: "v3: Aprovado por RW0X em 19/02/2026."
RE_HISTORICO_MPD = re.compile(
    r"v(\d+):\s*(Aprovado|Comentado|Rejeitado)\s+por\s+(\S+)\s+em\s+(\d{2}/\d{2}/\d{4})",
    re.IGNORECASE,
)

# Regex para comentários: "CSD-MPD - Versão: 1"
RE_COMENTARIO_CSD_MPD = re.compile(
    r"CSD-MPD\s*-\s*Vers[aã]o:\s*(\d+)",
    re.IGNORECASE,
)

# ── Logging ─────────────────────────────────────────────────────────────────

LOG_FILE = SCRIPT_DIR / "scraper_seqop.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("seqop_scraper")


# ── Dataclasses ─────────────────────────────────────────────────────────────

@dataclass
class SequenciaMPD:
    """Dados coletados na Fase 1 (tabela de listagem)."""
    poco: str
    titulo: str
    versao_tabela: str
    publicado_em: str
    url: str
    seq_id: str


@dataclass
class AprovacaoMPD:
    """Dados extraídos na Fase 2 (página de detalhe)."""
    poco: str
    seq_id: str
    titulo: str
    versao: str
    aprovado_por: str
    aprovado_em: str
    fonte_url: str
    carimbo_extracao_utc: str = field(
        default_factory=lambda: datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    )


# ── Índices das colunas (0-based) ──────────────────────────────────────────
COL_POCO = 2
COL_TITULO = 4
COL_VERSAO = 5
COL_PUBLICADO = 6
COL_APROVADORES = 7
COL_ACOES = 8


# ── Funções auxiliares ──────────────────────────────────────────────────────

def carregar_pocos() -> list[str]:
    if ENTRADA_XLSX.exists():
        log.info(f"Carregando poços de {ENTRADA_XLSX}")
        df = pd.read_excel(ENTRADA_XLSX, engine="openpyxl")
        col = None
        for c in df.columns:
            if c.strip().lower().replace("ç", "c") in ("poco", "poço"):
                col = c
                break
        if col is None:
            log.warning("Coluna 'poco' não encontrada. Usando lista padrão.")
            return POCOS_DEFAULT
        pocos = df[col].dropna().astype(str).str.strip().tolist()
        log.info(f"  → {len(pocos)} poços carregados")
        return pocos
    log.info(f"Usando lista padrão ({len(POCOS_DEFAULT)} poços).")
    return POCOS_DEFAULT


def _edge_profile_dir() -> Path:
    p = Path(os.environ.get("LOCALAPPDATA", "")) / "SeqopScraper_Edge"
    p.mkdir(parents=True, exist_ok=True)
    return p


def criar_driver() -> webdriver.Edge:
    try:
        log.info(f"Conectando ao Edge na porta {DEBUGGING_PORT}...")
        opts = EdgeOptions()
        opts.debugger_address = f"127.0.0.1:{DEBUGGING_PORT}"
        d = webdriver.Edge(options=opts)
        log.info("Conectado ao Edge existente.")
        return d
    except WebDriverException:
        log.info("Abrindo nova instância do Edge...")

    opts = EdgeOptions()
    profile = _edge_profile_dir()
    opts.add_argument(f"--user-data-dir={profile}")
    opts.add_argument("--profile-directory=Default")
    opts.add_argument("--start-maximized")
    opts.add_argument(f"--remote-debugging-port={DEBUGGING_PORT}")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    d = webdriver.Edge(options=opts)
    log.info("Edge aberto (perfil dedicado).")
    return d


# ── Login / navegação ──────────────────────────────────────────────────────

def _esta_logado(driver):
    try:
        menus = driver.find_elements(By.CSS_SELECTOR, "aside.menu")
        if menus and menus[0].is_displayed():
            for b in driver.find_elements(
                By.XPATH,
                "//button[contains(text(),'Entrar')] | //a[contains(text(),'Entrar')]",
            ):
                if b.is_displayed():
                    return False
            return True
        for t in driver.find_elements(By.TAG_NAME, "table"):
            try:
                if t.is_displayed():
                    return True
            except StaleElementReferenceException:
                continue
    except WebDriverException:
        pass
    return False


def aguardar_login(driver):
    driver.get(BASE_URL)
    log.info("Página SEQOP aberta.")
    time.sleep(4)
    if _esta_logado(driver):
        log.info("Já logado!")
        return
    log.info("=" * 50)
    log.info("  AGUARDANDO LOGIN MANUAL")
    log.info("  Faça login no navegador que abriu.")
    log.info("=" * 50)
    for t in range(120):
        time.sleep(5)
        if _esta_logado(driver):
            log.info("Login detectado!")
            time.sleep(2)
            return
        if t > 0 and t % 12 == 0:
            log.info(f"  ... aguardando ({t * 5 // 60} min)")
    raise TimeoutError("Timeout aguardando login (10 min).")


def navegar_arquivadas(driver):
    if "arquivadas" in driver.current_url.lower():
        return
    log.info("Navegando para Arquivadas...")
    try:
        link = driver.find_element(By.CSS_SELECTOR, 'a[href="/seqop/#/arquivadas2"]')
        driver.execute_script("arguments[0].click();", link)
        time.sleep(3)
        if "arquivadas" in driver.current_url.lower():
            return
    except (NoSuchElementException, WebDriverException):
        pass
    try:
        driver.execute_script("window.location.hash = '#/arquivadas2';")
        time.sleep(3)
        if "arquivadas" in driver.current_url.lower():
            return
    except WebDriverException:
        pass
    driver.get(ARQUIVADAS_URL)
    time.sleep(5)
    if not _esta_logado(driver):
        aguardar_login(driver)
        navegar_arquivadas(driver)


# ── Filtros / pesquisa ─────────────────────────────────────────────────────

def _limpar_filtros(driver):
    try:
        btn = driver.find_element(By.CSS_SELECTOR, "button.btn.btn-warning")
        driver.execute_script("arguments[0].click();", btn)
        time.sleep(1)
    except (NoSuchElementException, WebDriverException):
        pass


def pesquisar_poco(driver, wait, poco):
    log.info(f"Pesquisando poço: {poco}")
    _limpar_filtros(driver)
    time.sleep(0.5)
    campo = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input#poco")))
    campo.click()
    campo.send_keys(Keys.CONTROL, "a")
    campo.send_keys(Keys.DELETE)
    time.sleep(0.3)
    campo.send_keys(poco)
    time.sleep(0.6)
    try:
        btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.btn.btn-success")))
    except TimeoutException:
        btn = driver.find_element(By.CSS_SELECTOR, "button.btn.btn-success")
        driver.execute_script(
            "arguments[0].removeAttribute('disabled');"
            "arguments[0].classList.remove('disabled');", btn)
        time.sleep(0.3)
    driver.execute_script("arguments[0].click();", btn)
    time.sleep(2)
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr td")))
    except TimeoutException:
        time.sleep(3)


# ── Paginação ──────────────────────────────────────────────────────────────

def avancar_pagina(driver):
    try:
        btns = driver.find_elements(By.CSS_SELECTOR, 'button[aria-label="Go to next page"]')
        if btns:
            li = btns[0].find_element(By.XPATH, "./ancestor::li[1]")
            if "disabled" in (li.get_attribute("class") or ""):
                return False
            driver.execute_script("arguments[0].click();", btns[0])
            time.sleep(PAUSA + 1)
            return True
        return False
    except (NoSuchElementException, StaleElementReferenceException, WebDriverException):
        return False


# ==========================================================================
#  FASE 1 – Coletar URLs das sequências MPD (via nova aba)
# ==========================================================================

def _extrair_url_sequencia(driver, link_element) -> str | None:
    """Extrai a URL da sequência do link 'Visualizar'.
    Tenta ler o href diretamente (instantâneo) antes de abrir nova aba."""

    # ── Tentativa rápida: ler href direto via JS (sem abrir aba) ─────
    try:
        href = driver.execute_script(
            "return arguments[0].href || arguments[0].getAttribute('href') || '';",
            link_element,
        )
        if href and href.startswith("http"):
            return href
        # href relativo → montar URL completa
        if href and href.startswith("/"):
            origin = driver.execute_script("return window.location.origin;")
            return origin + href
        if href and "#" in href:
            origin = driver.execute_script("return window.location.origin;")
            return origin + "/seqop/" + href
    except Exception:
        pass

    # ── Fallback: abrir nova aba e capturar URL ──────────────────────
    original = driver.current_window_handle
    handles_antes = set(driver.window_handles)

    try:
        driver.execute_script("arguments[0].click();", link_element)
    except Exception as e:
        log.warning(f"    Erro ao clicar Visualizar: {e}")
        return None

    nova_abriu = False
    try:
        WebDriverWait(driver, 6).until(
            lambda d: len(d.window_handles) > len(handles_antes)
        )
        nova_abriu = True
    except TimeoutException:
        pass

    if not nova_abriu:
        try:
            ActionChains(driver).key_down(Keys.CONTROL).click(
                link_element
            ).key_up(Keys.CONTROL).perform()
            WebDriverWait(driver, 6).until(
                lambda d: len(d.window_handles) > len(handles_antes)
            )
            nova_abriu = True
        except Exception:
            pass

    if not nova_abriu:
        log.warning("    Não conseguiu abrir nova aba")
        return None

    nova_aba = (set(driver.window_handles) - handles_antes).pop()
    driver.switch_to.window(nova_aba)
    time.sleep(1)  # só precisamos da URL, não do conteúdo completo
    url = driver.current_url

    driver.close()
    driver.switch_to.window(original)
    time.sleep(0.3)

    return url


def fase1_coletar_sequencias(driver, wait, poco: str) -> list[SequenciaMPD]:
    """Varre todas as páginas e coleta URLs de sequências com badge MPD."""
    sequencias = []
    urls_vistas = set()  # evitar duplicatas
    pagina = 1

    while True:
        log.info(f"  Fase 1 – Página {pagina}")
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        num_rows = len(rows)
        log.info(f"    {num_rows} linhas na tabela")

        for row_idx in range(num_rows):
            # Re-localizar linhas (o DOM pode mudar após fechar aba)
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            if row_idx >= len(rows):
                break
            row = rows[row_idx]

            try:
                tds = row.find_elements(By.TAG_NAME, "td")
                if len(tds) <= COL_ACOES:
                    continue

                # Checar se tem badge MPD
                aprov_spans = tds[COL_APROVADORES].find_elements(
                    By.CSS_SELECTOR, "span.aprovador"
                )
                has_mpd = any(
                    sp.text.strip().upper() == "MPD" for sp in aprov_spans
                )
                if not has_mpd:
                    continue

                # Dados básicos
                titulo_spans = tds[COL_TITULO].find_elements(By.TAG_NAME, "span")
                titulo = (
                    (titulo_spans[0].get_attribute("title") or titulo_spans[0].text.strip())
                    if titulo_spans else tds[COL_TITULO].text.strip()
                )
                versao_spans = tds[COL_VERSAO].find_elements(By.TAG_NAME, "span")
                versao = (versao_spans[0].text.strip() if versao_spans
                          else tds[COL_VERSAO].text.strip())
                publicado = tds[COL_PUBLICADO].text.strip()

                log.info(f"    MPD [{row_idx+1}/{num_rows}]: {titulo[:60]}  (v{versao})")

                # Extrair URL (rápido via href, fallback p/ nova aba)
                link_viz = tds[COL_ACOES].find_element(
                    By.CSS_SELECTOR, 'a[title*="Visualizar"]'
                )
                url = _extrair_url_sequencia(driver, link_viz)

                if not url:
                    log.warning(f"    URL não capturada para: {titulo[:50]}")
                    continue

                if url in urls_vistas:
                    log.info(f"    URL duplicada, pulando: {url}")
                    continue
                urls_vistas.add(url)

                # Extrair ID numérico da URL
                m_id = re.search(r"(\d{3,})", url.split("#")[-1])
                seq_id = m_id.group(1) if m_id else ""

                log.info(f"    → URL: {url}  (ID: {seq_id})")

                sequencias.append(SequenciaMPD(
                    poco=poco, titulo=titulo, versao_tabela=versao,
                    publicado_em=publicado, url=url, seq_id=seq_id,
                ))

            except StaleElementReferenceException:
                log.warning(f"    Stale na linha {row_idx}, pulando.")
            except Exception as e:
                log.warning(f"    Erro na linha {row_idx}: {e}")

        if not avancar_pagina(driver):
            break
        pagina += 1

    log.info(f"  Fase 1: {len(sequencias)} sequências MPD para {poco}")
    return sequencias


# ==========================================================================
#  FASE 2 – Visitar cada URL e extrair histórico completo MPD
# ==========================================================================

_diag_salvo = False


def fase2_extrair_historico(driver, wait, seq: SequenciaMPD) -> list[AprovacaoMPD]:
    r"""Abre a URL e extrai o histórico COMPLETO de aprovações MPD.

    Estrutura real da página de detalhe:
      - div.ma-3 > h6 "Histórico para CSD-MPD:" seguida de divs com:
        "v3: Aprovado por RW0X em 19/02/2026."
        "v2: Aprovado por RW0X em 18/02/2026."
        "v1: Comentado por RW0X em 17/02/2026."
      - Comentários em div.grupoComentario com:
        span "Nome Completo há X dias"
        span "CSD-MPD - Versão: N"

    Estratégias (em ordem):
      1. Seção "Histórico para CSD-MPD" (extrai versão, ação, código, data)
      2. Comentários CSD-MPD (extrai nome completo, versão)
      3. Regex no texto completo da página
    """
    global _diag_salvo

    log.info(f"  Fase 2 – {seq.titulo[:50]}  →  {seq.url}")
    resultados = []

    # Extrair o ID mongo da URL para validar que a página certa carregou
    url_hash = seq.url.split("#")[-1]  # "/view/69971ff5b47a5f00126e85ef"
    m_mongo = re.search(r"([a-f0-9]{24})", url_hash)
    expected_id = m_mongo.group(1) if m_mongo else ""

    # Navegar — em SPA com hash, precisamos garantir que o conteúdo mudou
    old_id = ""
    try:
        old_el = driver.find_element(By.CSS_SELECTOR, "div[idsequencia]")
        old_id = old_el.get_attribute("idsequencia") or ""
    except Exception:
        pass

    driver.get(seq.url)

    # Se o ID esperado é igual ao que já estava, forçar reload
    if expected_id and expected_id == old_id:
        # Mesmo ID, página já está certa — sem espera
        pass
    elif expected_id:
        # Aguardar até o atributo idsequencia mudar para o ID esperado
        try:
            WebDriverWait(driver, 15).until(
                lambda d: d.find_element(
                    By.CSS_SELECTOR, f'div[idsequencia="{expected_id}"]'
                )
            )
        except TimeoutException:
            # Fallback: forçar navegação via JS
            driver.execute_script(f"window.location.hash = '#/view/{expected_id}';")
            time.sleep(2)
    else:
        # Sem ID mongo, esperar genérico
        try:
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.dadosVersao"))
            )
        except TimeoutException:
            time.sleep(2)

    # ── Diagnóstico: salvar HTML da 1ª página de detalhe ─────────────
    if not _diag_salvo:
        try:
            DIAG_HTML.write_text(driver.page_source, encoding="utf-8")
            log.info(f"    [DIAG] HTML salvo em {DIAG_HTML}")
            _diag_salvo = True
        except Exception as e:
            log.warning(f"    [DIAG] Falha: {e}")

    # ── Capturar título e poço da página ─────────────────────────────
    titulo_pagina = seq.titulo
    try:
        h5s = driver.find_elements(By.CSS_SELECTOR, "div.SideBar h5")
        if len(h5s) >= 2:
            titulo_pagina = h5s[1].text.strip() or titulo_pagina
    except Exception:
        pass

    # ── Mapa de nomes (sempre inicializado) ─────────────────────────
    mapa_nomes = {}

    # ── Estratégia 1: Seção "Histórico para CSD-MPD" ────────────────
    historico_encontrado = _extrair_historico_mpd_sidebar(driver)

    if historico_encontrado:
        mapa_nomes = _extrair_nomes_mpd_comentarios(driver)

        for entry in historico_encontrado:
            versao = entry["versao"]
            acao = entry["acao"]
            codigo = entry["codigo"]
            data = entry["data"]
            nome = mapa_nomes.get(codigo, codigo)

            resultados.append(AprovacaoMPD(
                poco=seq.poco, seq_id=seq.seq_id, titulo=titulo_pagina,
                versao=versao,
                aprovado_por=f"{nome} ({acao})" if acao.lower() != "aprovado" else nome,
                aprovado_em=data, fonte_url=seq.url,
            ))

        log.info(f"    → {len(resultados)} entradas MPD (sidebar)")
        return resultados

    # ── Estratégia 2: Comentários CSD-MPD ────────────────────────────
    resultados = _extrair_aprovacoes_comentarios(driver, seq)

    if resultados:
        log.info(f"    → {len(resultados)} entradas MPD (comentários)")
        return resultados

    # ── Estratégia 3: Regex no texto completo ────────────────────────
    try:
        page_text = driver.find_element(By.TAG_NAME, "body").text
    except Exception:
        page_text = ""

    for m in RE_HISTORICO_MPD.finditer(page_text):
        versao, acao, codigo, data = m.group(1), m.group(2), m.group(3), m.group(4)
        nome = mapa_nomes.get(codigo, codigo)
        resultados.append(AprovacaoMPD(
            poco=seq.poco, seq_id=seq.seq_id, titulo=titulo_pagina,
            versao=versao,
            aprovado_por=f"{nome} ({acao})" if acao.lower() != "aprovado" else nome,
            aprovado_em=data, fonte_url=seq.url,
        ))

    if resultados:
        log.info(f"    → {len(resultados)} entradas MPD (regex texto)")
    else:
        log.warning(f"    → Nenhuma aprovação MPD encontrada")
        resultados.append(AprovacaoMPD(
            poco=seq.poco, seq_id=seq.seq_id, titulo=titulo_pagina,
            versao=seq.versao_tabela, aprovado_por="(ver página manualmente)",
            aprovado_em="", fonte_url=seq.url,
        ))

    return resultados


def _extrair_historico_mpd_sidebar(driver) -> list[dict]:
    r"""Extrai entradas da seção 'Histórico para CSD-MPD:' na sidebar.

    Retorna lista de dicts: {versao, acao, codigo, data}
    """
    entradas = []

    try:
        # Usar JS para encontrar o texto do container de uma vez (mais rápido)
        texto_container = driver.execute_script("""
            var h6s = document.querySelectorAll('div.ma-3 h6, h6');
            for (var i = 0; i < h6s.length; i++) {
                var t = h6s[i].textContent.toUpperCase();
                if (t.indexOf('CSD-MPD') >= 0 || (t.indexOf('HIST') >= 0 && t.indexOf('MPD') >= 0)) {
                    return h6s[i].parentElement.textContent;
                }
            }
            return '';
        """)

        if not texto_container:
            return entradas

        for m in RE_HISTORICO_MPD.finditer(texto_container):
            entradas.append({
                "versao": m.group(1),
                "acao": m.group(2),
                "codigo": m.group(3),
                "data": m.group(4),
            })

    except Exception as e:
        log.warning(f"    Erro ao extrair histórico sidebar: {e}")

    return entradas


def _extrair_nomes_mpd_comentarios(driver) -> dict:
    """Mapeia códigos de usuário → nomes completos a partir dos comentários CSD-MPD.

    Procura em div.grupoComentario > div.emLinhaAutor.autor-versao:
      span[0] = "Nome Completo há X dias"
      span[1] = "CSD-MPD - Versão: N"
    """
    mapa = {}

    try:
        grupos = driver.find_elements(By.CSS_SELECTOR, "div.grupoComentario")
        for grupo in grupos:
            try:
                autor_div = grupo.find_element(
                    By.CSS_SELECTOR, "div.emLinhaAutor.autor-versao"
                )
                spans = autor_div.find_elements(By.TAG_NAME, "span")
                if len(spans) < 2:
                    continue

                info_grupo = spans[1].text.strip()  # "CSD-MPD - Versão: 1"
                if "CSD-MPD" not in info_grupo.upper():
                    continue

                nome_raw = spans[0].text.strip()  # "Nome Completo há 7 dias"
                # Remover "há X dias/horas/minutos"
                nome = re.sub(
                    r"\s+h[aá]\s+\d+\s+\w+$", "", nome_raw, flags=re.IGNORECASE
                ).strip()

                if nome:
                    # Não temos o código diretamente, mas armazenamos
                    # para tentativa de cross-reference
                    v_m = RE_COMENTARIO_CSD_MPD.search(info_grupo)
                    versao = v_m.group(1) if v_m else ""
                    # Guardar por versão (pode ajudar no cross-ref)
                    mapa[f"v{versao}"] = nome
                    mapa[nome] = nome  # identidade

                    log.info(f"    Comentário CSD-MPD: {nome} (Versão {versao})")

            except (NoSuchElementException, StaleElementReferenceException):
                continue

    except Exception as e:
        log.warning(f"    Erro ao extrair nomes dos comentários: {e}")

    return mapa


def _extrair_aprovacoes_comentarios(driver, seq: SequenciaMPD) -> list[AprovacaoMPD]:
    """Extrai aprovações MPD dos comentários (fallback se sidebar falhar)."""
    resultados = []

    try:
        grupos = driver.find_elements(By.CSS_SELECTOR, "div.grupoComentario")
        for grupo in grupos:
            try:
                autor_div = grupo.find_element(
                    By.CSS_SELECTOR, "div.emLinhaAutor.autor-versao"
                )
                spans = autor_div.find_elements(By.TAG_NAME, "span")
                if len(spans) < 2:
                    continue

                info_grupo = spans[1].text.strip()
                if "CSD-MPD" not in info_grupo.upper():
                    continue

                nome_raw = spans[0].text.strip()
                # Extrair data aproximada de "há X dias/horas"
                data_aprox = ""
                m_ha = re.search(r"h[aá]\s+(\d+)\s+(\w+)", nome_raw, re.IGNORECASE)
                if m_ha:
                    qtd = int(m_ha.group(1))
                    unidade = m_ha.group(2).lower()
                    from datetime import timedelta
                    hoje = datetime.now()
                    if "dia" in unidade:
                        data_calc = hoje - timedelta(days=qtd)
                    elif "hora" in unidade:
                        data_calc = hoje - timedelta(hours=qtd)
                    elif "min" in unidade:
                        data_calc = hoje - timedelta(minutes=qtd)
                    elif "sem" in unidade:  # semana
                        data_calc = hoje - timedelta(weeks=qtd)
                    elif "m" in unidade and "es" in unidade:  # mês/meses
                        data_calc = hoje - timedelta(days=qtd * 30)
                    else:
                        data_calc = hoje - timedelta(days=qtd)
                    data_aprox = data_calc.strftime("%d/%m/%Y")

                nome = re.sub(
                    r"\s+h[aá]\s+\d+\s+\w+$", "", nome_raw, flags=re.IGNORECASE
                ).strip()

                v_m = RE_COMENTARIO_CSD_MPD.search(info_grupo)
                versao = v_m.group(1) if v_m else ""

                # Tentar pegar o texto do comentário para ver se aprovou
                texto_comentario = ""
                try:
                    card = grupo.find_element(By.CSS_SELECTOR, "div.card-text")
                    texto_comentario = card.text.strip().lower()
                except Exception:
                    pass

                acao = "Comentou"
                if any(w in texto_comentario for w in
                       ["aprovad", "aprovo", "de acordo", "sem ressalvas",
                        "sem comentários adicionais", "sequência aprovada",
                        "seqop aprovada", "sem objeção", "sem objecao"]):
                    acao = "Aprovado"

                resultados.append(AprovacaoMPD(
                    poco=seq.poco, seq_id=seq.seq_id, titulo=seq.titulo,
                    versao=versao,
                    aprovado_por=f"{nome} ({acao})" if acao != "Aprovado" else nome,
                    aprovado_em=data_aprox if data_aprox else "",
                    fonte_url=seq.url,
                ))

            except (NoSuchElementException, StaleElementReferenceException):
                continue

    except Exception as e:
        log.warning(f"    Erro ao extrair dos comentários: {e}")

    return resultados


# ── Excel ──────────────────────────────────────────────────────────────────

def salvar_excel(resultados: list[AprovacaoMPD]):
    if not resultados:
        return

    df = pd.DataFrame([asdict(r) for r in resultados])
    df.sort_values(by=["poco", "titulo", "versao"], inplace=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "aprovacoes_mpd"

    colunas = [
        ("Poço", 22), ("Seq ID", 10), ("Título", 55), ("Versão", 10),
        ("Aprovado Por (MPD)", 40), ("Aprovado Em", 18),
        ("URL Fonte", 55), ("Extração (UTC)", 22),
    ]

    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for ci, (nome, larg) in enumerate(colunas, 1):
        c = ws.cell(row=1, column=ci, value=nome)
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        ws.column_dimensions[c.column_letter].width = larg

    for ri, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for ci, val in enumerate(row, 1):
            # Converter datas string → datetime para o Excel reconhecer
            if ci == 6 and isinstance(val, str) and val:  # coluna "Aprovado Em"
                for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M"):
                    try:
                        val = datetime.strptime(val, fmt)
                        break
                    except (ValueError, TypeError):
                        continue
            if ci == 8 and isinstance(val, str) and val:  # coluna "Extração (UTC)"
                try:
                    val = datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
                except (ValueError, TypeError):
                    pass
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = border
            c.alignment = Alignment(vertical="center")
            if ci == 6 and isinstance(val, datetime):
                c.number_format = "DD/MM/YYYY"
            elif ci == 8 and isinstance(val, datetime):
                c.number_format = "YYYY-MM-DD HH:MM:SS"

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(SAIDA_XLSX)
    log.info(f"Excel salvo: {SAIDA_XLSX}  ({len(resultados)} linhas)")


# ── Scraper principal ──────────────────────────────────────────────────────

def executar_scraper():
    log.info("=" * 60)
    log.info("SCRAPER SEQOP v3 – 2 Fases (URLs + Histórico)")
    log.info(f"Data/hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    log.info("=" * 60)

    pocos = carregar_pocos()
    log.info(f"Poços a processar: {len(pocos)}")

    todas_sequencias: list[SequenciaMPD] = []
    todos_resultados: list[AprovacaoMPD] = []

    driver = criar_driver()
    wait = WebDriverWait(driver, WAIT_TIMEOUT,
                         ignored_exceptions=[StaleElementReferenceException])

    try:
        aguardar_login(driver)
        navegar_arquivadas(driver)

        # ━━━━ FASE 1: Coletar URLs ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        log.info("=" * 60)
        log.info("FASE 1 – Coletando URLs de sequências MPD")
        log.info("=" * 60)

        for idx, poco in enumerate(pocos, 1):
            log.info("-" * 50)
            log.info(f"[{idx}/{len(pocos)}] Poço: {poco}")
            try:
                navegar_arquivadas(driver)
                pesquisar_poco(driver, wait, poco)
                seqs = fase1_coletar_sequencias(driver, wait, poco)
                todas_sequencias.extend(seqs)
                log.info(f"  Total acumulado: {len(todas_sequencias)} sequências")
            except Exception as e:
                log.error(f"Erro Fase 1 – poço {poco}: {e}", exc_info=True)
                try:
                    driver.get(ARQUIVADAS_URL)
                    time.sleep(3)
                except Exception:
                    driver = _recuperar_driver(driver)

        log.info("=" * 60)
        log.info(f"FASE 1 COMPLETA: {len(todas_sequencias)} sequências MPD")
        for s in todas_sequencias:
            log.info(f"  {s.poco} | {s.seq_id} | {s.titulo[:50]} | {s.url}")
        log.info("=" * 60)

        # ━━━━ FASE 2: Extrair históricos ━━━━━━━━━━━━━━━━━━━━━━━━━━━
        log.info("FASE 2 – Extraindo histórico de aprovações MPD")
        log.info("=" * 60)

        for idx, seq in enumerate(todas_sequencias, 1):
            log.info(f"[{idx}/{len(todas_sequencias)}] {seq.poco} – {seq.titulo[:50]}")
            try:
                aprovacoes = fase2_extrair_historico(driver, wait, seq)
                todos_resultados.extend(aprovacoes)
            except Exception as e:
                log.error(f"Erro Fase 2 – seq {seq.seq_id}: {e}", exc_info=True)
                todos_resultados.append(AprovacaoMPD(
                    poco=seq.poco, seq_id=seq.seq_id, titulo=seq.titulo,
                    versao=seq.versao_tabela, aprovado_por=f"(ERRO: {e})",
                    aprovado_em="", fonte_url=seq.url,
                ))

            if idx % 5 == 0 and todos_resultados:
                salvar_excel(todos_resultados)
                log.info(f"  Progresso: {len(todos_resultados)} aprovações")

    except KeyboardInterrupt:
        log.info("Interrompido (Ctrl+C)")
    except Exception as e:
        log.error(f"Erro fatal: {e}", exc_info=True)
    finally:
        if todos_resultados:
            salvar_excel(todos_resultados)
        log.info("=" * 60)
        log.info(f"Total: {len(todos_resultados)} aprovações MPD → {SAIDA_XLSX}")
        log.info("=" * 60)
        try:
            input("\nPressione Enter para fechar o navegador...")
        except EOFError:
            pass
        driver.quit()


def _recuperar_driver(driver):
    log.error("Recriando driver...")
    try:
        driver.quit()
    except Exception:
        pass
    d = criar_driver()
    aguardar_login(d)
    return d


if __name__ == "__main__":
    executar_scraper()

"""
Gerador de Checklist por Abas — Fluxo Operacional MPD
=====================================================
Gera Excel + HTML organizados pelas 11 operações do fluxo cronológico MPD,
usando os dados enriquecidos de 256 SEQOPs (seqops_enriquecidas.json).

Cada aba contém:
  • Blocos MPD consolidados (com frequência de ocorrência)
  • Comentários dos revisores CSD-MPD (DIRETA/INDIRETA)
  • Pontos de verificação extraídos pela IA
  • Temas recorrentes com frequência

Entrada:
    seqops_enriquecidas.json  (256 SEQOPs com analise_ia)
    fluxo_operacoes_mpd.py    (11 operações com aba_id e sub_tipos)

Saída:
    checklist_operacional_mpd.xlsx  (11 abas + Resumo + Dados)
    checklist_operacional_mpd.html  (página com tabs navegáveis)

Uso:
    C:\\SharedPython\\venv\\Scripts\\python.exe gerar_checklist_abas.py
"""

# ── Re-launch com venv compartilhado ───────────────────────────────────────
import subprocess, sys
from pathlib import Path

_SHARED_PYTHON = Path(r"C:\SharedPython\venv\Scripts\python.exe")
if _SHARED_PYTHON.exists() and Path(sys.executable).resolve() != _SHARED_PYTHON.resolve():
    result = subprocess.run([str(_SHARED_PYTHON)] + sys.argv, check=False)
    sys.exit(result.returncode)

import json, re, logging, io
from collections import Counter, defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.open(sys.stdout.fileno(), mode="w", encoding="utf-8", closefd=False)

# ── Configuração ───────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
ENRIQUECIDAS_JSON = SCRIPT_DIR / "seqops_enriquecidas.json"
SAIDA_XLSX = SCRIPT_DIR / "checklist_operacional_mpd.xlsx"
SAIDA_HTML = SCRIPT_DIR / "checklist_operacional_mpd.html"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(
            open(sys.stdout.fileno(), mode="w", encoding="utf-8", closefd=False)
        ),
    ],
)
log = logging.getLogger("checklist_abas")

# ── Importar fluxo de operações ────────────────────────────────────────────
from fluxo_operacoes_mpd import FLUXO_OPERACOES, METADADOS


# ==========================================================================
#  MAPEAMENTO tipo_operacao → aba_id
# ==========================================================================

# Padrões regex para cada aba, priorizando do mais específico ao mais geral.
# A primeira correspondência ganha.
MAPEAMENTO_REGEX = [
    # 11 - Completação / PACI / Cauda
    ("completacao_paci", [
        r"completacao|paci|cauda|vif|shifter|abandono|tampao|amortecimento|"
        r"estimulacao_acida|nrv|recuperacao_wb|recuperacao_packer|"
        r"fechamento_vif|manobra_stx"
    ]),
    # 10 - Contingência / FMCD / PMCD / Pescaria
    ("contingencia_pmcd_fmcd", [
        r"contingencia|pescaria|combate_perda|well_defend|"
        r"pmcd(?!.*perf)|fmcd(?!.*perf)|conversao.*pmcd|conversao.*fmcd|"
        r"bullheading|corte_bpp"
    ]),
    # 7 - Testemunhagem
    ("testemunhagem", [
        r"testemunhagem"
    ]),
    # 9 - Teste influxo / Teste BOP periódico
    ("teste_influxo", [
        r"teste_influxo|teste_bop(?!.*descida)(?!.*instalacao)|"
        r"teste_bop_bsrs|teste_bop_com|retirada_lmrp|"
        r"desconexao.*bop|desconexao.*reconexao|"
        r"desconexao_bop_dmm|desconexao_reinstalacao"
    ]),
    # 8 - Retirada BHA / Manobra
    ("retirada_bha", [
        r"retirada_bha|retirada_coluna|manobra_mpd|troca_broca|"
        r"descida_pata_elefante"
    ]),
    # 6 - Perfuração com MPD
    ("perfuracao", [
        r"perfuracao(?!.*fase)|perf_mpd|perf_.*mpd|sidetrack.*perf|"
        r"perfuracao_mpd|perfuracao_pmcd|perfuracao_fmcd|"
        r"perfuracao_poco|troca_fluido_perf|perfuracao_fase"
    ]),
    # 5 - Corte cimento / FIT / DLOT
    ("corte_cimento_fit", [
        r"corte_cimento|fit_pre|dfit|dlot|microfrac|"
        r"corte_bpp|tampao_injecao|teste_mpd_pos_cimentacao|"
        r"teste_mpd_corte|teste_sistema_mpd_pos_corte"
    ]),
    # 4 - Troca de fluido / Condicionamento
    ("troca_fluido", [
        r"troca_fluido(?!.*finger)|condicionamento|substituicao_fluido|"
        r"condicionamento_pos_perfilagem|condicionamento_reservatorio"
    ]),
    # 3 - Fingerprint / Treinamento
    ("fingerprint", [
        r"fingerprint|treinamento|choke_drill|mpd_drill|calibracao|"
        r"troca_fluido.*finger|troca_fluido.*drill"
    ]),
    # 1 - Instalação BOP
    ("instalacao_bop", [
        r"instalacao_bop|descida_bop|assentamento_bop|dmm.*bop|"
        r"navegacao.*bop|reconexao_lmrp.*junta|descida_lmrp|"
        r"instalacao_cvu|retirada_bop_descida_bop|"
        r"DMM_instalacao|dmm_descida_bop"
    ]),
    # 2 - Descida BHA / Teste MPD
    ("descida_bha_teste", [
        r"descida_bha|montagem_bha|descida_bha_sidetrack|preparacao_teste|"
        r"montagem_bha_teste|navegacao.*descida_bop.*comissionamento"
    ]),
]


def mapear_tipo_para_aba(tipo_op: str) -> str:
    """Mapeia o tipo_operacao da IA para um aba_id do fluxo."""
    tipo_lower = tipo_op.lower().replace("-", "_")
    for aba_id, patterns in MAPEAMENTO_REGEX:
        for pat in patterns:
            if re.search(pat, tipo_lower):
                return aba_id
    # Fallback para termos genéricos
    if "perf" in tipo_lower:
        return "perfuracao"
    if "bop" in tipo_lower or "junta" in tipo_lower:
        return "instalacao_bop"
    if "bha" in tipo_lower:
        return "descida_bha_teste"
    if "fluido" in tipo_lower or "condicion" in tipo_lower:
        return "troca_fluido"
    if "finger" in tipo_lower or "treina" in tipo_lower:
        return "fingerprint"
    if "fmcd" in tipo_lower or "pmcd" in tipo_lower:
        return "contingencia_pmcd_fmcd"
    if "complet" in tipo_lower or "paci" in tipo_lower:
        return "completacao_paci"
    return "perfuracao"  # default


# ==========================================================================
#  CARREGAR E AGREGAR  DADOS
# ==========================================================================

def carregar_dados():
    """Carrega o JSON enriquecido e agrega por aba."""
    log.info(f"Carregando {ENRIQUECIDAS_JSON}...")
    with open(ENRIQUECIDAS_JSON, "r", encoding="utf-8") as f:
        seqops = json.load(f)

    log.info(f"  {len(seqops)} SEQOPs carregadas")

    # Estrutura de agregação por aba
    abas = {}
    for op in FLUXO_OPERACOES:
        abas[op["aba_id"]] = {
            "info": op,
            "seqops": [],
            "blocos_mpd": [],
            "comentarios": [],
            "temas": Counter(),
            "pocos": Counter(),
        }

    # Distribuir SEQOPs nas abas
    sem_mapeamento = []
    for s in seqops:
        ai = s.get("analise_ia")
        if not isinstance(ai, dict) or "blocos_mpd" not in ai:
            continue

        tipo_op = ai.get("tipo_operacao", "")
        aba_id = mapear_tipo_para_aba(tipo_op)

        if aba_id not in abas:
            sem_mapeamento.append((tipo_op, s.get("poco"), s.get("titulo", "")[:40]))
            continue

        aba = abas[aba_id]
        aba["seqops"].append(s)
        aba["pocos"][s.get("poco", "?")] += 1

        # Blocos MPD
        for b in ai.get("blocos_mpd", []):
            aba["blocos_mpd"].append({
                "nome": b.get("nome", ""),
                "relevancia": b.get("rel", ""),
                "motivo": b.get("motivo", ""),
                "itens": b.get("itens", b.get("itens_seqop", [])),
                "poco": s.get("poco", ""),
                "titulo_seqop": s.get("titulo", ""),
            })

        # Comentários (chave compacta da IA: i, rel, bloco, temas, resumo, check)
        for c in ai.get("comentarios", ai.get("comentarios_analisados", [])):
            rel = c.get("rel", c.get("relevancia_mpd", ""))
            if rel in ("DIRETA", "INDIRETA"):
                aba["comentarios"].append({
                    "indice": c.get("i", c.get("indice", "")),
                    "rel": rel,
                    "bloco": c.get("bloco", c.get("bloco_referenciado", "")),
                    "temas": c.get("temas", c.get("temas_mpd", [])),
                    "resumo": c.get("resumo", ""),
                    "check": c.get("check", c.get("ponto_verificacao", "")),
                    "poco": s.get("poco", ""),
                    "titulo_seqop": s.get("titulo", ""),
                })
                for t in c.get("temas", c.get("temas_mpd", [])):
                    aba["temas"][t] += 1

    if sem_mapeamento:
        log.warning(f"  {len(sem_mapeamento)} tipos sem mapeamento:")
        for t, p, tit in sem_mapeamento[:5]:
            log.warning(f"    {t} ({p} - {tit})")

    # Stats
    total_seqops = sum(len(a["seqops"]) for a in abas.values())
    total_blocos = sum(len(a["blocos_mpd"]) for a in abas.values())
    total_coments = sum(len(a["comentarios"]) for a in abas.values())

    log.info(f"  Mapeadas: {total_seqops} SEQOPs → {sum(1 for a in abas.values() if a['seqops'])} abas")
    log.info(f"  Blocos MPD: {total_blocos}")
    log.info(f"  Comentários relevantes: {total_coments}")

    for op in FLUXO_OPERACOES:
        aba = abas[op["aba_id"]]
        log.info(f"    {op['aba_nome_curto']:<14} → {len(aba['seqops']):>3} SEQOPs, "
                 f"{len(aba['blocos_mpd']):>4} blocos, {len(aba['comentarios']):>3} coments")

    return abas


def consolidar_blocos(blocos: list[dict]) -> list[dict]:
    """Consolida blocos MPD similares, calculando frequência."""
    agrupados = defaultdict(lambda: {
        "nome": "", "freq": 0, "relevancias": Counter(),
        "motivos": [], "pocos": set(), "itens_exemplo": [],
    })

    for b in blocos:
        nome_key = re.sub(r'\s+', ' ', b["nome"].strip().lower())
        # Agrupar por nome normalizado
        g = agrupados[nome_key]
        g["nome"] = b["nome"]
        g["freq"] += 1
        g["relevancias"][b.get("relevancia", "MEDIA")] += 1
        if b.get("motivo") and len(g["motivos"]) < 3:
            g["motivos"].append(b["motivo"][:150])
        g["pocos"].add(b.get("poco", ""))
        if b.get("itens") and len(g["itens_exemplo"]) < 2:
            g["itens_exemplo"].extend(b["itens"][:2])

    # Ordenar por frequência
    resultado = []
    for key, g in sorted(agrupados.items(), key=lambda x: -x[1]["freq"]):
        rel_top = g["relevancias"].most_common(1)[0][0] if g["relevancias"] else "MEDIA"
        resultado.append({
            "nome": g["nome"],
            "frequencia": g["freq"],
            "relevancia": rel_top,
            "n_pocos": len(g["pocos"]),
            "motivos": list(set(g["motivos"]))[:3],
            "itens_exemplo": g["itens_exemplo"][:3],
        })

    return resultado


# ==========================================================================
#  GERAR EXCEL
# ==========================================================================

# Paleta de cores
COR_AZUL_ESCURO = "0D2137"
COR_AZUL_MEDIO  = "1A5276"
COR_AZUL_CLARO  = "D4E6F1"
COR_CINZA_CLARO = "F5F6FA"
COR_VERDE       = "27AE60"
COR_AMARELO     = "F39C12"
COR_VERMELHO    = "E74C3C"
COR_BRANCO      = "FFFFFF"
COR_CINZA       = "BDC3C7"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _estilo_hdr():
    return {
        "fill": PatternFill(start_color=COR_AZUL_ESCURO, end_color=COR_AZUL_ESCURO, fill_type="solid"),
        "font": Font(color=COR_BRANCO, bold=True, size=10),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": THIN_BORDER,
    }


def _estilo_subhdr():
    return {
        "fill": PatternFill(start_color=COR_AZUL_MEDIO, end_color=COR_AZUL_MEDIO, fill_type="solid"),
        "font": Font(color=COR_BRANCO, bold=True, size=10),
        "alignment": Alignment(vertical="center", wrap_text=True),
        "border": THIN_BORDER,
    }


def _aplicar_estilo(cell, estilo: dict):
    for k, v in estilo.items():
        setattr(cell, k, v)


def _cor_relevancia(rel: str) -> PatternFill:
    if rel == "ALTA":
        return PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    if rel == "MEDIA":
        return PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    return PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")


def gerar_excel(abas: dict):
    """Gera o Excel com 11+ abas."""
    log.info("Gerando Excel...")
    wb = Workbook()

    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Aba Resumo ────────────────────────────────────────────────────
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo.sheet_properties.tabColor = COR_AZUL_ESCURO

    # Título
    ws_resumo.merge_cells("A1:H1")
    t = ws_resumo.cell(row=1, column=1,
        value="Checklist Operacional MPD — Fluxo Cronológico por Abas")
    t.font = Font(bold=True, size=16, color=COR_AZUL_ESCURO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo.row_dimensions[1].height = 40

    ws_resumo.merge_cells("A2:H2")
    ws_resumo.cell(row=2, column=1,
        value=f"Gerado: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
              f"{METADADOS['total_seqops']} SEQOPs | "
              f"{METADADOS['total_pocos']} poços | "
              f"Modelo: gpt-5.1"
    ).font = Font(italic=True, size=10, color="666666")

    # Tabela resumo
    cols_resumo = [
        ("#", 5), ("Aba", 14), ("Operação", 35), ("SEQOPs", 10),
        ("Blocos MPD", 12), ("Comentários", 12), ("Temas", 10), ("Poços", 8),
    ]
    hr = 4
    for ci, (nome, larg) in enumerate(cols_resumo, 1):
        c = ws_resumo.cell(row=hr, column=ci, value=nome)
        _aplicar_estilo(c, _estilo_hdr())
        ws_resumo.column_dimensions[get_column_letter(ci)].width = larg

    r = hr + 1
    total_blocos = 0
    total_coments = 0
    for op in FLUXO_OPERACOES:
        aba = abas[op["aba_id"]]
        blocos_cons = consolidar_blocos(aba["blocos_mpd"])
        n_b = len(aba["blocos_mpd"])
        n_c = len(aba["comentarios"])
        total_blocos += n_b
        total_coments += n_c

        vals = [
            op["ordem"], op["aba_nome_curto"], op["aba_nome"],
            len(aba["seqops"]), n_b, n_c,
            len(aba["temas"]), len(aba["pocos"]),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws_resumo.cell(row=r, column=ci, value=val)
            c.border = THIN_BORDER
            c.alignment = center if ci <= 1 else wrap
            if ci == 3:
                c.alignment = wrap
            c.font = Font(size=10)
            # Zebra
            if r % 2 == 0:
                c.fill = PatternFill(start_color=COR_CINZA_CLARO, end_color=COR_CINZA_CLARO, fill_type="solid")
        r += 1

    # Linha total
    totais = ["", "", "TOTAL", sum(len(a["seqops"]) for a in abas.values()),
              total_blocos, total_coments, "", ""]
    for ci, val in enumerate(totais, 1):
        c = ws_resumo.cell(row=r, column=ci, value=val)
        c.border = THIN_BORDER
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill(start_color=COR_AZUL_CLARO, end_color=COR_AZUL_CLARO, fill_type="solid")

    # ── Abas por Operação ──────────────────────────────────────────────
    tab_colors = [
        "1ABC9C", "2E86AB", "A569BD", "E67E22", "3498DB",
        "E74C3C", "16A085", "8E44AD", "D35400", "C0392B", "2ECC71",
    ]

    for idx, op in enumerate(FLUXO_OPERACOES):
        aba = abas[op["aba_id"]]
        aba_short = re.sub(r'[/\\*?\[\]:]', '-', op["aba_nome_curto"][:20])
        ws = wb.create_sheet(title=f"{op['ordem']:02d}-{aba_short}")
        ws.sheet_properties.tabColor = tab_colors[idx % len(tab_colors)]

        blocos_cons = consolidar_blocos(aba["blocos_mpd"])
        temas_top = aba["temas"].most_common(20)
        comentarios = aba["comentarios"]
        comentarios_dir = [c for c in comentarios if c.get("rel") == "DIRETA"]
        comentarios_ind = [c for c in comentarios if c.get("rel") == "INDIRETA"]

        # ─── Cabeçalho da aba ───────────────────────────────────
        ws.merge_cells("A1:H1")
        t = ws.cell(row=1, column=1,
            value=f"{op['ordem']}. {op['aba_nome']}")
        t.font = Font(bold=True, size=14, color=COR_AZUL_ESCURO)
        t.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:H2")
        ws.cell(row=2, column=1,
            value=op["descricao"]
        ).font = Font(italic=True, size=10, color="555555")

        ws.merge_cells("A3:H3")
        ws.cell(row=3, column=1,
            value=f"SEQOPs: {len(aba['seqops'])} | "
                  f"Blocos MPD: {len(aba['blocos_mpd'])} ({len(blocos_cons)} únicos) | "
                  f"Comentários: {len(comentarios)} ({len(comentarios_dir)} diretos, {len(comentarios_ind)} indiretos) | "
                  f"Poços: {len(aba['pocos'])}"
        ).font = Font(size=9, color="888888")

        row = 5

        # ─── SEÇÃO 1: Blocos MPD (checklist principal) ──────────
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value="BLOCOS MPD — PONTOS DE VERIFICAÇÃO")
        _aplicar_estilo(c, _estilo_subhdr())
        for ci in range(2, 9):
            _aplicar_estilo(ws.cell(row=row, column=ci), _estilo_subhdr())
        row += 1

        # Cabeçalhos da tabela de blocos
        cols_blocos = [
            ("✓", 4), ("Bloco MPD", 45), ("Freq", 7), ("Rel.", 10),
            ("Poços", 7), ("Motivo / Justificativa", 45), ("Exemplos de Itens", 50), ("Obs", 15),
        ]
        for ci, (nome, larg) in enumerate(cols_blocos, 1):
            c = ws.cell(row=row, column=ci, value=nome)
            _aplicar_estilo(c, _estilo_hdr())
            ws.column_dimensions[get_column_letter(ci)].width = max(
                ws.column_dimensions[get_column_letter(ci)].width or 0, larg)
        row += 1

        for bi, bloco in enumerate(blocos_cons):
            vals = [
                "☐",
                bloco["nome"],
                bloco["frequencia"],
                bloco["relevancia"],
                bloco["n_pocos"],
                "; ".join(bloco["motivos"][:2]),
                "\n".join(it[:120] for it in bloco["itens_exemplo"][:2]),
                "",
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.border = THIN_BORDER
                c.alignment = wrap if ci in (2, 6, 7) else center
                c.font = Font(size=10 if ci == 2 else 9)
                if ci == 1:
                    c.font = Font(size=14)
                # Cor por relevância
                if ci == 4:
                    c.fill = _cor_relevancia(bloco["relevancia"])
                    c.font = Font(size=9, bold=True)
                # Zebra
                elif bi % 2 == 1:
                    c.fill = PatternFill(start_color=COR_CINZA_CLARO, end_color=COR_CINZA_CLARO, fill_type="solid")
            ws.row_dimensions[row].height = max(30, min(80, 15 * len(bloco["itens_exemplo"])))
            row += 1

        row += 1

        # ─── SEÇÃO 2: Comentários dos Revisores ─────────────────
        if comentarios:
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row=row, column=1,
                value=f"COMENTÁRIOS DOS REVISORES CSD-MPD ({len(comentarios)})")
            _aplicar_estilo(c, _estilo_subhdr())
            for ci in range(2, 9):
                _aplicar_estilo(ws.cell(row=row, column=ci), _estilo_subhdr())
            row += 1

            cols_coment = [
                ("#", 4), ("Poço", 18), ("Rel.", 10), ("Bloco Referenciado", 30),
                ("Resumo", 50), ("Ponto de Verificação", 50), ("Temas", 25), ("SEQOP", 25),
            ]
            for ci, (nome, larg) in enumerate(cols_coment, 1):
                c = ws.cell(row=row, column=ci, value=nome)
                _aplicar_estilo(c, _estilo_hdr())
                ws.column_dimensions[get_column_letter(ci)].width = max(
                    ws.column_dimensions[get_column_letter(ci)].width or 0, larg)
            row += 1

            # Primeiro DIRETA, depois INDIRETA
            coments_ordenados = sorted(comentarios,
                key=lambda c: (0 if c.get("rel") == "DIRETA" else 1,
                               c.get("poco", "")))

            for ci_idx, com in enumerate(coments_ordenados):
                vals = [
                    ci_idx + 1,
                    com.get("poco", ""),
                    com.get("rel", ""),
                    com.get("bloco", ""),
                    com.get("resumo", ""),
                    com.get("check", ""),
                    ", ".join(com.get("temas", [])[:4]),
                    com.get("titulo_seqop", "")[:40],
                ]
                for ci, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=ci, value=val)
                    c.border = THIN_BORDER
                    c.alignment = wrap if ci >= 4 else center
                    c.font = Font(size=9)
                    # Cor por relevância
                    if ci == 3:
                        if val == "DIRETA":
                            c.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                        else:
                            c.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                    elif ci_idx % 2 == 1:
                        c.fill = PatternFill(start_color=COR_CINZA_CLARO, end_color=COR_CINZA_CLARO, fill_type="solid")
                ws.row_dimensions[row].height = 40
                row += 1

            row += 1

        # ─── SEÇÃO 3: Temas Recorrentes ──────────────────────────
        if temas_top:
            ws.merge_cells(f"A{row}:D{row}")
            c = ws.cell(row=row, column=1, value="TEMAS RECORRENTES")
            _aplicar_estilo(c, _estilo_subhdr())
            for ci in range(2, 5):
                _aplicar_estilo(ws.cell(row=row, column=ci), _estilo_subhdr())
            row += 1

            for ci, (nome, larg) in enumerate([("#", 4), ("Tema", 40), ("Frequência", 12), ("Barra", 30)], 1):
                c = ws.cell(row=row, column=ci, value=nome)
                _aplicar_estilo(c, _estilo_hdr())
            row += 1

            max_freq = temas_top[0][1] if temas_top else 1
            for ti, (tema, freq) in enumerate(temas_top):
                barra = "█" * int(20 * freq / max_freq) + f" {freq}"
                vals = [ti + 1, tema.replace("_", " ").title(), freq, barra]
                for ci, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=ci, value=val)
                    c.border = THIN_BORDER
                    c.font = Font(size=9)
                    c.alignment = wrap if ci == 2 else center
                row += 1

    # Salvar
    wb.save(SAIDA_XLSX)
    log.info(f"  Excel salvo: {SAIDA_XLSX}")
    log.info(f"    {len(FLUXO_OPERACOES) + 1} abas geradas")


# ==========================================================================
#  GERAR HTML
# ==========================================================================

def gerar_html(abas: dict):
    """Gera HTML com tabs navegáveis."""
    log.info("Gerando HTML...")

    # Prep dados
    tabs_html = []
    for op in FLUXO_OPERACOES:
        aba = abas[op["aba_id"]]
        blocos_cons = consolidar_blocos(aba["blocos_mpd"])
        temas_top = aba["temas"].most_common(15)
        comentarios = aba["comentarios"]
        comentarios_dir = [c for c in comentarios if c.get("rel") == "DIRETA"]
        comentarios_ind = [c for c in comentarios if c.get("rel") == "INDIRETA"]

        tabs_html.append({
            "id": op["aba_id"],
            "ordem": op["ordem"],
            "nome": op["aba_nome"],
            "nome_curto": op["aba_nome_curto"],
            "descricao": op["descricao"],
            "n_seqops": len(aba["seqops"]),
            "n_blocos": len(aba["blocos_mpd"]),
            "n_blocos_unicos": len(blocos_cons),
            "n_coments": len(comentarios),
            "n_dir": len(comentarios_dir),
            "n_ind": len(comentarios_ind),
            "n_pocos": len(aba["pocos"]),
            "blocos": blocos_cons,
            "comentarios": sorted(comentarios,
                key=lambda c: (0 if c.get("rel") == "DIRETA" else 1)),
            "temas": temas_top,
        })

    total_blocos = sum(t["n_blocos"] for t in tabs_html)
    total_coments = sum(t["n_coments"] for t in tabs_html)
    total_seqops = sum(t["n_seqops"] for t in tabs_html)

    # Gerar blocos section HTML
    def _blocos_html(tab):
        if not tab["blocos"]:
            return '<p class="empty">Nenhum bloco MPD identificado nesta operação.</p>'
        rows = []
        for i, b in enumerate(tab["blocos"]):
            rel_class = b["relevancia"].lower()
            motivo_html = "<br>".join(_esc(m) for m in b["motivos"][:2])
            itens_html = "<br>".join(f"• {_esc(it[:150])}" for it in b["itens_exemplo"][:2])
            rows.append(f"""
                <tr class="{'zebra' if i % 2 else ''}">
                    <td class="check">☐</td>
                    <td class="nome">{_esc(b['nome'])}</td>
                    <td class="center">{b['frequencia']}</td>
                    <td class="rel {rel_class}">{b['relevancia']}</td>
                    <td class="center">{b['n_pocos']}</td>
                    <td class="motivo">{motivo_html}</td>
                    <td class="itens">{itens_html}</td>
                </tr>""")
        return f"""
            <table class="blocos">
                <thead><tr>
                    <th width="30">✓</th><th>Bloco MPD</th><th width="50">Freq</th>
                    <th width="70">Rel.</th><th width="50">Poços</th>
                    <th>Motivo</th><th>Exemplos de Itens</th>
                </tr></thead>
                <tbody>{''.join(rows)}</tbody>
            </table>"""

    def _coments_html(tab):
        if not tab["comentarios"]:
            return ""
        rows = []
        for i, c in enumerate(tab["comentarios"]):
            rel = c.get("rel", "")
            rel_class = "direta" if rel == "DIRETA" else "indireta"
            temas = ", ".join(c.get("temas", [])[:3])
            rows.append(f"""
                <tr class="{'zebra' if i % 2 else ''}">
                    <td class="center">{i+1}</td>
                    <td>{_esc(c.get('poco',''))}</td>
                    <td class="rel {rel_class}">{rel}</td>
                    <td>{_esc(c.get('bloco',''))}</td>
                    <td>{_esc(c.get('resumo',''))}</td>
                    <td class="check-ponto">{_esc(c.get('check',''))}</td>
                    <td class="temas">{_esc(temas)}</td>
                </tr>""")
        return f"""
            <h3>Comentários dos Revisores CSD-MPD ({tab['n_coments']})</h3>
            <table class="comentarios">
                <thead><tr>
                    <th width="30">#</th><th width="100">Poço</th><th width="70">Rel.</th>
                    <th>Bloco Ref.</th><th>Resumo</th>
                    <th>Ponto de Verificação</th><th width="150">Temas</th>
                </tr></thead>
                <tbody>{''.join(rows)}</tbody>
            </table>"""

    def _temas_html(tab):
        if not tab["temas"]:
            return ""
        max_freq = tab["temas"][0][1]
        rows = []
        for i, (tema, freq) in enumerate(tab["temas"]):
            pct = int(100 * freq / max_freq)
            rows.append(f"""
                <tr class="{'zebra' if i % 2 else ''}">
                    <td>{_esc(tema.replace('_',' ').title())}</td>
                    <td class="center">{freq}</td>
                    <td><div class="bar" style="width:{pct}%">{freq}</div></td>
                </tr>""")
        return f"""
            <h3>Temas Recorrentes</h3>
            <table class="temas">
                <thead><tr><th>Tema</th><th width="60">Freq</th><th>Distribuição</th></tr></thead>
                <tbody>{''.join(rows)}</tbody>
            </table>"""

    def _esc(text):
        if not text:
            return ""
        return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    # Tabs nav
    nav_items = []
    tab_contents = []
    for i, tab in enumerate(tabs_html):
        active = " active" if i == 0 else ""
        badge_class = "badge-blue" if tab["n_coments"] > 0 else "badge-gray"
        nav_items.append(
            f'<button class="tab-btn{active}" data-tab="{tab["id"]}">'
            f'<span class="tab-num">{tab["ordem"]}</span> {tab["nome_curto"]}'
            f'<span class="badge {badge_class}">{tab["n_coments"]}</span></button>'
        )

        tab_contents.append(f"""
        <div class="tab-content{active}" id="tab-{tab['id']}">
            <div class="tab-header">
                <h2>{tab['ordem']}. {tab['nome']}</h2>
                <p class="descricao">{_esc(tab['descricao'])}</p>
                <div class="stats-bar">
                    <span class="stat"><b>{tab['n_seqops']}</b> SEQOPs</span>
                    <span class="stat"><b>{tab['n_blocos']}</b> blocos ({tab['n_blocos_unicos']} únicos)</span>
                    <span class="stat"><b>{tab['n_dir']}</b> diretos</span>
                    <span class="stat"><b>{tab['n_ind']}</b> indiretos</span>
                    <span class="stat"><b>{tab['n_pocos']}</b> poços</span>
                </div>
            </div>
            <h3>Blocos MPD — Pontos de Verificação ({tab['n_blocos_unicos']})</h3>
            {_blocos_html(tab)}
            {_coments_html(tab)}
            {_temas_html(tab)}
        </div>""")

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Checklist Operacional MPD — Fluxo por Abas</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: 'Segoe UI', system-ui, sans-serif; background: #f0f2f5; color: #333; }}
.container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}

/* Header */
.header {{ background: linear-gradient(135deg, {COR_AZUL_ESCURO} 0%, {COR_AZUL_MEDIO} 100%);
           color: white; padding: 25px 30px; border-radius: 12px 12px 0 0; }}
.header h1 {{ font-size: 1.6em; margin-bottom: 8px; }}
.header .meta {{ font-size: 0.85em; opacity: 0.8; }}
.header .summary {{ display: flex; gap: 20px; margin-top: 12px; flex-wrap: wrap; }}
.header .summary .item {{ background: rgba(255,255,255,0.15); padding: 8px 16px;
                          border-radius: 8px; font-size: 0.9em; }}
.header .summary .item b {{ font-size: 1.3em; display: block; }}

/* Tabs */
.tabs-nav {{ display: flex; gap: 2px; background: #e0e3e8; padding: 4px 4px 0;
             border-radius: 0; overflow-x: auto; flex-wrap: wrap; }}
.tab-btn {{ border: none; background: #d5d8de; color: #555; padding: 10px 14px;
            cursor: pointer; font-size: 0.82em; border-radius: 8px 8px 0 0;
            display: flex; align-items: center; gap: 6px; transition: all 0.2s;
            white-space: nowrap; }}
.tab-btn:hover {{ background: #c5c9d0; }}
.tab-btn.active {{ background: white; color: {COR_AZUL_ESCURO}; font-weight: 600;
                   box-shadow: 0 -2px 6px rgba(0,0,0,0.08); }}
.tab-num {{ background: {COR_AZUL_ESCURO}; color: white; border-radius: 50%;
            width: 22px; height: 22px; display: inline-flex; align-items: center;
            justify-content: center; font-size: 0.75em; font-weight: 700; }}
.tab-btn.active .tab-num {{ background: {COR_AZUL_MEDIO}; }}
.badge {{ font-size: 0.7em; padding: 2px 6px; border-radius: 10px; font-weight: 600; }}
.badge-blue {{ background: #3498db; color: white; }}
.badge-gray {{ background: #bdc3c7; color: white; }}

/* Tab content */
.tabs-body {{ background: white; border-radius: 0 0 12px 12px; padding: 0;
              box-shadow: 0 4px 12px rgba(0,0,0,0.08); }}
.tab-content {{ display: none; padding: 25px 30px; }}
.tab-content.active {{ display: block; }}
.tab-header {{ margin-bottom: 20px; }}
.tab-header h2 {{ color: {COR_AZUL_ESCURO}; font-size: 1.4em; margin-bottom: 6px; }}
.tab-header .descricao {{ color: #666; font-size: 0.95em; margin-bottom: 10px; }}
.stats-bar {{ display: flex; gap: 15px; flex-wrap: wrap; }}
.stats-bar .stat {{ background: #f5f6fa; padding: 6px 14px; border-radius: 6px;
                    font-size: 0.85em; color: #555; }}
.stats-bar .stat b {{ color: {COR_AZUL_ESCURO}; }}

h3 {{ color: {COR_AZUL_MEDIO}; font-size: 1.1em; margin: 20px 0 10px;
      border-bottom: 2px solid {COR_AZUL_CLARO}; padding-bottom: 6px; }}

/* Tables */
table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 0.88em; }}
thead th {{ background: {COR_AZUL_ESCURO}; color: white; padding: 10px 8px;
            text-align: left; font-weight: 600; position: sticky; top: 0; }}
tbody td {{ padding: 8px; border-bottom: 1px solid #eee; vertical-align: top; }}
tr.zebra td {{ background: #fafbfc; }}
tbody tr:hover td {{ background: #e8f4fd !important; }}

.check {{ font-size: 1.3em; text-align: center; cursor: pointer; }}
.nome {{ font-weight: 500; }}
.center {{ text-align: center; }}
.rel {{ font-weight: 600; text-align: center; border-radius: 4px; }}
.rel.alta {{ background: #fadbd8; color: #c0392b; }}
.rel.media {{ background: #fef9e7; color: #b7950b; }}
.rel.baixa {{ background: #eafaf1; color: #27ae60; }}
.rel.direta {{ background: #fadbd8; color: #c0392b; }}
.rel.indireta {{ background: #fef9e7; color: #b7950b; }}
.motivo, .itens {{ font-size: 0.9em; color: #555; }}
.check-ponto {{ font-weight: 500; color: {COR_AZUL_MEDIO}; }}
.temas {{ font-size: 0.85em; color: #777; }}
.empty {{ color: #999; font-style: italic; padding: 20px; }}

/* Bars */
.bar {{ background: linear-gradient(90deg, #3498db, #2ecc71); color: white;
        padding: 3px 8px; border-radius: 4px; font-size: 0.8em; min-width: 20px;
        text-align: right; }}

/* Interactive checks */
.check.checked {{ color: #27ae60; }}
.check.checked::after {{ content: "✓"; }}

/* Print */
@media print {{
    .tabs-nav {{ display: none; }}
    .tab-content {{ display: block !important; page-break-before: always; }}
    .header {{ background: white; color: black; border: 2px solid #333; }}
    .tab-num {{ background: #333; }}
}}
</style>
</head>
<body>
<div class="container">
    <div class="header">
        <h1>Checklist Operacional MPD — Fluxo Cronológico por Abas</h1>
        <div class="meta">
            Gerado: {datetime.now().strftime('%Y-%m-%d %H:%M')} |
            Modelo: gpt-5.1 |
            Fonte: {METADADOS['total_seqops']} SEQOPs de {METADADOS['total_pocos']} poços
        </div>
        <div class="summary">
            <div class="item"><b>{total_seqops}</b>SEQOPs</div>
            <div class="item"><b>{total_blocos}</b>Blocos MPD</div>
            <div class="item"><b>{total_coments}</b>Comentários</div>
            <div class="item"><b>{len(FLUXO_OPERACOES)}</b>Operações</div>
            <div class="item"><b>{METADADOS['total_pocos']}</b>Poços</div>
        </div>
    </div>

    <div class="tabs-nav">
        {''.join(nav_items)}
    </div>

    <div class="tabs-body">
        {''.join(tab_contents)}
    </div>
</div>

<script>
document.querySelectorAll('.tab-btn').forEach(btn => {{
    btn.addEventListener('click', () => {{
        document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
        document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
        btn.classList.add('active');
        document.getElementById('tab-' + btn.dataset.tab).classList.add('active');
    }});
}});
document.querySelectorAll('.check').forEach(cell => {{
    cell.addEventListener('click', () => {{
        cell.classList.toggle('checked');
        cell.textContent = cell.classList.contains('checked') ? '✓' : '☐';
    }});
}});
</script>
</body>
</html>"""

    with open(SAIDA_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    log.info(f"  HTML salvo: {SAIDA_HTML}")


# ==========================================================================
#  MAIN
# ==========================================================================

def main():
    log.info("=" * 70)
    log.info("GERADOR DE CHECKLIST POR ABAS — FLUXO OPERACIONAL MPD")
    log.info("=" * 70)

    abas = carregar_dados()
    gerar_excel(abas)
    gerar_html(abas)

    log.info("=" * 70)
    log.info("CONCLUÍDO!")
    log.info(f"  Excel: {SAIDA_XLSX}")
    log.info(f"  HTML:  {SAIDA_HTML}")
    log.info("=" * 70)


if __name__ == "__main__":
    main()

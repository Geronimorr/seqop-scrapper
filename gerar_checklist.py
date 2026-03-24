"""
Gerador de Checklist para Revisão de SEQOPs (CSD-MPD)
=====================================================
Analisa comentários MPD coletados pelo scraper e usa IA generativa
para produzir um checklist estruturado de revisão.

Fases:
  1. Enriquecer dados: visitar URLs e coletar comentários + conteúdo SEQOP
  2. Analisar com IA: enviar ao claude-sonnet-4-6 para identificar padrões
  3. Gerar checklist: produzir Excel + relatório estruturado

Uso:
    C:\\SharedPython\\venv\\Scripts\\python.exe gerar_checklist.py [--skip-enrich]

Pré-requisitos:
    - resultado_aprovacoes_mpd.xlsx (saída do scraper principal)
    - Acesso à API de IA Petrobras
    - Edge + Selenium (para fase de enriquecimento)
"""

# ── Re-launch com o venv compartilhado ──────────────────────────────────────
import subprocess
import sys
from pathlib import Path

_SHARED_PYTHON = Path(r"C:\SharedPython\venv\Scripts\python.exe")
if _SHARED_PYTHON.exists() and Path(sys.executable).resolve() != _SHARED_PYTHON.resolve():
    print(f"[LAUNCH] Relançando com {_SHARED_PYTHON}")
    result = subprocess.run([str(_SHARED_PYTHON)] + sys.argv, check=False)
    sys.exit(result.returncode)

# ── Imports ─────────────────────────────────────────────────────────────────
import json
import logging
import os
import re
import time
import argparse
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from textwrap import dedent

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Configuração ────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent

# Arquivos de entrada/saída
ENTRADA_XLSX = SCRIPT_DIR / "resultado_aprovacoes_mpd.xlsx"
COMENTARIOS_JSON = SCRIPT_DIR / "comentarios_mpd.json"
CHECKLIST_XLSX = SCRIPT_DIR / "checklist_revisao_mpd.xlsx"
CHECKLIST_HTML = SCRIPT_DIR / "checklist_revisao_mpd.html"

# API de IA
API_BASE = "https://apid.petrobras.com.br/ia/generativos/v1/litellm-chat-petrobras/litellm"
API_KEY = "b320f1f58c9743e9a74048ce64717c89"
MODELO_IA = "claude-sonnet-4-6"

# Selenium
DEBUGGING_PORT = 9222
WAIT_TIMEOUT = 15

# ── Logging ─────────────────────────────────────────────────────────────────

LOG_FILE = SCRIPT_DIR / "checklist.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("checklist")

# Suprimir warnings SSL
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ==========================================================================
#  FASE 1: ENRIQUECER – Visitar URLs e coletar comentários + conteúdo
# ==========================================================================

def _criar_driver():
    """Cria ou conecta ao Edge."""
    from selenium import webdriver
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.common.exceptions import WebDriverException

    try:
        log.info(f"Conectando ao Edge na porta {DEBUGGING_PORT}...")
        opts = EdgeOptions()
        opts.debugger_address = f"127.0.0.1:{DEBUGGING_PORT}"
        d = webdriver.Edge(options=opts)
        log.info("Conectado ao Edge existente.")
        return d
    except WebDriverException:
        log.info("Abrindo nova instância do Edge...")

    opts = EdgeOptions()
    profile = Path(os.environ.get("LOCALAPPDATA", "")) / "SeqopScraper_Edge"
    profile.mkdir(parents=True, exist_ok=True)
    opts.add_argument(f"--user-data-dir={profile}")
    opts.add_argument("--profile-directory=Default")
    opts.add_argument("--start-maximized")
    opts.add_argument(f"--remote-debugging-port={DEBUGGING_PORT}")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    d = webdriver.Edge(options=opts)
    log.info("Edge aberto (perfil dedicado).")
    return d


def _extrair_comentarios_pagina(driver) -> list[dict]:
    """Extrai TODOS os comentários da página de detalhe da SEQOP (não só MPD)."""
    from selenium.webdriver.common.by import By
    from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException

    comentarios = []

    try:
        grupos = driver.find_elements(By.CSS_SELECTOR, "div.grupoComentario")
        for grupo in grupos:
            try:
                # Texto do comentário principal
                texto = ""
                try:
                    card = grupo.find_element(By.CSS_SELECTOR, "div.card-text")
                    texto = card.text.strip()
                except Exception:
                    pass

                # Autor e tipo (CSD-MPD, CSD-SMAB, etc.)
                autor = ""
                tipo_csd = ""
                versao = ""
                try:
                    autor_div = grupo.find_element(
                        By.CSS_SELECTOR, "div.emLinhaAutor.autor-versao"
                    )
                    spans = autor_div.find_elements(By.TAG_NAME, "span")
                    if len(spans) >= 1:
                        nome_raw = spans[0].text.strip()
                        autor = re.sub(
                            r"\s+h[aá]\s+\d+\s+\w+$", "", nome_raw,
                            flags=re.IGNORECASE,
                        ).strip()
                    if len(spans) >= 2:
                        info = spans[1].text.strip()  # "CSD-MPD - Versão: 2"
                        m_tipo = re.match(r"(CSD-\w+|Fiscal)\s*-\s*Vers[aã]o:\s*(\d+)", info, re.IGNORECASE)
                        if m_tipo:
                            tipo_csd = m_tipo.group(1)
                            versao = m_tipo.group(2)
                        else:
                            tipo_csd = info
                except Exception:
                    pass

                # Respostas
                respostas = []
                try:
                    replies = grupo.find_elements(By.CSS_SELECTOR, "div.replyComments")
                    for reply in replies:
                        try:
                            reply_text = ""
                            try:
                                rc = reply.find_element(By.CSS_SELECTOR, "div.card-text")
                                reply_text = rc.text.strip()
                            except Exception:
                                pass
                            reply_autor = ""
                            try:
                                ra = reply.find_element(By.CSS_SELECTOR, "div.emLinhaAutor.autor-versao")
                                rs = ra.find_elements(By.TAG_NAME, "span")
                                if rs:
                                    reply_autor = re.sub(
                                        r"\s+h[aá]\s+\d+\s+\w+$", "",
                                        rs[0].text.strip(), flags=re.IGNORECASE,
                                    ).strip()
                            except Exception:
                                pass
                            if reply_text:
                                respostas.append({"autor": reply_autor, "texto": reply_text})
                        except Exception:
                            continue
                except Exception:
                    pass

                if texto or respostas:
                    comentarios.append({
                        "autor": autor,
                        "tipo_csd": tipo_csd,
                        "versao": versao,
                        "texto": texto,
                        "respostas": respostas,
                    })

            except (NoSuchElementException, StaleElementReferenceException):
                continue

    except Exception as e:
        log.warning(f"  Erro ao extrair comentários: {e}")

    return comentarios


def _extrair_conteudo_seqop(driver) -> str:
    """Extrai o conteúdo da SEQOP (itens operacionais) da página de detalhe.

    Tenta expandir 'Detalhes da versão' e capturar a tabela de itens.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    conteudo = ""

    try:
        # Tentar clicar em "Detalhes da versão" para expandir
        try:
            btn_detalhes = driver.find_element(
                By.XPATH,
                "//span[contains(text(),'Detalhes da vers')]/.."
            )
            driver.execute_script("arguments[0].click();", btn_detalhes)
            time.sleep(2)
        except Exception:
            pass

        # Capturar conteúdo da área principal (col-9)
        try:
            main_area = driver.find_element(By.CSS_SELECTOR, "div.col-9")
            conteudo = main_area.text.strip()
        except Exception:
            pass

        # Se não encontrou, tenta dadosVersao
        if not conteudo:
            try:
                dados = driver.find_element(By.CSS_SELECTOR, "div.dadosVersao")
                conteudo = dados.text.strip()
            except Exception:
                pass

        # Se não encontrou, tenta todo o content
        if not conteudo:
            try:
                content = driver.find_element(By.CSS_SELECTOR, "div.content")
                conteudo = content.text.strip()
            except Exception:
                pass

    except Exception as e:
        log.warning(f"  Erro ao extrair conteúdo SEQOP: {e}")

    return conteudo


def fase1_enriquecer(urls_unicas: list[dict], max_urls: int = 0) -> list[dict]:
    """Visita URLs da SEQOP e coleta comentários + conteúdo.

    Args:
        urls_unicas: lista de dicts com {url, poco, titulo, seq_id}
        max_urls: limite de URLs a visitar (0 = todas)

    Returns:
        lista de dicts com comentários e conteúdo enriquecido
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    if max_urls > 0:
        urls_unicas = urls_unicas[:max_urls]

    log.info(f"FASE 1: Enriquecendo {len(urls_unicas)} URLs")

    driver = _criar_driver()

    # Aguardar login
    driver.get("https://csdpocos.petrobras.com.br/seqop")
    time.sleep(3)

    # Verificar login
    try:
        menus = driver.find_elements(By.CSS_SELECTOR, "aside.menu")
        if not menus:
            log.info("Aguardando login manual...")
            for _ in range(120):
                time.sleep(5)
                menus = driver.find_elements(By.CSS_SELECTOR, "aside.menu")
                if menus:
                    break
    except Exception:
        pass

    time.sleep(2)
    dados_enriquecidos = []

    for idx, info in enumerate(urls_unicas, 1):
        url = info["url"]
        log.info(f"  [{idx}/{len(urls_unicas)}] {info['poco']} – {info['titulo'][:50]}")
        log.info(f"    URL: {url}")

        try:
            driver.get(url)
            time.sleep(3)

            # Aguardar página carregar
            try:
                WebDriverWait(driver, 10).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
            except Exception:
                pass
            time.sleep(1)

            # Extrair dados
            comentarios = _extrair_comentarios_pagina(driver)
            conteudo = _extrair_conteudo_seqop(driver)

            # Título da página
            titulo_pagina = info["titulo"]
            try:
                h5s = driver.find_elements(By.CSS_SELECTOR, "div.SideBar h5")
                if len(h5s) >= 2:
                    titulo_pagina = h5s[1].text.strip() or titulo_pagina
            except Exception:
                pass

            dados = {
                "url": url,
                "poco": info["poco"],
                "titulo": titulo_pagina,
                "seq_id": info["seq_id"],
                "comentarios": comentarios,
                "conteudo_seqop": conteudo[:10000],  # limitar tamanho
                "comentarios_mpd": [c for c in comentarios if "MPD" in c.get("tipo_csd", "").upper()],
                "total_comentarios": len(comentarios),
                "total_mpd": sum(1 for c in comentarios if "MPD" in c.get("tipo_csd", "").upper()),
            }
            dados_enriquecidos.append(dados)

            log.info(f"    → {len(comentarios)} comentários ({dados['total_mpd']} MPD), "
                     f"conteúdo: {len(conteudo)} chars")

        except Exception as e:
            log.error(f"    Erro: {e}")
            dados_enriquecidos.append({
                "url": url, "poco": info["poco"], "titulo": info["titulo"],
                "seq_id": info["seq_id"], "comentarios": [], "conteudo_seqop": "",
                "comentarios_mpd": [], "total_comentarios": 0, "total_mpd": 0,
                "erro": str(e),
            })

        # Salvar progresso a cada 10
        if idx % 10 == 0:
            _salvar_json(dados_enriquecidos, COMENTARIOS_JSON)
            log.info(f"  Progresso salvo ({idx}/{len(urls_unicas)})")

    _salvar_json(dados_enriquecidos, COMENTARIOS_JSON)
    log.info(f"FASE 1 completa: {len(dados_enriquecidos)} SEQOPs enriquecidas → {COMENTARIOS_JSON}")

    return dados_enriquecidos


def _salvar_json(dados, caminho):
    """Salva dados em JSON com formatação."""
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


# ==========================================================================
#  FASE 2: ANALISAR – Enviar ao claude-sonnet-4-6 para análise
# ==========================================================================

MODELOS_FALLBACK = [
    MODELO_IA,
    "gpt-4o",
    "claude-sonnet-4-5",
    "gpt-4.1",
    "claude-3-7-sonnet",
    "gpt-5-mini",
]


def _chamar_ia(mensagens: list[dict], max_tokens: int = 4096, temperature: float = 0.3) -> str:
    """Chama a API de IA generativa e retorna o texto da resposta.
    Tenta múltiplos modelos em caso de falha."""
    headers = {
        "Content-Type": "application/json",
        "api-key": API_KEY,
    }
    payload = {
        "messages": mensagens,
        "max_tokens": max_tokens,
        "temperature": temperature,
    }

    for modelo in MODELOS_FALLBACK:
        url = f"{API_BASE}/engines/{modelo}/chat/completions"
        log.info(f"    Chamando modelo: {modelo}")

        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=180, verify=False)

            if resp.status_code != 200:
                log.warning(f"    Modelo {modelo}: HTTP {resp.status_code} – {resp.text[:200]}")
                continue

            data = resp.json()
            choices = data.get("choices", [])
            if not choices:
                log.warning(f"    Modelo {modelo}: Resposta sem choices")
                continue

            content = choices[0].get("message", {}).get("content", "")
            usage = data.get("usage", {})
            log.info(f"    ✓ {modelo} – tokens: prompt={usage.get('prompt_tokens')}, "
                     f"completion={usage.get('completion_tokens')}")
            return content

        except requests.exceptions.Timeout:
            log.warning(f"    Modelo {modelo}: Timeout")
        except Exception as e:
            log.warning(f"    Modelo {modelo}: Erro – {e}")

    raise RuntimeError(f"Todos os modelos falharam: {MODELOS_FALLBACK}")


def fase2_analisar(dados_enriquecidos: list[dict]) -> dict:
    """Analisa os comentários MPD com IA e gera dados para o checklist.

    Etapas:
      1. Agregar todos os comentários MPD
      2. Enviar para IA em batches para identificar categorias/padrões
      3. Enviar para IA para gerar o checklist final
    """
    log.info("=" * 60)
    log.info("FASE 2: Análise com IA (claude-sonnet-4-6)")
    log.info("=" * 60)

    # ── Agregar comentários MPD ──────────────────────────────────────
    comentarios_mpd = []
    for seq in dados_enriquecidos:
        for c in seq.get("comentarios_mpd", []):
            if c.get("texto"):
                texto_completo = c["texto"]
                if c.get("respostas"):
                    for r in c["respostas"]:
                        texto_completo += f"\n[Resposta de {r.get('autor', '?')}]: {r['texto']}"
                comentarios_mpd.append({
                    "poco": seq["poco"],
                    "titulo": seq["titulo"],
                    "versao": c.get("versao", ""),
                    "autor": c.get("autor", ""),
                    "texto": texto_completo,
                })

    # Adicionar também comentários de outros CSDs que mencionam MPD
    for seq in dados_enriquecidos:
        for c in seq.get("comentarios", []):
            if "MPD" not in c.get("tipo_csd", "").upper():
                # Verificar se o texto menciona MPD
                texto = c.get("texto", "").upper()
                if "MPD" in texto or "SBP" in texto or "BEARING" in texto:
                    texto_completo = c.get("texto", "")
                    if c.get("respostas"):
                        for r in c["respostas"]:
                            texto_completo += f"\n[Resposta de {r.get('autor', '?')}]: {r['texto']}"
                    comentarios_mpd.append({
                        "poco": seq["poco"],
                        "titulo": seq["titulo"],
                        "versao": c.get("versao", ""),
                        "autor": c.get("autor", ""),
                        "tipo_csd": c.get("tipo_csd", ""),
                        "texto": texto_completo,
                    })

    log.info(f"  Total de comentários para análise: {len(comentarios_mpd)}")

    if not comentarios_mpd:
        log.warning("  Nenhum comentário MPD encontrado!")
        return {"categorias": [], "itens_checklist": [], "analise_raw": ""}

    # ── Etapa 1: Identificar categorias e padrões ────────────────────
    log.info("  Etapa 1: Identificando categorias e padrões...")

    # Montar texto dos comentários (limitando tamanho para o prompt)
    textos_para_analise = []
    chars_total = 0
    for c in comentarios_mpd:
        bloco = f"[{c['poco']} | {c['titulo']} | v{c['versao']} | {c.get('autor','')}]\n{c['texto']}\n---"
        if chars_total + len(bloco) > 80000:  # limite seguro para o prompt
            break
        textos_para_analise.append(bloco)
        chars_total += len(bloco)

    log.info(f"  Usando {len(textos_para_analise)} comentários ({chars_total} chars)")

    prompt_categorias = dedent("""\
    Você é um especialista em perfuração de poços de petróleo, especificamente em operações MPD (Managed Pressure Drilling).

    Abaixo estão comentários reais de revisores CSD-MPD sobre sequências operacionais (SEQOPs) de poços da Petrobras.
    Cada comentário está no formato: [poço | título da seqop | versão | autor]

    TAREFA: Analise todos os comentários e identifique:
    1. As CATEGORIAS principais de itens que os revisores MPD verificam/cobram
    2. Os PADRÕES recorrentes (itens que aparecem em múltiplas revisões)
    3. Os ERROS COMUNS que costumam ser corrigidos
    4. Os PADRÕES NORMATIVOS mencionados (ex: PE-2POC-01113)

    Responda em JSON com a seguinte estrutura:
    {
      "categorias": [
        {
          "nome": "Nome da Categoria",
          "descricao": "Descrição breve",
          "frequencia": "alta/media/baixa",
          "exemplos_comentarios": ["resumo do comentário 1", "resumo 2"]
        }
      ],
      "padroes_recorrentes": [
        {
          "padrao": "Descrição do padrão",
          "frequencia": N,
          "categoria": "categoria relacionada"
        }
      ],
      "erros_comuns": ["erro 1", "erro 2"],
      "normas_mencionadas": ["PE-2POC-01113", ...]
    }

    COMENTÁRIOS PARA ANÁLISE:
    """)

    analise_categorias = _chamar_ia([
        {"role": "system", "content": "Você é um especialista em perfuração MPD da Petrobras. Responda sempre em JSON válido, em português."},
        {"role": "user", "content": prompt_categorias + "\n".join(textos_para_analise)},
    ], max_tokens=8000, temperature=0.2)

    log.info("  Etapa 1 concluída.")

    # ── Etapa 2: Gerar checklist estruturado ─────────────────────────
    log.info("  Etapa 2: Gerando checklist estruturado...")

    prompt_checklist = dedent(f"""\
    Você é um especialista em perfuração MPD na Petrobras.

    Com base na análise de categorias abaixo (extraída de {len(comentarios_mpd)} comentários reais de revisores CSD-MPD),
    gere um CHECKLIST COMPLETO para revisão de SEQOPs que envolvam MPD.

    ANÁLISE DE CATEGORIAS:
    {analise_categorias}

    TAREFA: Gere um checklist estruturado em JSON com a seguinte estrutura:
    {{
      "titulo": "Checklist de Revisão de SEQOPs - CSD-MPD",
      "versao": "1.0",
      "data_geracao": "{datetime.now().strftime('%d/%m/%Y')}",
      "baseado_em": "{len(comentarios_mpd)} comentários de {len(set(c['poco'] for c in comentarios_mpd))} poços",
      "categorias": [
        {{
          "id": "CAT-01",
          "nome": "Nome da Categoria",
          "descricao": "Descrição da categoria",
          "prioridade": "ALTA/MEDIA/BAIXA",
          "itens": [
            {{
              "id": "CAT-01-001",
              "descricao": "Item do checklist (pergunta ou verificação)",
              "detalhes": "Detalhes e justificativa baseada nos comentários reais",
              "referencia_normativa": "Padrão normativo relacionado (se houver)",
              "exemplo_comentario": "Exemplo real de comentário de revisor que motivou este item",
              "criticidade": "CRITICA/IMPORTANTE/RECOMENDADA"
            }}
          ]
        }}
      ],
      "observacoes_gerais": ["observação 1", "observação 2"]
    }}

    REQUISITOS:
    - Cada item deve ser uma PERGUNTA ou VERIFICAÇÃO clara e objetiva
    - Inclua pelo menos 5 categorias
    - Cada categoria deve ter pelo menos 3 itens
    - Priorize itens baseados em frequência de aparição nos comentários
    - Inclua referências normativas quando mencionadas
    - A linguagem deve ser técnica e direta
    - Todos os itens devem ser diretamente derivados dos comentários reais analisados
    """)

    checklist_json = _chamar_ia([
        {"role": "system", "content": "Você é um especialista em perfuração MPD da Petrobras. Gere o checklist em JSON válido, em português."},
        {"role": "user", "content": prompt_checklist},
    ], max_tokens=16000, temperature=0.2)

    log.info("  Etapa 2 concluída.")

    # Parsear JSON
    resultado = {
        "analise_categorias_raw": analise_categorias,
        "checklist_raw": checklist_json,
    }

    try:
        # Extrair JSON da resposta (pode ter texto antes/depois)
        json_match = re.search(r'\{.*\}', analise_categorias, re.DOTALL)
        if json_match:
            resultado["analise_categorias"] = json.loads(json_match.group())

        json_match = re.search(r'\{.*\}', checklist_json, re.DOTALL)
        if json_match:
            resultado["checklist"] = json.loads(json_match.group())
    except json.JSONDecodeError as e:
        log.warning(f"  Erro ao parsear JSON da IA: {e}")

    return resultado


# ==========================================================================
#  FASE 3: GERAR – Produzir Excel e HTML do checklist
# ==========================================================================

def fase3_gerar_excel(resultado: dict):
    """Gera o checklist em formato Excel profissional."""
    checklist = resultado.get("checklist", {})

    if not checklist:
        log.warning("Checklist vazio, tentando usar raw JSON")
        try:
            json_match = re.search(r'\{.*\}', resultado.get("checklist_raw", ""), re.DOTALL)
            if json_match:
                checklist = json.loads(json_match.group())
        except Exception:
            log.error("Não foi possível parsear o checklist")
            return

    wb = Workbook()

    # ── Aba 1: Checklist ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Checklist MPD"

    # Estilos
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    cat_font = Font(bold=True, size=11, color="1F4E79")
    crit_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    imp_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    rec_fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Título
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value=checklist.get("titulo", "Checklist de Revisão MPD"))
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Metadados
    ws.merge_cells("A2:G2")
    meta = f"Gerado em: {checklist.get('data_geracao', datetime.now().strftime('%d/%m/%Y'))} | " \
           f"Baseado em: {checklist.get('baseado_em', 'N/A')} | Versão: {checklist.get('versao', '1.0')}"
    ws.cell(row=2, column=1, value=meta).font = Font(italic=True, size=9, color="666666")

    # Cabeçalhos
    colunas = [
        ("ID", 14), ("✓", 5), ("Verificação", 60), ("Detalhes / Justificativa", 55),
        ("Ref. Normativa", 22), ("Criticidade", 14), ("Observações do Revisor", 40),
    ]
    for ci, (nome, larg) in enumerate(colunas, 1):
        c = ws.cell(row=4, column=ci, value=nome)
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = larg

    row_num = 5
    categorias = checklist.get("categorias", [])

    for cat in categorias:
        # Linha de categoria
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        cat_cell = ws.cell(row=row_num, column=1,
                           value=f"{cat.get('id', '')} – {cat.get('nome', '')} [{cat.get('prioridade', '')}]")
        cat_cell.fill = cat_fill
        cat_cell.font = cat_font
        cat_cell.alignment = Alignment(vertical="center")
        for ci in range(1, 8):
            ws.cell(row=row_num, column=ci).border = border
        ws.row_dimensions[row_num].height = 22
        row_num += 1

        # Descrição da categoria
        if cat.get("descricao"):
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
            desc_cell = ws.cell(row=row_num, column=1, value=f"  {cat['descricao']}")
            desc_cell.font = Font(italic=True, size=9, color="444444")
            desc_cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # Itens
        for item in cat.get("itens", []):
            criticidade = item.get("criticidade", "").upper()

            ws.cell(row=row_num, column=1, value=item.get("id", "")).border = border
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")

            # Checkbox vazia
            check_cell = ws.cell(row=row_num, column=2, value="☐")
            check_cell.border = border
            check_cell.alignment = Alignment(horizontal="center", vertical="center")
            check_cell.font = Font(size=14)

            # Verificação
            verif_cell = ws.cell(row=row_num, column=3, value=item.get("descricao", ""))
            verif_cell.border = border
            verif_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Detalhes
            detalhes = item.get("detalhes", "")
            exemplo = item.get("exemplo_comentario", "")
            texto_detalhes = detalhes
            if exemplo:
                texto_detalhes += f"\n\nEx: \"{exemplo}\""
            det_cell = ws.cell(row=row_num, column=4, value=texto_detalhes)
            det_cell.border = border
            det_cell.alignment = Alignment(wrap_text=True, vertical="top")
            det_cell.font = Font(size=9)

            # Ref normativa
            ref_cell = ws.cell(row=row_num, column=5, value=item.get("referencia_normativa", ""))
            ref_cell.border = border
            ref_cell.alignment = Alignment(wrap_text=True, vertical="center")

            # Criticidade
            crit_cell = ws.cell(row=row_num, column=6, value=criticidade)
            crit_cell.border = border
            crit_cell.alignment = Alignment(horizontal="center", vertical="center")
            if "CRIT" in criticidade:
                crit_cell.fill = crit_fill
                crit_cell.font = Font(bold=True, color="FFFFFF")
            elif "IMPORT" in criticidade:
                crit_cell.fill = imp_fill
            elif "RECOM" in criticidade:
                crit_cell.fill = rec_fill

            # Observações (vazia para o revisor preencher)
            obs_cell = ws.cell(row=row_num, column=7, value="")
            obs_cell.border = border

            ws.row_dimensions[row_num].height = 45
            row_num += 1

    # Observações gerais
    row_num += 1
    obs_gerais = checklist.get("observacoes_gerais", [])
    if obs_gerais:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        ws.cell(row=row_num, column=1, value="OBSERVAÇÕES GERAIS").font = Font(bold=True, size=11)
        row_num += 1
        for obs in obs_gerais:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
            ws.cell(row=row_num, column=1, value=f"• {obs}").alignment = Alignment(wrap_text=True)
            row_num += 1

    ws.auto_filter.ref = f"A4:G{row_num}"
    ws.freeze_panes = "A5"

    # ── Aba 2: Análise de Categorias ─────────────────────────────────
    ws2 = wb.create_sheet("Análise Categorias")
    analise = resultado.get("analise_categorias", {})

    ws2.cell(row=1, column=1, value="Análise de Categorias - Comentários MPD").font = Font(bold=True, size=12)

    if isinstance(analise, dict):
        row2 = 3
        # Categorias
        ws2.cell(row=row2, column=1, value="CATEGORIAS IDENTIFICADAS").font = Font(bold=True)
        row2 += 1
        for cat_a in analise.get("categorias", []):
            ws2.cell(row=row2, column=1, value=cat_a.get("nome", ""))
            ws2.cell(row=row2, column=2, value=cat_a.get("descricao", ""))
            ws2.cell(row=row2, column=3, value=cat_a.get("frequencia", ""))
            row2 += 1

        row2 += 1
        ws2.cell(row=row2, column=1, value="PADRÕES RECORRENTES").font = Font(bold=True)
        row2 += 1
        for p in analise.get("padroes_recorrentes", []):
            ws2.cell(row=row2, column=1, value=p.get("padrao", ""))
            ws2.cell(row=row2, column=2, value=str(p.get("frequencia", "")))
            ws2.cell(row=row2, column=3, value=p.get("categoria", ""))
            row2 += 1

        row2 += 1
        ws2.cell(row=row2, column=1, value="ERROS COMUNS").font = Font(bold=True)
        row2 += 1
        for e in analise.get("erros_comuns", []):
            ws2.cell(row=row2, column=1, value=e)
            row2 += 1

        row2 += 1
        ws2.cell(row=row2, column=1, value="NORMAS MENCIONADAS").font = Font(bold=True)
        row2 += 1
        for n in analise.get("normas_mencionadas", []):
            ws2.cell(row=row2, column=1, value=n)
            row2 += 1

    wb.save(CHECKLIST_XLSX)
    log.info(f"Checklist Excel salvo: {CHECKLIST_XLSX}")


def fase3_gerar_html(resultado: dict):
    """Gera o checklist em formato HTML interativo."""
    checklist = resultado.get("checklist", {})
    if not checklist:
        return

    categorias = checklist.get("categorias", [])
    # Contar itens por criticidade
    total_itens = sum(len(cat.get("itens", [])) for cat in categorias)
    criticos = sum(1 for cat in categorias for i in cat.get("itens", [])
                   if "CRIT" in i.get("criticidade", "").upper())
    importantes = sum(1 for cat in categorias for i in cat.get("itens", [])
                      if "IMPORT" in i.get("criticidade", "").upper())

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <title>{checklist.get('titulo', 'Checklist MPD')}</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: 'Segoe UI', Tahoma, sans-serif; background: #f0f2f5; color: #333; padding: 20px; }}
        .container {{ max-width: 1100px; margin: 0 auto; }}
        .header {{ background: linear-gradient(135deg, #1F4E79, #2980b9); color: white; padding: 30px; border-radius: 10px; margin-bottom: 20px; }}
        .header h1 {{ font-size: 1.8em; margin-bottom: 8px; }}
        .header .meta {{ opacity: 0.85; font-size: 0.9em; }}
        .stats {{ display: flex; gap: 15px; margin: 20px 0; }}
        .stat {{ background: white; padding: 15px 25px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); text-align: center; flex: 1; }}
        .stat .number {{ font-size: 2em; font-weight: bold; color: #1F4E79; }}
        .stat .label {{ font-size: 0.85em; color: #666; }}
        .category {{ background: white; border-radius: 8px; margin-bottom: 15px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); overflow: hidden; }}
        .cat-header {{ background: #D6E4F0; padding: 15px 20px; cursor: pointer; display: flex; justify-content: space-between; align-items: center; }}
        .cat-header h2 {{ font-size: 1.1em; color: #1F4E79; }}
        .cat-header .badge {{ padding: 3px 10px; border-radius: 12px; font-size: 0.75em; font-weight: bold; }}
        .badge-alta {{ background: #ff6b6b; color: white; }}
        .badge-media {{ background: #ffd93d; color: #333; }}
        .badge-baixa {{ background: #6bcb77; color: white; }}
        .cat-body {{ padding: 0; }}
        .item {{ padding: 12px 20px; border-bottom: 1px solid #eee; display: flex; gap: 15px; align-items: flex-start; }}
        .item:hover {{ background: #f8f9fa; }}
        .item input[type=checkbox] {{ width: 20px; height: 20px; margin-top: 3px; cursor: pointer; }}
        .item-content {{ flex: 1; }}
        .item-desc {{ font-size: 0.95em; font-weight: 500; }}
        .item-details {{ font-size: 0.82em; color: #666; margin-top: 4px; }}
        .item-ref {{ font-size: 0.78em; color: #1F4E79; margin-top: 3px; }}
        .item .crit {{ padding: 2px 8px; border-radius: 10px; font-size: 0.7em; font-weight: bold; white-space: nowrap; }}
        .crit-CRITICA {{ background: #ff6b6b; color: white; }}
        .crit-IMPORTANTE {{ background: #ffd93d; color: #333; }}
        .crit-RECOMENDADA {{ background: #6bcb77; color: white; }}
        .checked .item-desc {{ text-decoration: line-through; color: #999; }}
        .footer {{ text-align: center; color: #999; font-size: 0.8em; margin-top: 30px; padding: 20px; }}
        .progress-bar {{ height: 6px; background: #e0e0e0; border-radius: 3px; margin: 15px 0; }}
        .progress-fill {{ height: 100%; background: linear-gradient(90deg, #6bcb77, #1F4E79); border-radius: 3px; transition: width 0.3s; }}
        .obs-gerais {{ background: white; border-radius: 8px; padding: 20px; margin-top: 15px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        .obs-gerais h3 {{ color: #1F4E79; margin-bottom: 10px; }}
        .obs-gerais li {{ margin: 5px 0; padding-left: 10px; font-size: 0.9em; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{checklist.get('titulo', 'Checklist de Revisão MPD')}</h1>
            <div class="meta">
                Gerado em: {checklist.get('data_geracao', '')} |
                {checklist.get('baseado_em', '')} |
                Versão: {checklist.get('versao', '1.0')}
            </div>
        </div>

        <div class="stats">
            <div class="stat"><div class="number">{total_itens}</div><div class="label">Itens</div></div>
            <div class="stat"><div class="number">{len(categorias)}</div><div class="label">Categorias</div></div>
            <div class="stat"><div class="number" style="color:#ff6b6b">{criticos}</div><div class="label">Críticos</div></div>
            <div class="stat"><div class="number" style="color:#ffd93d">{importantes}</div><div class="label">Importantes</div></div>
        </div>

        <div class="progress-bar"><div class="progress-fill" id="progress" style="width: 0%"></div></div>
        <div style="text-align:center;font-size:0.85em;color:#666" id="progressText">0 / {total_itens} verificados</div>
"""

    for cat in categorias:
        prioridade = cat.get("prioridade", "MEDIA").upper()
        badge_cls = "badge-alta" if "ALT" in prioridade else ("badge-media" if "MED" in prioridade else "badge-baixa")

        html += f"""
        <div class="category">
            <div class="cat-header" onclick="this.nextElementSibling.style.display = this.nextElementSibling.style.display === 'none' ? '' : 'none'">
                <h2>{cat.get('id', '')} – {cat.get('nome', '')}</h2>
                <span class="badge {badge_cls}">{prioridade}</span>
            </div>
            <div class="cat-body">
"""
        for item in cat.get("itens", []):
            crit = item.get("criticidade", "").upper()
            ref = item.get("referencia_normativa", "")
            detalhes = item.get("detalhes", "")
            exemplo = item.get("exemplo_comentario", "")

            html += f"""
                <div class="item" id="{item.get('id', '')}">
                    <input type="checkbox" onchange="toggleItem(this)">
                    <div class="item-content">
                        <div class="item-desc">{item.get('descricao', '')}</div>
                        <div class="item-details">{detalhes}</div>
"""
            if exemplo:
                html += f'                        <div class="item-details" style="font-style:italic">Ex: "{exemplo}"</div>\n'
            if ref:
                html += f'                        <div class="item-ref">📋 {ref}</div>\n'

            html += f"""                    </div>
                    <span class="crit crit-{crit}">{crit}</span>
                </div>
"""
        html += "            </div>\n        </div>\n"

    # Observações gerais
    obs = checklist.get("observacoes_gerais", [])
    if obs:
        html += '        <div class="obs-gerais"><h3>Observações Gerais</h3><ul>\n'
        for o in obs:
            html += f"            <li>{o}</li>\n"
        html += "        </ul></div>\n"

    html += f"""
        <div class="footer">
            Checklist de Revisão CSD-MPD – Gerado automaticamente por análise de IA<br>
            Modelo: {MODELO_IA} | {datetime.now().strftime('%d/%m/%Y %H:%M')}
        </div>
    </div>
    <script>
        const total = {total_itens};
        let checked = 0;
        function toggleItem(cb) {{
            const item = cb.closest('.item');
            if (cb.checked) {{ item.classList.add('checked'); checked++; }}
            else {{ item.classList.remove('checked'); checked--; }}
            document.getElementById('progress').style.width = (checked/total*100) + '%';
            document.getElementById('progressText').textContent = checked + ' / ' + total + ' verificados';
        }}
    </script>
</body>
</html>"""

    CHECKLIST_HTML.write_text(html, encoding="utf-8")
    log.info(f"Checklist HTML salvo: {CHECKLIST_HTML}")


# ==========================================================================
#  PIPELINE PRINCIPAL
# ==========================================================================

def main():
    parser = argparse.ArgumentParser(description="Gerador de Checklist para Revisão de SEQOPs (CSD-MPD)")
    parser.add_argument("--skip-enrich", action="store_true",
                        help="Pular fase de enriquecimento (usar dados já coletados)")
    parser.add_argument("--max-urls", type=int, default=0,
                        help="Limite de URLs para enriquecer (0 = todas)")
    parser.add_argument("--from-json", type=str, default="",
                        help="Carregar dados de JSON em vez de enriquecer")
    args = parser.parse_args()

    log.info("=" * 60)
    log.info("GERADOR DE CHECKLIST – Revisão de SEQOPs CSD-MPD")
    log.info(f"Data/hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    log.info(f"Modelo IA: {MODELO_IA}")
    log.info("=" * 60)

    dados_enriquecidos = []

    # ── Carregar dados enriquecidos existentes ou enriquecer ─────────
    if args.from_json:
        json_path = Path(args.from_json)
        log.info(f"Carregando dados de: {json_path}")
        with open(json_path, "r", encoding="utf-8") as f:
            dados_enriquecidos = json.load(f)
    elif args.skip_enrich and COMENTARIOS_JSON.exists():
        log.info(f"Carregando dados enriquecidos de: {COMENTARIOS_JSON}")
        with open(COMENTARIOS_JSON, "r", encoding="utf-8") as f:
            dados_enriquecidos = json.load(f)
    else:
        # Carregar URLs do Excel existente
        if not ENTRADA_XLSX.exists():
            log.error(f"Arquivo não encontrado: {ENTRADA_XLSX}")
            log.error("Execute primeiro o scraper principal (scraper_seqop.py)")
            return

        df = pd.read_excel(ENTRADA_XLSX, engine="openpyxl")
        log.info(f"Excel carregado: {len(df)} linhas")

        # Extrair URLs únicas
        urls_unicas = []
        urls_vistas = set()
        for _, row in df.iterrows():
            url = str(row.get("URL Fonte", ""))
            if url and url.startswith("http") and url not in urls_vistas:
                urls_vistas.add(url)
                urls_unicas.append({
                    "url": url,
                    "poco": str(row.get("Poço", "")),
                    "titulo": str(row.get("Título", "")),
                    "seq_id": str(row.get("Seq ID", "")),
                })

        log.info(f"URLs únicas: {len(urls_unicas)}")

        # Fase 1: Enriquecer
        dados_enriquecidos = fase1_enriquecer(urls_unicas, max_urls=args.max_urls)

    if not dados_enriquecidos:
        log.error("Sem dados enriquecidos. Abortando.")
        return

    log.info(f"Dados para análise: {len(dados_enriquecidos)} SEQOPs")
    total_mpd = sum(d.get("total_mpd", 0) for d in dados_enriquecidos)
    total_com = sum(d.get("total_comentarios", 0) for d in dados_enriquecidos)
    log.info(f"  Comentários totais: {total_com} ({total_mpd} MPD)")

    # ── Fase 2: Análise com IA ───────────────────────────────────────
    resultado = fase2_analisar(dados_enriquecidos)

    # Salvar resultado bruto
    resultado_json = SCRIPT_DIR / "checklist_resultado_ia.json"
    with open(resultado_json, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)
    log.info(f"Resultado IA salvo: {resultado_json}")

    # ── Fase 3: Gerar saídas ─────────────────────────────────────────
    fase3_gerar_excel(resultado)
    fase3_gerar_html(resultado)

    log.info("=" * 60)
    log.info("CHECKLIST GERADO COM SUCESSO!")
    log.info(f"  Excel: {CHECKLIST_XLSX}")
    log.info(f"  HTML:  {CHECKLIST_HTML}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()

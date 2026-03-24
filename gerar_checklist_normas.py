"""
Gerador de Checklist Normativo para Revisão de SEQOPs (CSD-MPD)
================================================================
Analisa o corpus de padrões normativos Petrobras (Markdown convertido de PDFs)
e usa IA generativa para produzir um checklist estruturado de revisão.

Fases:
  1. Carregar corpus: ler os arquivos .md do diretório corpus/
  2. Extrair requisitos por lote: enviar ao gpt-4o em batches para extrair
     pontos de verificação de cada grupo normativo
  3. Consolidar checklist: unificar os requisitos em um checklist coerente
  4. Gerar saídas: Excel profissional + HTML interativo

Uso:
    C:\\SharedPython\\venv\\Scripts\\python.exe gerar_checklist_normas.py [opções]

    Opções:
      --skip-extract    Pular extração (usar JSON intermediário já salvo)
      --corpus DIR      Diretório do corpus (default: corpus/)
      --modelo MODELO   Modelo de IA (default: gpt-4o)

Pré-requisitos:
    - Corpus de normas convertidas em Markdown (corpus/*.md)
    - Acesso à API de IA Petrobras
"""

# ── Re-launch com o venv compartilhado ──────────────────────────────────────
import subprocess
import sys
from pathlib import Path

_SHARED_PYTHON = Path(r"C:\SharedPython\venv\Scripts\python.exe")
if _SHARED_PYTHON.exists() and Path(sys.executable).resolve() != _SHARED_PYTHON.resolve():
    print(f"[LAUNCH] Relançando com {_SHARED_PYTHON}")
    result = subprocess.run([str(_SHARED_PYTHON)] + sys.argv, check=False)
    sys.exit(result.returncode)

# ── Imports ─────────────────────────────────────────────────────────────────
import json
import logging
import re
import argparse
from dataclasses import dataclass, field
from datetime import datetime
from textwrap import dedent

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Configuração ────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
CORPUS_DIR = SCRIPT_DIR / "corpus"

# Saídas
REQUISITOS_JSON = SCRIPT_DIR / "normas_requisitos_extraidos.json"
CHECKLIST_JSON = SCRIPT_DIR / "checklist_normativo_resultado.json"
CHECKLIST_XLSX = SCRIPT_DIR / "checklist_normativo_mpd.xlsx"
CHECKLIST_HTML = SCRIPT_DIR / "checklist_normativo_mpd.html"

# API de IA
API_BASE = "https://apid.petrobras.com.br/ia/generativos/v1/litellm-chat-petrobras/litellm"
API_KEY = "b320f1f58c9743e9a74048ce64717c89"
MODELO_IA = "gpt-4o"

MODELOS_FALLBACK = [
    MODELO_IA,
    "gpt-4.1",
    "gpt-5-mini",
    "claude-sonnet-4-5",
    "claude-3-7-sonnet",
]

# Limite de caracteres por lote (seguro para ~100K tokens de contexto)
MAX_CHARS_POR_LOTE = 90_000

# ── Logging ─────────────────────────────────────────────────────────────────

LOG_FILE = SCRIPT_DIR / "checklist_normas.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(
            open(sys.stdout.fileno(), mode='w', encoding='utf-8', closefd=False)
        ),
    ],
)
log = logging.getLogger("checklist_normas")

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ==========================================================================
#  DEFINIÇÃO DOS LOTES DE DOCUMENTOS (por relevância e tamanho)
# ==========================================================================

# Grupos de documentos organizados por tema para processamento inteligente.
# Cada grupo será enviado como prompt separado à IA.
LOTES_DOCUMENTOS = [
    {
        "id": "LOTE-01",
        "nome": "Padrão Principal – Detecção e Circulação de Influxo MPD",
        "descricao": "PE-2POC-01113: O padrão mais referenciado, trata de detecção, controle e circulação de influxos com sistema MPD (SBP).",
        "prioridade": "CRITICA",
        "arquivos": [
            "PE-2POC-01113_0e949926.md",
        ],
    },
    {
        "id": "LOTE-02",
        "nome": "Checklists Existentes (Anexos A e B)",
        "descricao": "Checklists oficiais de insumos e projetos MPD já publicados pela Petrobras.",
        "prioridade": "CRITICA",
        "arquivos": [
            "Anexo_A_-_Checklist_Insumos_MPD_38a09c15.md",
            "Anexo_B_-_Checklist_MPD_-_PROJETOS_18d2bb8f.md",
            "Anexo_B_-_Check-list_instalação_da_junta_integrada_4a183bee.md",
        ],
    },
    {
        "id": "LOTE-03",
        "nome": "Testes de Pressão e Janela Operacional (DPPT/DFIT/DLOT)",
        "descricao": "PE-2POC-01232: Procedimentos detalhados para determinação da janela operacional com MPD.",
        "prioridade": "CRITICA",
        "arquivos": [
            "PE-2POC-01232_848639b4.md",
        ],
    },
    {
        "id": "LOTE-04",
        "nome": "Mecanismos de Segurança contra Sobrepressurização",
        "descricao": "Configurações de PRVs, SPL, BIAS, High Limit e ajustes por fornecedor (Halliburton/SLB/Weatherford).",
        "prioridade": "CRITICA",
        "arquivos": [
            "Anexo_D_-_Ajuste_dos_mecanismos_contra_sobrepressurização_c81087ba.md",
        ],
    },
    {
        "id": "LOTE-05",
        "nome": "Manual de Projeto MPD (PE-2POC-01204)",
        "descricao": "Manual completo de projeto MPD: insumos, estratégia, fluidos, BHA, simulações, testes, fingerprint, completação.",
        "prioridade": "ALTA",
        "arquivos": [
            "Manual_de_projeto_MPD_113d6534.md",
        ],
    },
    {
        "id": "LOTE-06",
        "nome": "Perfuração FMCD (PE-2POC-01392)",
        "descricao": "Procedimentos para perfuração em FMCD: preparação, alinhamentos, bombas, BHA, contingências.",
        "prioridade": "ALTA",
        "arquivos": [
            "PE-2POC-01392_8131d840.md",
        ],
    },
    {
        "id": "LOTE-07",
        "nome": "Equipamentos e Operação MPD",
        "descricao": "PE-2POC-01247 (equipamentos MPD): RCD, choke, manifold, MGS, parâmetros operacionais. "
                     "PE-2POC-01257 (completação com MPD).",
        "prioridade": "ALTA",
        "arquivos": [
            "PE-2POC-01247_3024543a.md",
            "PE-2POC-01257_0965985d.md",
        ],
    },
    {
        "id": "LOTE-08",
        "nome": "Critérios de Aprovação e Treinamento",
        "descricao": "Critérios de aprovação HP MPD, matriz de treinamento, e padrões complementares.",
        "prioridade": "MEDIA",
        "arquivos": [
            "Critérios_aprovação_HP_MPD_a2e48c3f.md",
            "Matriz_de_treinamento_MPD_e7332a3b.md",
            "Ficha_Registro_Treinamento_Prático_MPD_mai22_03236203.md",
        ],
    },
    {
        "id": "LOTE-09",
        "nome": "Documentos Complementares e P&ID",
        "descricao": "Modelo de projeto MPD, buffer manifold, P&IDs genéricos, envelope operacional, árvore de decisão de perda.",
        "prioridade": "MEDIA",
        "arquivos": [
            "Anexo_E_-_Modelo_de_Projeto_MPD_c35fa95f.md",
            "Anexo_C_-_Buffer_Manifold_e_P&ID_genéricos_MPD_8c3d7dd7.md",
            "Anexo_D_-_Árvore_de_decisão_de_perda_3ed54519.md",
            "Anexo_E_-_Envelope_Operacional_dos_Conjuntos_de_Vedação_48a6bd3c.md",
        ],
    },
    {
        "id": "LOTE-10",
        "nome": "Padrões Complementares (PEs adicionais)",
        "descricao": "PE-2POC-01203 (teste de pressão linhas flutuantes), PE-2POC-01204 (projeto MPD), "
                     "PE-1PBR-00486, PE-1PBR-01123, PE-1PBR-01326 (fingerprint), PE-1PBR-01580.",
        "prioridade": "MEDIA",
        "arquivos": [
            "PE-2POC-01203_e20f3a77.md",
            "PE-2POC-01204_3126b152.md",
            "PE-1PBR-00486_64d96a37.md",
            "PE-1PBR-01123_a2d54195.md",
            "PE-1PBR-01326_3b2b98f8.md",
            "PE-1PBR-01580_96230712.md",
        ],
    },
]


# ==========================================================================
#  FUNÇÕES AUXILIARES
# ==========================================================================

def _salvar_json(dados, caminho: Path):
    """Salva dados em JSON com formatação."""
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    log.info(f"  JSON salvo: {caminho}")


def _carregar_json(caminho: Path) -> dict | list | None:
    """Carrega JSON de arquivo."""
    if caminho.exists():
        with open(caminho, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def _chamar_ia(mensagens: list[dict], max_tokens: int = 8000, temperature: float = 0.2) -> str:
    """Chama a API de IA generativa. Tenta múltiplos modelos em caso de falha."""
    headers = {
        "Content-Type": "application/json",
        "api-key": API_KEY,
    }
    payload = {
        "messages": mensagens,
        "max_tokens": max_tokens,
        "temperature": temperature,
    }

    for modelo in MODELOS_FALLBACK:
        url = f"{API_BASE}/engines/{modelo}/chat/completions"
        log.info(f"    Chamando modelo: {modelo}")

        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=300, verify=False)

            if resp.status_code != 200:
                log.warning(f"    Modelo {modelo}: HTTP {resp.status_code} – {resp.text[:300]}")
                continue

            data = resp.json()
            choices = data.get("choices", [])
            if not choices:
                log.warning(f"    Modelo {modelo}: Resposta sem choices")
                continue

            content = choices[0].get("message", {}).get("content", "")
            usage = data.get("usage", {})
            log.info(f"    ✓ {modelo} – tokens: prompt={usage.get('prompt_tokens')}, "
                     f"completion={usage.get('completion_tokens')}")
            return content

        except requests.exceptions.Timeout:
            log.warning(f"    Modelo {modelo}: Timeout (300s)")
        except Exception as e:
            log.warning(f"    Modelo {modelo}: Erro – {e}")

    raise RuntimeError(f"Todos os modelos falharam: {MODELOS_FALLBACK}")


def _extrair_json(texto: str) -> dict | list | None:
    """Extrai JSON de uma resposta de texto (pode ter markdown ou texto ao redor)."""
    # Tentar extrair de code block primeiro
    m = re.search(r'```(?:json)?\s*\n?([\s\S]*?)\n?```', texto)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass

    # Tentar extrair objeto ou array mais externo
    for pattern in [r'\{[\s\S]*\}', r'\[[\s\S]*\]']:
        m = re.search(pattern, texto)
        if m:
            try:
                return json.loads(m.group())
            except json.JSONDecodeError:
                continue

    return None


# ==========================================================================
#  FASE 1: CARREGAR CORPUS
# ==========================================================================

def fase1_carregar_corpus(corpus_dir: Path) -> dict[str, str]:
    """Carrega todos os arquivos .md do corpus.

    Returns:
        Dict mapeando nome_arquivo → conteúdo_texto
    """
    log.info("=" * 60)
    log.info("FASE 1: Carregando corpus de normas")
    log.info(f"  Diretório: {corpus_dir}")

    arquivos = {}
    for md_file in sorted(corpus_dir.glob("*.md")):
        # Pular COMBINED.md e INDEX/PID_INDEX
        if md_file.name in ("COMBINED.md", "INDEX.md", "PID_INDEX.md"):
            continue
        try:
            texto = md_file.read_text(encoding="utf-8")
            arquivos[md_file.name] = texto
            log.info(f"  ✓ {md_file.name} ({len(texto):,} chars)")
        except Exception as e:
            log.warning(f"  ✗ {md_file.name}: {e}")

    log.info(f"  Total: {len(arquivos)} documentos, "
             f"{sum(len(t) for t in arquivos.values()):,} chars")
    return arquivos


# ==========================================================================
#  FASE 2: EXTRAIR REQUISITOS POR LOTE
# ==========================================================================

PROMPT_SISTEMA_EXTRACAO = dedent("""\
Você é um especialista em perfuração de poços de petróleo offshore com profundo
conhecimento dos padrões normativos Petrobras para operações MPD (Managed Pressure Drilling).

Sua tarefa é analisar documentos normativos e extrair TODOS os pontos de verificação
relevantes para revisão de SEQOPs (sequências operacionais).

Responda SEMPRE em JSON válido, em português brasileiro.
""")


def _prompt_extracao_lote(lote: dict, textos: list[str]) -> str:
    """Monta o prompt para extrair requisitos de um lote de documentos."""
    return dedent(f"""\
    Analise os documentos normativos abaixo e extraia TODOS os requisitos, verificações
    e pontos de checklist relevantes para revisão de SEQOPs MPD.

    CONTEXTO DO LOTE: {lote['nome']}
    {lote['descricao']}

    TAREFA: Para cada requisito encontrado, extraia:
    - categoria: tema geral (ex: "Testes de Pressão", "Equipamentos", "Controle de Poço")
    - requisito: descrição clara do que deve ser verificado na SEQOP
    - detalhes: informações adicionais, limites numéricos, condições
    - secao: seção/parágrafo do documento (se identificável)
    - documento_fonte: nome do PE/Anexo de onde foi extraído
    - criticidade: CRITICA (segurança/mandatory), IMPORTANTE (best practice), RECOMENDADA (nice-to-have)

    Responda em JSON:
    {{
      "lote_id": "{lote['id']}",
      "lote_nome": "{lote['nome']}",
      "requisitos": [
        {{
          "categoria": "...",
          "requisito": "...",
          "detalhes": "...",
          "secao": "§X.Y",
          "documento_fonte": "PE-XXXX-XXXXX",
          "criticidade": "CRITICA|IMPORTANTE|RECOMENDADA"
        }}
      ]
    }}

    INSTRUÇÕES IMPORTANTES:
    - Extraia TODOS os requisitos, não apenas os mais importantes
    - Inclua limites numéricos específicos (pressões, volumes, percentuais)
    - Mantenha a rastreabilidade: indique a seção de origem
    - Foque em itens verificáveis em uma SEQOP
    - Não invente requisitos — extraia apenas o que está nos documentos

    DOCUMENTOS PARA ANÁLISE:
    {"=" * 60}
    {"=" * 60 + chr(10)}""") + "\n".join(textos)


def fase2_extrair_requisitos(corpus: dict[str, str], corpus_dir: Path) -> list[dict]:
    """Processa cada lote de documentos e extrai requisitos via IA.

    Returns:
        Lista de dicts com requisitos extraídos por lote
    """
    log.info("=" * 60)
    log.info("FASE 2: Extraindo requisitos dos documentos normativos")
    log.info("=" * 60)

    todos_resultados = []

    for lote in LOTES_DOCUMENTOS:
        log.info(f"\n  ── {lote['id']}: {lote['nome']} ──")

        # Carregar textos do lote
        textos = []
        chars_lote = 0
        for arq_nome in lote["arquivos"]:
            if arq_nome in corpus:
                texto = corpus[arq_nome]
                # Truncar se necessário para caber no lote
                if chars_lote + len(texto) > MAX_CHARS_POR_LOTE:
                    disponivel = MAX_CHARS_POR_LOTE - chars_lote
                    if disponivel > 5000:  # só inclui se ainda tiver espaço significativo
                        texto = texto[:disponivel] + "\n\n[... DOCUMENTO TRUNCADO ...]"
                        textos.append(f"\n--- DOCUMENTO: {arq_nome} ---\n{texto}")
                        chars_lote += len(texto)
                    else:
                        log.warning(f"    ⚠ {arq_nome} não coube no lote ({len(texto):,} chars)")
                    continue

                textos.append(f"\n--- DOCUMENTO: {arq_nome} ---\n{texto}")
                chars_lote += len(texto)
                log.info(f"    ✓ {arq_nome} ({len(texto):,} chars)")
            else:
                log.warning(f"    ✗ {arq_nome} não encontrado no corpus")

        if not textos:
            log.warning(f"    Nenhum documento encontrado para este lote!")
            continue

        log.info(f"    Total do lote: {chars_lote:,} chars")

        # Chamar IA
        log.info(f"    Enviando para IA...")
        prompt = _prompt_extracao_lote(lote, textos)

        try:
            resposta = _chamar_ia([
                {"role": "system", "content": PROMPT_SISTEMA_EXTRACAO},
                {"role": "user", "content": prompt},
            ], max_tokens=12000, temperature=0.15)

            resultado = _extrair_json(resposta)
            if resultado:
                n_req = len(resultado.get("requisitos", []))
                log.info(f"    ✓ {n_req} requisitos extraídos")
                resultado["lote_prioridade"] = lote["prioridade"]
                todos_resultados.append(resultado)
            else:
                log.warning(f"    ✗ Não foi possível parsear JSON da resposta")
                todos_resultados.append({
                    "lote_id": lote["id"],
                    "lote_nome": lote["nome"],
                    "requisitos": [],
                    "erro": "JSON inválido",
                    "resposta_raw": resposta[:2000],
                })

        except Exception as e:
            log.error(f"    ✗ Erro: {e}")
            todos_resultados.append({
                "lote_id": lote["id"],
                "lote_nome": lote["nome"],
                "requisitos": [],
                "erro": str(e),
            })

    # Salvar resultados intermediários
    _salvar_json(todos_resultados, REQUISITOS_JSON)

    total_req = sum(len(r.get("requisitos", [])) for r in todos_resultados)
    log.info(f"\nFASE 2 completa: {total_req} requisitos de {len(todos_resultados)} lotes")
    return todos_resultados


# ==========================================================================
#  FASE 3: CONSOLIDAR CHECKLIST
# ==========================================================================

def fase3_consolidar(requisitos_por_lote: list[dict]) -> dict:
    """Envia os requisitos extraídos para a IA consolidar em um checklist final.

    Usa duas chamadas à IA:
      - Passo A: Primeira metade das categorias
      - Passo B: Segunda metade das categorias
    Depois faz merge local.
    """
    log.info("=" * 60)
    log.info("FASE 3: Consolidando checklist com IA")
    log.info("=" * 60)

    # Montar texto consolidado dos requisitos
    todos_requisitos = []
    for lote in requisitos_por_lote:
        for req in lote.get("requisitos", []):
            req["_lote"] = lote.get("lote_nome", "")
            req["_lote_prioridade"] = lote.get("lote_prioridade", "MEDIA")
            todos_requisitos.append(req)

    log.info(f"  Total de requisitos para consolidação: {len(todos_requisitos)}")

    if not todos_requisitos:
        log.error("  Nenhum requisito extraído! Abortando.")
        return {}

    # Preparar texto resumido dos requisitos
    linhas_requisitos = []
    chars = 0
    for i, req in enumerate(todos_requisitos, 1):
        linha = (
            f"[{i}] ({req.get('criticidade', '?')}) "
            f"[{req.get('documento_fonte', '?')} {req.get('secao', '')}] "
            f"Cat: {req.get('categoria', '?')} | "
            f"{req.get('requisito', '')} | "
            f"Det: {req.get('detalhes', '')[:200]}"
        )
        if chars + len(linha) > 70000:
            log.warning(f"  Truncando requisitos em {i-1} de {len(todos_requisitos)}")
            break
        linhas_requisitos.append(linha)
        chars += len(linha)

    log.info(f"  Usando {len(linhas_requisitos)} requisitos ({chars:,} chars)")

    # ── Categorias por tipo de SEQOP ─────────────────────────────────
    # Dividimos em 3 passos para garantir completude na geração
    categorias_passo_a = [
        ("Perfuração MPD (SBP)",
         "Itens de verificação para SEQOPs de perfuração com sistema MPD usando técnica SBP. "
         "Inclui ajuste de back-pressure, alarmes, monitoramento, modelo hidráulico, envelope operacional."),
        ("Perfuração FMCD",
         "Itens para SEQOPs de perfuração em Full Mud Cap Drilling (perda total). "
         "Inclui alinhamentos, bombas booster, BHA, combate a perda, conversão SBP→FMCD."),
        ("Perfuração PMCD",
         "Itens para SEQOPs de Pressurized Mud Cap Drilling. "
         "Inclui bombeio pelo anular, monitoramento de nível, contingências de migração de gás."),
        ("Instalação da Junta Integrada",
         "Itens para SEQOPs de instalação da junta integrada MPD (RCD + DSIT). "
         "Inclui testes, torques, vedações, tally de riser, mangueiras, umbilicais."),
    ]
    categorias_passo_b = [
        ("Fingerprint",
         "Itens para SEQOPs de fingerprinting MPD. "
         "Inclui calibração, flowcheck dinâmico, compressibilidade do fluido, simulado MPD Drill."),
        ("DPPT",
         "Itens para SEQOPs de Dynamic Pore Pressure Test — determinação do limite inferior da janela operacional. "
         "Inclui steps de despressurização, critérios de detecção, limites."),
        ("DFIT",
         "Itens para SEQOPs de DFIT/DLOT — determinação do limite superior da janela operacional. "
         "Inclui steps de pressurização, ajuste de alarmes, envelope RCD."),
        ("Circulação de Influxo MPD",
         "Itens para SEQOPs de circulação de influxo pelo sistema MPD. "
         "Inclui os 9 critérios obrigatórios do PE-2POC-01113 §3.7.3.3, limites de volume/pressão, MGS."),
    ]
    categorias_passo_c = [
        ("Descida de Cauda em MPD",
         "Itens para SEQOPs de descida de coluna de completação/cauda usando sistema MPD (SBP). "
         "Inclui testes de pressão, monitoramento, manobra com back-pressure."),
        ("Descida de Cauda em FMCD",
         "Itens para SEQOPs de descida de cauda em cenário FMCD. "
         "Inclui alinhamentos, bombeio pelo anular, BHA, NRVs."),
        ("Descida de Cauda em PMCD",
         "Itens para SEQOPs de descida de cauda em cenário PMCD. "
         "Inclui monitoramento de nível, bombeio contínuo, contingências."),
    ]

    requisitos_texto = "\n".join(linhas_requisitos)

    def _gerar_passo(cats_lista: list[tuple[str, str]], id_inicio: int, passo_label: str) -> list[dict]:
        cats_str = "\n".join(f"  - **{nome}**: {desc}" for nome, desc in cats_lista)
        prompt = dedent(f"""\
        Você é um especialista em perfuração MPD na Petrobras, revisando sequências operacionais (SEQOPs).

        Abaixo estão {len(linhas_requisitos)} requisitos extraídos de documentos normativos MPD Petrobras.

        TAREFA: Gere um checklist de revisão para os seguintes TIPOS DE SEQOP:

{cats_str}

        Para CADA tipo de SEQOP, gere de 5 a 12 itens de verificação que um revisor CSD-MPD
        deve conferir ao analisar esse tipo de sequência operacional.

        REGRAS:
        - Cada item deve ser uma VERIFICAÇÃO CONCRETA que o revisor pode marcar SIM/NÃO na SEQOP
        - Use linguagem imperativa: "Verificar se...", "Confirmar que...", "A SEQOP inclui..."
        - INCLUA limites numéricos específicos (pressões em psi, volumes em bbl, percentuais)
        - Preserve referências normativas (PE-2POC-XXXXX §X.Y)
        - Criticidade: CRITICA (segurança/mandatory), IMPORTANTE (norma), RECOMENDADA (boa prática)
        - NÃO invente requisitos — baseie-se APENAS nos dados normativos fornecidos
        - Foque no que deve CONSTAR NA SEQOP, não em procedimentos operacionais genéricos

        Responda em JSON:
        {{
          "categorias": [
            {{
              "id": "CAT-{id_inicio:02d}",
              "nome": "Nome do tipo de SEQOP",
              "descricao": "Descrição do escopo",
              "prioridade": "ALTA/MEDIA/BAIXA",
              "itens": [
                {{
                  "id": "CAT-{id_inicio:02d}-001",
                  "descricao": "Verificação concreta para a SEQOP",
                  "detalhes": "Detalhes, limites numéricos e condições",
                  "referencia_normativa": "PE-XXXX §X.Y",
                  "criticidade": "CRITICA/IMPORTANTE/RECOMENDADA",
                  "aplicabilidade": "Condição de aplicação (se houver)"
                }}
              ]
            }}
          ]
        }}

        REQUISITOS NORMATIVOS EXTRAÍDOS:
        """) + requisitos_texto

        log.info(f"  Passo {passo_label}: gerando {len(cats_lista)} categorias...")
        resposta = _chamar_ia([
            {"role": "system", "content": (
                "Você é um especialista em perfuração MPD da Petrobras que revisa SEQOPs. "
                "Gere o checklist em JSON válido, em português brasileiro. "
                "Seja COMPLETO: inclua TODOS os itens relevantes para cada tipo de SEQOP. "
                "Produza de 5 a 12 itens por categoria. Cada item deve ser verificável na SEQOP."
            )},
            {"role": "user", "content": prompt},
        ], max_tokens=16000, temperature=0.15)

        resultado = _extrair_json(resposta)
        if resultado and "categorias" in resultado:
            cats = resultado["categorias"]
            n = sum(len(c.get("itens", [])) for c in cats)
            log.info(f"    ✓ Passo {passo_label}: {len(cats)} categorias, {n} itens")
            return cats
        else:
            log.warning(f"    ✗ Passo {passo_label}: JSON inválido")
            log.warning(f"      Resposta: {resposta[:500] if resposta else '(vazia)'}")
            return []

    # ── Executar os três passos ──────────────────────────────────────
    cats_a = _gerar_passo(categorias_passo_a, id_inicio=1, passo_label="A (Perfuração + Junta)")
    cats_b = _gerar_passo(categorias_passo_b, id_inicio=5, passo_label="B (FP + Testes + Influxo)")
    cats_c = _gerar_passo(categorias_passo_c, id_inicio=9, passo_label="C (Descidas de Cauda)")

    # Merge
    todas_categorias = cats_a + cats_b + cats_c

    # Re-numerar IDs de forma consistente
    for ci, cat in enumerate(todas_categorias, 1):
        cat["id"] = f"CAT-{ci:02d}"
        for ii, item in enumerate(cat.get("itens", []), 1):
            item["id"] = f"CAT-{ci:02d}-{ii:03d}"

    # Extrair documentos_fonte
    docs_set = set()
    for req in todos_requisitos:
        doc = req.get("documento_fonte", "")
        if doc and "?" not in doc:
            docs_set.add(doc)

    resultado = {
        "titulo": "Checklist Normativo de Revisão de SEQOPs – CSD-MPD",
        "versao": "1.0",
        "data_geracao": datetime.now().strftime('%d/%m/%Y'),
        "baseado_em": f"{len(todos_requisitos)} requisitos de {len(requisitos_por_lote)} documentos normativos Petrobras",
        "categorias": todas_categorias,
        "observacoes_gerais": [
            "Este checklist foi gerado automaticamente a partir da análise de documentos normativos Petrobras.",
            "Todos os itens devem ser validados por especialista antes de uso em produção.",
            "Referências normativas devem ser consultadas para detalhes completos.",
            "Itens marcados como CRITICA estão diretamente relacionados à segurança de poço.",
        ],
        "documentos_fonte": sorted(docs_set),
        "_total_requisitos_fonte": len(todos_requisitos),
        "_total_lotes": len(requisitos_por_lote),
        "_modelo": MODELO_IA,
    }

    _salvar_json(resultado, CHECKLIST_JSON)

    n_cats = len(todas_categorias)
    n_itens = sum(len(c.get("itens", [])) for c in todas_categorias)
    log.info(f"FASE 3 completa: {n_cats} categorias, {n_itens} itens no checklist")

    return resultado


# ==========================================================================
#  FASE 4: GERAR EXCEL PROFISSIONAL
# ==========================================================================

def fase4_gerar_excel(checklist: dict):
    """Gera o checklist normativo em formato Excel profissional."""
    log.info("=" * 60)
    log.info("FASE 4a: Gerando Excel")

    if not checklist.get("categorias"):
        log.error("  Checklist vazio!")
        return

    wb = Workbook()

    # ── Aba 1: Checklist ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Checklist Normativo MPD"

    # Estilos
    hdr_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    cat_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    cat_font = Font(bold=True, size=11, color="1B3A5C")
    crit_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    imp_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    rec_fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Título
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1,
                         value=checklist.get("titulo", "Checklist Normativo de Revisão de SEQOPs – CSD-MPD"))
    title_cell.font = Font(bold=True, size=14, color="1B3A5C")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Metadados (2 linhas)
    ws.merge_cells("A2:H2")
    meta = (f"Gerado em: {checklist.get('data_geracao', datetime.now().strftime('%d/%m/%Y'))} | "
            f"Baseado em: {checklist.get('baseado_em', 'N/A')} | "
            f"Versão: {checklist.get('versao', '1.0')} | Modelo: {MODELO_IA}")
    ws.cell(row=2, column=1, value=meta).font = Font(italic=True, size=9, color="666666")

    docs_fonte = checklist.get("documentos_fonte", [])
    if docs_fonte:
        ws.merge_cells("A3:H3")
        ws.cell(row=3, column=1,
                value=f"Documentos: {', '.join(docs_fonte[:15])}").font = Font(italic=True, size=8, color="888888")

    # Cabeçalhos
    colunas = [
        ("ID", 14), ("✓", 5), ("Verificação", 55), ("Detalhes / Limites", 50),
        ("Ref. Normativa", 22), ("Criticidade", 13), ("Aplicabilidade", 25),
        ("Observações do Revisor", 35),
    ]
    header_row = 5
    for ci, (nome, larg) in enumerate(colunas, 1):
        c = ws.cell(row=header_row, column=ci, value=nome)
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[c.column_letter].width = larg

    row_num = header_row + 1
    categorias = checklist.get("categorias", [])

    for cat in categorias:
        # Linha de categoria
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
        cat_cell = ws.cell(
            row=row_num, column=1,
            value=f"{cat.get('id', '')} – {cat.get('nome', '')} [{cat.get('prioridade', '')}]"
        )
        cat_cell.fill = cat_fill
        cat_cell.font = cat_font
        cat_cell.alignment = Alignment(vertical="center")
        for ci in range(1, 9):
            ws.cell(row=row_num, column=ci).border = border
        ws.row_dimensions[row_num].height = 24
        row_num += 1

        # Descrição da categoria
        if cat.get("descricao"):
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            desc_cell = ws.cell(row=row_num, column=1, value=f"  {cat['descricao']}")
            desc_cell.font = Font(italic=True, size=9, color="444444")
            desc_cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # Itens
        for item in cat.get("itens", []):
            criticidade = item.get("criticidade", "").upper()

            # ID
            c_id = ws.cell(row=row_num, column=1, value=item.get("id", ""))
            c_id.border = border
            c_id.alignment = Alignment(horizontal="center", vertical="center")
            c_id.font = Font(size=9)

            # Checkbox
            check = ws.cell(row=row_num, column=2, value="☐")
            check.border = border
            check.alignment = center
            check.font = Font(size=14)

            # Verificação
            verif = ws.cell(row=row_num, column=3, value=item.get("descricao", ""))
            verif.border = border
            verif.alignment = wrap
            verif.font = Font(size=10)

            # Detalhes
            det = ws.cell(row=row_num, column=4, value=item.get("detalhes", ""))
            det.border = border
            det.alignment = wrap
            det.font = Font(size=9)

            # Ref normativa
            ref = ws.cell(row=row_num, column=5, value=item.get("referencia_normativa", ""))
            ref.border = border
            ref.alignment = Alignment(wrap_text=True, vertical="center")
            ref.font = Font(size=9)

            # Criticidade
            crit = ws.cell(row=row_num, column=6, value=criticidade)
            crit.border = border
            crit.alignment = center
            crit.font = Font(size=9)
            if "CRIT" in criticidade:
                crit.fill = crit_fill
                crit.font = Font(bold=True, color="FFFFFF", size=9)
            elif "IMPORT" in criticidade:
                crit.fill = imp_fill
            elif "RECOM" in criticidade:
                crit.fill = rec_fill

            # Aplicabilidade
            aplic = ws.cell(row=row_num, column=7, value=item.get("aplicabilidade", ""))
            aplic.border = border
            aplic.alignment = wrap
            aplic.font = Font(size=9)

            # Observações (vazia)
            obs = ws.cell(row=row_num, column=8, value="")
            obs.border = border

            ws.row_dimensions[row_num].height = 50
            row_num += 1

    # Observações gerais
    row_num += 1
    obs_gerais = checklist.get("observacoes_gerais", [])
    if obs_gerais:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
        ws.cell(row=row_num, column=1, value="OBSERVAÇÕES GERAIS").font = Font(bold=True, size=11)
        row_num += 1
        for o in obs_gerais:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            ws.cell(row=row_num, column=1, value=f"• {o}").alignment = Alignment(wrap_text=True)
            row_num += 1

    ws.auto_filter.ref = f"A{header_row}:H{row_num}"
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Aba 2: Rastreabilidade ────────────────────────────────────────
    ws2 = wb.create_sheet("Rastreabilidade")
    ws2.cell(row=1, column=1, value="Rastreabilidade – Requisitos por Documento Normativo").font = Font(bold=True, size=12)

    # Cabeçalhos
    for ci, (nome, larg) in enumerate([
        ("Lote", 10), ("Documento", 25), ("Prioridade", 12),
        ("Requisitos", 10), ("Categorias", 40)
    ], 1):
        c = ws2.cell(row=3, column=ci, value=nome)
        c.fill = hdr_fill
        c.font = hdr_font
        c.border = border
        ws2.column_dimensions[c.column_letter].width = larg

    row2 = 4
    requisitos_raw = _carregar_json(REQUISITOS_JSON)
    if requisitos_raw:
        for lote in requisitos_raw:
            reqs = lote.get("requisitos", [])
            cats_set = set(r.get("categoria", "?") for r in reqs)
            ws2.cell(row=row2, column=1, value=lote.get("lote_id", "")).border = border
            ws2.cell(row=row2, column=2, value=lote.get("lote_nome", "")[:50]).border = border
            ws2.cell(row=row2, column=3, value=lote.get("lote_prioridade", "")).border = border
            ws2.cell(row=row2, column=4, value=len(reqs)).border = border
            ws2.cell(row=row2, column=5, value=", ".join(sorted(cats_set))).border = border
            ws2.cell(row=row2, column=5).alignment = wrap
            ws2.row_dimensions[row2].height = 22
            row2 += 1

    wb.save(CHECKLIST_XLSX)
    log.info(f"  Excel salvo: {CHECKLIST_XLSX}")


# ==========================================================================
#  FASE 4b: GERAR HTML INTERATIVO
# ==========================================================================

def fase4_gerar_html(checklist: dict):
    """Gera o checklist normativo em formato HTML interativo."""
    log.info("  Gerando HTML interativo...")

    categorias = checklist.get("categorias", [])
    if not categorias:
        log.error("  Checklist vazio!")
        return

    total_itens = sum(len(cat.get("itens", [])) for cat in categorias)
    criticos = sum(1 for cat in categorias for i in cat.get("itens", [])
                   if "CRIT" in i.get("criticidade", "").upper())
    importantes = sum(1 for cat in categorias for i in cat.get("itens", [])
                      if "IMPORT" in i.get("criticidade", "").upper())
    recomendados = total_itens - criticos - importantes

    docs_fonte = checklist.get("documentos_fonte", [])

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{checklist.get('titulo', 'Checklist Normativo MPD')}</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: 'Segoe UI', Tahoma, sans-serif; background: #f0f2f5; color: #333; padding: 20px; }}
        .container {{ max-width: 1200px; margin: 0 auto; }}

        /* Header */
        .header {{
            background: linear-gradient(135deg, #1B3A5C, #2E86AB);
            color: white; padding: 30px; border-radius: 12px; margin-bottom: 20px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.15);
        }}
        .header h1 {{ font-size: 1.7em; margin-bottom: 8px; }}
        .header .meta {{ opacity: 0.85; font-size: 0.88em; line-height: 1.5; }}
        .header .docs {{ opacity: 0.7; font-size: 0.78em; margin-top: 8px; }}

        /* Stats */
        .stats {{ display: flex; gap: 12px; margin: 15px 0; flex-wrap: wrap; }}
        .stat {{
            background: white; padding: 15px 20px; border-radius: 10px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08); text-align: center; flex: 1; min-width: 120px;
        }}
        .stat .number {{ font-size: 2em; font-weight: bold; color: #1B3A5C; }}
        .stat .label {{ font-size: 0.82em; color: #888; }}

        /* Filter buttons */
        .filters {{ margin: 15px 0; display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }}
        .filters label {{ font-size: 0.85em; color: #666; margin-right: 5px; }}
        .filter-btn {{
            padding: 6px 14px; border: 1px solid #ccc; background: white; border-radius: 20px;
            cursor: pointer; font-size: 0.82em; transition: all 0.2s;
        }}
        .filter-btn:hover {{ border-color: #1B3A5C; color: #1B3A5C; }}
        .filter-btn.active {{ background: #1B3A5C; color: white; border-color: #1B3A5C; }}

        /* Progress */
        .progress-bar {{ height: 8px; background: #e0e0e0; border-radius: 4px; margin: 12px 0; }}
        .progress-fill {{
            height: 100%; background: linear-gradient(90deg, #6bcb77, #2E86AB, #1B3A5C);
            border-radius: 4px; transition: width 0.4s ease;
        }}
        .progress-text {{ text-align: center; font-size: 0.85em; color: #666; }}

        /* Categories */
        .category {{
            background: white; border-radius: 10px; margin-bottom: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08); overflow: hidden;
        }}
        .cat-header {{
            background: #C5D9F1; padding: 14px 20px; cursor: pointer;
            display: flex; justify-content: space-between; align-items: center;
            transition: background 0.2s;
        }}
        .cat-header:hover {{ background: #b0cce8; }}
        .cat-header h2 {{ font-size: 1.05em; color: #1B3A5C; }}
        .cat-header .cat-info {{
            display: flex; gap: 8px; align-items: center;
        }}
        .cat-header .cat-count {{
            font-size: 0.8em; color: #666; background: white; padding: 2px 10px;
            border-radius: 12px;
        }}
        .badge {{ padding: 3px 10px; border-radius: 12px; font-size: 0.72em; font-weight: 600; }}
        .badge-alta {{ background: #ff6b6b; color: white; }}
        .badge-media {{ background: #ffd93d; color: #333; }}
        .badge-baixa {{ background: #6bcb77; color: white; }}

        .cat-body {{ padding: 0; }}

        /* Items */
        .item {{
            padding: 12px 20px; border-bottom: 1px solid #f0f0f0;
            display: flex; gap: 14px; align-items: flex-start; transition: background 0.2s;
        }}
        .item:hover {{ background: #f8f9fa; }}
        .item:last-child {{ border-bottom: none; }}
        .item input[type=checkbox] {{
            width: 20px; height: 20px; margin-top: 3px; cursor: pointer;
            accent-color: #1B3A5C;
        }}
        .item-content {{ flex: 1; }}
        .item-desc {{ font-size: 0.93em; font-weight: 500; line-height: 1.4; }}
        .item-details {{ font-size: 0.82em; color: #666; margin-top: 4px; line-height: 1.3; }}
        .item-ref {{ font-size: 0.78em; color: #1B3A5C; margin-top: 4px; font-weight: 500; }}
        .item-aplic {{ font-size: 0.75em; color: #888; margin-top: 2px; font-style: italic; }}
        .item .crit {{
            padding: 3px 10px; border-radius: 12px; font-size: 0.7em;
            font-weight: 600; white-space: nowrap; flex-shrink: 0;
        }}
        .crit-CRITICA {{ background: #ff6b6b; color: white; }}
        .crit-IMPORTANTE {{ background: #ffd93d; color: #333; }}
        .crit-RECOMENDADA {{ background: #6bcb77; color: white; }}
        .checked .item-desc {{ text-decoration: line-through; color: #999; }}
        .checked .item-details {{ color: #bbb; }}
        .hidden {{ display: none !important; }}

        /* Footer */
        .footer {{ text-align: center; color: #aaa; font-size: 0.78em; margin-top: 30px; padding: 20px; }}

        /* Obs gerais */
        .obs-gerais {{
            background: white; border-radius: 10px; padding: 20px; margin-top: 15px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }}
        .obs-gerais h3 {{ color: #1B3A5C; margin-bottom: 10px; font-size: 1em; }}
        .obs-gerais li {{ margin: 6px 0; padding-left: 10px; font-size: 0.88em; color: #555; }}

        /* Print */
        @media print {{
            body {{ background: white; padding: 0; }}
            .header {{ background: #1B3A5C !important; -webkit-print-color-adjust: exact; }}
            .filters, .progress-bar, .progress-text {{ display: none; }}
            .category {{ page-break-inside: avoid; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📋 {checklist.get('titulo', 'Checklist Normativo MPD')}</h1>
            <div class="meta">
                Gerado em: {checklist.get('data_geracao', '')} |
                {checklist.get('baseado_em', '')} |
                Versão: {checklist.get('versao', '1.0')} | Modelo: {MODELO_IA}
            </div>
            <div class="docs">Documentos: {', '.join(docs_fonte[:15]) if docs_fonte else 'N/A'}</div>
        </div>

        <div class="stats">
            <div class="stat"><div class="number">{total_itens}</div><div class="label">Total de Itens</div></div>
            <div class="stat"><div class="number">{len(categorias)}</div><div class="label">Categorias</div></div>
            <div class="stat"><div class="number" style="color:#ff6b6b">{criticos}</div><div class="label">Críticos</div></div>
            <div class="stat"><div class="number" style="color:#e6a700">{importantes}</div><div class="label">Importantes</div></div>
            <div class="stat"><div class="number" style="color:#6bcb77">{recomendados}</div><div class="label">Recomendados</div></div>
        </div>

        <div class="filters">
            <label>Filtrar:</label>
            <button class="filter-btn active" onclick="filterItems('ALL')">Todos</button>
            <button class="filter-btn" onclick="filterItems('CRITICA')">🔴 Críticos</button>
            <button class="filter-btn" onclick="filterItems('IMPORTANTE')">🟡 Importantes</button>
            <button class="filter-btn" onclick="filterItems('RECOMENDADA')">🟢 Recomendados</button>
            <button class="filter-btn" onclick="filterItems('UNCHECKED')">⬜ Pendentes</button>
        </div>

        <div class="progress-bar"><div class="progress-fill" id="progress" style="width: 0%"></div></div>
        <div class="progress-text" id="progressText">0 / {total_itens} verificados (0%)</div>
"""

    for cat in categorias:
        prioridade = cat.get("prioridade", "MEDIA").upper()
        badge_cls = "badge-alta" if "ALT" in prioridade else ("badge-media" if "MED" in prioridade else "badge-baixa")
        n_itens_cat = len(cat.get("itens", []))

        html += f"""
        <div class="category">
            <div class="cat-header" onclick="toggleCategory(this)">
                <h2>{cat.get('id', '')} – {cat.get('nome', '')}</h2>
                <div class="cat-info">
                    <span class="cat-count">{n_itens_cat} itens</span>
                    <span class="badge {badge_cls}">{prioridade}</span>
                </div>
            </div>
            <div class="cat-body">
"""
        if cat.get("descricao"):
            html += f'                <div style="padding:8px 20px;font-size:0.85em;color:#666;font-style:italic;background:#f8f9fa">{cat["descricao"]}</div>\n'

        for item in cat.get("itens", []):
            crit = item.get("criticidade", "RECOMENDADA").upper()
            ref = item.get("referencia_normativa", "")
            detalhes = item.get("detalhes", "")
            aplic = item.get("aplicabilidade", "")

            html += f"""
                <div class="item" data-crit="{crit}">
                    <input type="checkbox" onchange="toggleItem(this)">
                    <div class="item-content">
                        <div class="item-desc">{item.get('descricao', '')}</div>
"""
            if detalhes:
                html += f'                        <div class="item-details">{detalhes}</div>\n'
            if ref:
                html += f'                        <div class="item-ref">📋 {ref}</div>\n'
            if aplic:
                html += f'                        <div class="item-aplic">📌 {aplic}</div>\n'

            html += f"""                    </div>
                    <span class="crit crit-{crit}">{crit}</span>
                </div>
"""
        html += "            </div>\n        </div>\n"

    # Observações gerais
    obs = checklist.get("observacoes_gerais", [])
    if obs:
        html += '        <div class="obs-gerais"><h3>📝 Observações Gerais</h3><ul>\n'
        for o in obs:
            html += f"            <li>{o}</li>\n"
        html += "        </ul></div>\n"

    html += f"""
        <div class="footer">
            Checklist Normativo de Revisão CSD-MPD – Gerado automaticamente por análise de IA<br>
            Modelo: {MODELO_IA} | {datetime.now().strftime('%d/%m/%Y %H:%M')} |
            {checklist.get('baseado_em', '')}
        </div>
    </div>

    <script>
        const total = {total_itens};
        let checkedCount = 0;
        let currentFilter = 'ALL';

        function toggleItem(cb) {{
            const item = cb.closest('.item');
            if (cb.checked) {{ item.classList.add('checked'); checkedCount++; }}
            else {{ item.classList.remove('checked'); checkedCount--; }}
            updateProgress();
            if (currentFilter === 'UNCHECKED') filterItems('UNCHECKED');
        }}

        function toggleCategory(header) {{
            const body = header.nextElementSibling;
            body.style.display = body.style.display === 'none' ? '' : 'none';
        }}

        function updateProgress() {{
            const pct = total > 0 ? (checkedCount / total * 100) : 0;
            document.getElementById('progress').style.width = pct.toFixed(1) + '%';
            document.getElementById('progressText').textContent =
                checkedCount + ' / ' + total + ' verificados (' + pct.toFixed(0) + '%)';
        }}

        function filterItems(filter) {{
            currentFilter = filter;
            // Update buttons
            document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
            event.target.classList.add('active');

            document.querySelectorAll('.item').forEach(item => {{
                const crit = item.dataset.crit;
                const isChecked = item.querySelector('input').checked;

                if (filter === 'ALL') item.classList.remove('hidden');
                else if (filter === 'UNCHECKED') {{
                    item.classList.toggle('hidden', isChecked);
                }} else {{
                    item.classList.toggle('hidden', crit !== filter);
                }}
            }});

            // Show/hide empty categories
            document.querySelectorAll('.category').forEach(cat => {{
                const visibleItems = cat.querySelectorAll('.item:not(.hidden)');
                cat.style.display = visibleItems.length > 0 ? '' : 'none';
            }});
        }}
    </script>
</body>
</html>"""

    CHECKLIST_HTML.write_text(html, encoding="utf-8")
    log.info(f"  HTML salvo: {CHECKLIST_HTML}")


# ==========================================================================
#  PIPELINE PRINCIPAL
# ==========================================================================

def main():
    global MODELO_IA, MODELOS_FALLBACK

    parser = argparse.ArgumentParser(
        description="Gerador de Checklist Normativo para Revisão de SEQOPs (CSD-MPD)"
    )
    parser.add_argument("--skip-extract", action="store_true",
                        help="Pular extração (usar JSON intermediário já salvo)")
    parser.add_argument("--corpus", type=str, default=str(CORPUS_DIR),
                        help=f"Diretório do corpus (default: {CORPUS_DIR})")
    parser.add_argument("--modelo", type=str, default=MODELO_IA,
                        help=f"Modelo de IA (default: {MODELO_IA})")
    args = parser.parse_args()

    if args.modelo != MODELO_IA:
        MODELO_IA = args.modelo
        MODELOS_FALLBACK = [MODELO_IA] + [m for m in MODELOS_FALLBACK if m != MODELO_IA]

    corpus_dir = Path(args.corpus)

    log.info("=" * 70)
    log.info("GERADOR DE CHECKLIST NORMATIVO – Revisão de SEQOPs CSD-MPD")
    log.info(f"Data/hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    log.info(f"Modelo IA: {MODELO_IA}")
    log.info(f"Corpus: {corpus_dir}")
    log.info("=" * 70)

    # ── Fase 1: Carregar corpus ──────────────────────────────────────
    corpus = fase1_carregar_corpus(corpus_dir)
    if not corpus:
        log.error("Corpus vazio! Verifique o diretório.")
        return

    # ── Fase 2: Extrair requisitos ───────────────────────────────────
    if args.skip_extract and REQUISITOS_JSON.exists():
        log.info(f"Carregando requisitos de: {REQUISITOS_JSON}")
        requisitos_por_lote = _carregar_json(REQUISITOS_JSON)
    else:
        requisitos_por_lote = fase2_extrair_requisitos(corpus, corpus_dir)

    if not requisitos_por_lote:
        log.error("Nenhum requisito extraído! Abortando.")
        return

    total_req = sum(len(r.get("requisitos", [])) for r in requisitos_por_lote)
    log.info(f"\nRequisitos disponíveis: {total_req} de {len(requisitos_por_lote)} lotes")

    # ── Fase 3: Consolidar checklist ─────────────────────────────────
    checklist = fase3_consolidar(requisitos_por_lote)

    if not checklist.get("categorias"):
        log.error("Checklist vazio após consolidação!")
        return

    # ── Fase 4: Gerar saídas ─────────────────────────────────────────
    fase4_gerar_excel(checklist)
    fase4_gerar_html(checklist)

    # ── Resumo final ─────────────────────────────────────────────────
    n_cats = len(checklist.get("categorias", []))
    n_itens = sum(len(c.get("itens", [])) for c in checklist.get("categorias", []))

    log.info("\n" + "=" * 70)
    log.info("CHECKLIST NORMATIVO GERADO COM SUCESSO!")
    log.info(f"  Categorias: {n_cats}")
    log.info(f"  Itens:      {n_itens}")
    log.info(f"  Excel:      {CHECKLIST_XLSX}")
    log.info(f"  HTML:       {CHECKLIST_HTML}")
    log.info(f"  JSON:       {CHECKLIST_JSON}")
    log.info("=" * 70)


if __name__ == "__main__":
    main()

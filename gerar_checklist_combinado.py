"""
Gerador de Checklist COMBINADO para Revisão de SEQOPs (CSD-MPD) — Abordagem 3
===============================================================================
Combina:
  - Checklist normativo (68 itens de 25 documentos Petrobras) — Abordagem 2
  - Comentários reais dos revisores CSD-MPD (coletados via Selenium) — Abordagem 1
  - Lições aprendidas MPD (pré-processadas de CSV Lessons Petrobras) — Novo

O resultado é um checklist unificado com:
  • Priorização por frequência real de cobrança pelos revisores
  • Exemplos de comentários reais vinculados a cada item
  • Evidências de lições aprendidas (alertas técnicos, boas práticas, observações)
  • Itens NOVOS identificados exclusivamente nos comentários (não cobertos pelas normas)
  • Score de relevância combinando criticidade normativa + frequência prática + lições

Fases:
  1. Carregar dados existentes: checklist normativo + comentários coletados
  2. (Opcional) Enriquecer: coletar mais comentários via Selenium
  3. Cruzamento IA: mapear comentários → itens normativos + identificar lacunas
  4. Gerar checklist combinado: Excel + HTML + JSON

Uso:
    C:\\SharedPython\\venv\\Scripts\\python.exe gerar_checklist_combinado.py [opções]

    Opções:
      --enrich              Enriquecer: coletar mais comentários via Selenium
      --max-urls N          Limite de URLs para enriquecer (0 = todas)
      --skip-enrich         Usar comentários já salvos (comentarios_mpd.json)
      --modelo MODELO       Modelo de IA (default: gpt-4o)
      --normativo JSON      Caminho do checklist normativo (default: auto)
      --comentarios JSON    Caminho dos comentários (default: auto)
      --licoes JSON         Caminho do JSON de lições aprendidas (default: auto)
      --sem-licoes          Ignorar lições aprendidas

Pré-requisitos:
    - checklist_normativo_resultado.json   (saída da Abordagem 2)
    - comentarios_mpd.json                 (saída da Abordagem 1, ou será coletado)
    - licoes_aprendidas_mpd.json           (saída de preprocessar_licoes.py, opcional)
    - resultado_aprovacoes_mpd.xlsx        (para enriquecimento, se necessário)
    - Acesso à API de IA Petrobras
"""

# ── Re-launch com o venv compartilhado ──────────────────────────────────────
import subprocess
import sys
from pathlib import Path

_SHARED_PYTHON = Path(r"C:\SharedPython\venv\Scripts\python.exe")
if _SHARED_PYTHON.exists() and Path(sys.executable).resolve() != _SHARED_PYTHON.resolve():
    print(f"[LAUNCH] Relançando com {_SHARED_PYTHON}")
    result = subprocess.run([str(_SHARED_PYTHON)] + sys.argv, check=False)
    sys.exit(result.returncode)

# ── Imports ─────────────────────────────────────────────────────────────────
import json
import logging
import os
import re
import time
import argparse
from datetime import datetime
from textwrap import dedent

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Configuração ────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent

# Entradas
NORMATIVO_JSON = SCRIPT_DIR / "checklist_normativo_resultado.json"
COMENTARIOS_JSON = SCRIPT_DIR / "comentarios_mpd.json"
LICOES_JSON = SCRIPT_DIR / "licoes_aprendidas_mpd.json"
ENTRADA_XLSX = SCRIPT_DIR / "resultado_aprovacoes_mpd.xlsx"

# Saídas
COMBINADO_JSON = SCRIPT_DIR / "checklist_combinado_resultado.json"
COMBINADO_XLSX = SCRIPT_DIR / "checklist_combinado_mpd.xlsx"
COMBINADO_HTML = SCRIPT_DIR / "checklist_combinado_mpd.html"

# API de IA
API_BASE = "https://apid.petrobras.com.br/ia/generativos/v1/litellm-chat-petrobras/litellm"
API_KEY = "b320f1f58c9743e9a74048ce64717c89"
MODELO_IA = "gpt-4o"

MODELOS_FALLBACK = [
    MODELO_IA,
    "gpt-4.1",
    "gpt-5-mini",
    "claude-sonnet-4-5",
    "claude-3-7-sonnet",
]

# Selenium
DEBUGGING_PORT = 9222
WAIT_TIMEOUT = 15

# ── Logging ─────────────────────────────────────────────────────────────────

LOG_FILE = SCRIPT_DIR / "checklist_combinado.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(
            open(sys.stdout.fileno(), mode='w', encoding='utf-8', closefd=False)
        ),
    ],
)
log = logging.getLogger("checklist_combinado")

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ==========================================================================
#  FUNÇÕES AUXILIARES
# ==========================================================================

def _salvar_json(dados, caminho: Path):
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    log.info(f"  JSON salvo: {caminho}")


def _carregar_json(caminho: Path):
    if caminho.exists():
        with open(caminho, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def _chamar_ia(mensagens: list[dict], max_tokens: int = 8000, temperature: float = 0.2) -> str:
    """Chama a API de IA generativa. Tenta múltiplos modelos em caso de falha."""
    headers = {
        "Content-Type": "application/json",
        "api-key": API_KEY,
    }
    payload = {
        "messages": mensagens,
        "max_tokens": max_tokens,
        "temperature": temperature,
    }

    for modelo in MODELOS_FALLBACK:
        url = f"{API_BASE}/engines/{modelo}/chat/completions"
        log.info(f"    Chamando modelo: {modelo}")

        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=300, verify=False)

            if resp.status_code != 200:
                log.warning(f"    Modelo {modelo}: HTTP {resp.status_code} – {resp.text[:300]}")
                continue

            data = resp.json()
            choices = data.get("choices", [])
            if not choices:
                log.warning(f"    Modelo {modelo}: Resposta sem choices")
                continue

            content = choices[0].get("message", {}).get("content", "")
            usage = data.get("usage", {})
            log.info(f"    ✓ {modelo} – tokens: prompt={usage.get('prompt_tokens')}, "
                     f"completion={usage.get('completion_tokens')}")
            return content

        except requests.exceptions.Timeout:
            log.warning(f"    Modelo {modelo}: Timeout (300s)")
        except Exception as e:
            log.warning(f"    Modelo {modelo}: Erro – {e}")

    raise RuntimeError(f"Todos os modelos falharam: {MODELOS_FALLBACK}")


def _extrair_json(texto: str):
    """Extrai JSON de uma resposta de texto (pode ter markdown ou texto ao redor)."""
    m = re.search(r'```(?:json)?\s*\n?([\s\S]*?)\n?```', texto)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass

    for pattern in [r'\{[\s\S]*\}', r'\[[\s\S]*\]']:
        m = re.search(pattern, texto)
        if m:
            try:
                return json.loads(m.group())
            except json.JSONDecodeError:
                continue

    return None


# ==========================================================================
#  FASE 1: CARREGAR DADOS EXISTENTES
# ==========================================================================

def fase1_carregar_dados(normativo_path: Path, comentarios_path: Path, licoes_path: Path = None) -> tuple[dict, list, dict]:
    """Carrega o checklist normativo, os comentários e as lições aprendidas.

    Returns:
        (checklist_normativo, dados_comentarios, dados_licoes)
    """
    log.info("=" * 70)
    log.info("FASE 1: Carregando dados existentes")
    log.info("=" * 70)

    # ── Checklist normativo ──────────────────────────────────────────
    checklist_norm = _carregar_json(normativo_path)
    if not checklist_norm:
        log.error(f"  Checklist normativo não encontrado: {normativo_path}")
        log.error("  Execute primeiro: gerar_checklist_normas.py (Abordagem 2)")
        return {}, [], {}

    n_cats = len(checklist_norm.get("categorias", []))
    n_itens = sum(len(c.get("itens", [])) for c in checklist_norm.get("categorias", []))
    log.info(f"  Checklist normativo: {n_cats} categorias, {n_itens} itens")

    # ── Comentários ──────────────────────────────────────────────────
    dados_comentarios = _carregar_json(comentarios_path) or []

    if dados_comentarios:
        total_seqops = len(dados_comentarios)
        total_com = sum(d.get("total_comentarios", 0) for d in dados_comentarios)
        total_mpd = sum(d.get("total_mpd", 0) for d in dados_comentarios)
        log.info(f"  Comentários: {total_seqops} SEQOPs, {total_com} comentários ({total_mpd} MPD)")
    else:
        log.warning(f"  Nenhum comentário encontrado: {comentarios_path}")
        log.info("  Use --enrich para coletar comentários via Selenium")

    # ── Lições Aprendidas ────────────────────────────────────────────
    dados_licoes = {}
    if licoes_path and licoes_path.exists():
        dados_licoes = _carregar_json(licoes_path) or {}
        if dados_licoes:
            n_cats_lic = len(dados_licoes.get("categorias", []))
            n_lic = sum(c.get("total_licoes", 0) for c in dados_licoes.get("categorias", []))
            log.info(f"  Lições aprendidas: {n_cats_lic} categorias, {n_lic} lições")
            log.info(f"    Fonte CSV: {dados_licoes.get('total_registros_csv', '?')} registros, {dados_licoes.get('total_relevantes_mpd', '?')} relevantes MPD")
    elif licoes_path:
        log.warning(f"  Lições aprendidas não encontradas: {licoes_path}")
        log.info("  Execute preprocessar_licoes.py para gerar o arquivo")
    else:
        log.info("  Lições aprendidas: desabilitado (use --licoes ou remova --sem-licoes)")

    return checklist_norm, dados_comentarios, dados_licoes


# ==========================================================================
#  FASE 2 (OPCIONAL): ENRIQUECER COMENTÁRIOS VIA SELENIUM
# ==========================================================================

def _criar_driver():
    """Cria ou conecta ao Edge."""
    from selenium import webdriver
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.common.exceptions import WebDriverException

    try:
        log.info(f"Conectando ao Edge na porta {DEBUGGING_PORT}...")
        opts = EdgeOptions()
        opts.debugger_address = f"127.0.0.1:{DEBUGGING_PORT}"
        d = webdriver.Edge(options=opts)
        log.info("Conectado ao Edge existente.")
        return d
    except WebDriverException:
        log.info("Abrindo nova instância do Edge...")

    opts = EdgeOptions()
    profile = Path(os.environ.get("LOCALAPPDATA", "")) / "SeqopScraper_Edge"
    profile.mkdir(parents=True, exist_ok=True)
    opts.add_argument(f"--user-data-dir={profile}")
    opts.add_argument("--profile-directory=Default")
    opts.add_argument("--start-maximized")
    opts.add_argument(f"--remote-debugging-port={DEBUGGING_PORT}")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    d = webdriver.Edge(options=opts)
    log.info("Edge aberto (perfil dedicado).")
    return d


def _extrair_comentarios_pagina(driver) -> list[dict]:
    """Extrai TODOS os comentários da página de detalhe da SEQOP."""
    from selenium.webdriver.common.by import By
    from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException

    comentarios = []
    try:
        grupos = driver.find_elements(By.CSS_SELECTOR, "div.grupoComentario")
        for grupo in grupos:
            try:
                texto = ""
                try:
                    card = grupo.find_element(By.CSS_SELECTOR, "div.card-text")
                    texto = card.text.strip()
                except Exception:
                    pass

                autor = ""
                tipo_csd = ""
                versao = ""
                try:
                    autor_div = grupo.find_element(
                        By.CSS_SELECTOR, "div.emLinhaAutor.autor-versao"
                    )
                    spans = autor_div.find_elements(By.TAG_NAME, "span")
                    if len(spans) >= 1:
                        nome_raw = spans[0].text.strip()
                        autor = re.sub(
                            r"\s+h[aá]\s+\d+\s+\w+$", "", nome_raw,
                            flags=re.IGNORECASE,
                        ).strip()
                    if len(spans) >= 2:
                        info = spans[1].text.strip()
                        m_tipo = re.match(r"(CSD-\w+|Fiscal)\s*-\s*Vers[aã]o:\s*(\d+)", info, re.IGNORECASE)
                        if m_tipo:
                            tipo_csd = m_tipo.group(1)
                            versao = m_tipo.group(2)
                        else:
                            tipo_csd = info
                except Exception:
                    pass

                respostas = []
                try:
                    replies = grupo.find_elements(By.CSS_SELECTOR, "div.replyComments")
                    for reply in replies:
                        try:
                            reply_text = ""
                            try:
                                rc = reply.find_element(By.CSS_SELECTOR, "div.card-text")
                                reply_text = rc.text.strip()
                            except Exception:
                                pass
                            reply_autor = ""
                            try:
                                ra = reply.find_element(By.CSS_SELECTOR, "div.emLinhaAutor.autor-versao")
                                rs = ra.find_elements(By.TAG_NAME, "span")
                                if rs:
                                    reply_autor = re.sub(
                                        r"\s+h[aá]\s+\d+\s+\w+$", "",
                                        rs[0].text.strip(), flags=re.IGNORECASE,
                                    ).strip()
                            except Exception:
                                pass
                            if reply_text:
                                respostas.append({"autor": reply_autor, "texto": reply_text})
                        except Exception:
                            continue
                except Exception:
                    pass

                if texto or respostas:
                    comentarios.append({
                        "autor": autor,
                        "tipo_csd": tipo_csd,
                        "versao": versao,
                        "texto": texto,
                        "respostas": respostas,
                    })

            except (NoSuchElementException, StaleElementReferenceException):
                continue

    except Exception as e:
        log.warning(f"  Erro ao extrair comentários: {e}")

    return comentarios


def fase2_enriquecer(dados_existentes: list[dict], max_urls: int = 0) -> list[dict]:
    """Coleta mais comentários via Selenium, preservando os já coletados.

    Returns:
        lista atualizada de dados enriquecidos
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    log.info("=" * 70)
    log.info("FASE 2: Enriquecendo comentários via Selenium")
    log.info("=" * 70)

    if not ENTRADA_XLSX.exists():
        log.error(f"  Arquivo não encontrado: {ENTRADA_XLSX}")
        log.error("  Execute primeiro o scraper principal (scraper_seqop.py)")
        return dados_existentes

    df = pd.read_excel(ENTRADA_XLSX, engine="openpyxl")
    log.info(f"  Excel carregado: {len(df)} linhas")

    # Extrair URLs únicas
    urls_unicas = []
    urls_vistas = set()
    for _, row in df.iterrows():
        url = str(row.get("URL Fonte", ""))
        if url and url.startswith("http") and url not in urls_vistas:
            urls_vistas.add(url)
            urls_unicas.append({
                "url": url,
                "poco": str(row.get("Poço", "")),
                "titulo": str(row.get("Título", "")),
                "seq_id": str(row.get("Seq ID", "")),
            })

    log.info(f"  URLs únicas totais: {len(urls_unicas)}")

    # Filtrar URLs já coletadas
    urls_ja_coletadas = {d.get("url", "") for d in dados_existentes}
    urls_pendentes = [u for u in urls_unicas if u["url"] not in urls_ja_coletadas]

    if max_urls > 0:
        urls_pendentes = urls_pendentes[:max_urls]

    log.info(f"  URLs já coletadas: {len(urls_ja_coletadas)}")
    log.info(f"  URLs pendentes: {len(urls_pendentes)}")

    if not urls_pendentes:
        log.info("  Nenhuma URL pendente. Fase 2 pulada.")
        return dados_existentes

    driver = _criar_driver()

    # Aguardar login
    driver.get("https://csdpocos.petrobras.com.br/seqop")
    time.sleep(3)

    try:
        menus = driver.find_elements(By.CSS_SELECTOR, "aside.menu")
        if not menus:
            log.info("Aguardando login manual...")
            for _ in range(120):
                time.sleep(5)
                menus = driver.find_elements(By.CSS_SELECTOR, "aside.menu")
                if menus:
                    break
    except Exception:
        pass

    time.sleep(2)
    novos_dados = list(dados_existentes)  # cópia

    for idx, info in enumerate(urls_pendentes, 1):
        url = info["url"]
        log.info(f"  [{idx}/{len(urls_pendentes)}] {info['poco']} – {info['titulo'][:50]}")

        try:
            driver.get(url)
            time.sleep(3)

            try:
                WebDriverWait(driver, 10).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
            except Exception:
                pass
            time.sleep(1)

            comentarios = _extrair_comentarios_pagina(driver)

            titulo_pagina = info["titulo"]
            try:
                h5s = driver.find_elements(By.CSS_SELECTOR, "div.SideBar h5")
                if len(h5s) >= 2:
                    titulo_pagina = h5s[1].text.strip() or titulo_pagina
            except Exception:
                pass

            dados = {
                "url": url,
                "poco": info["poco"],
                "titulo": titulo_pagina,
                "seq_id": info["seq_id"],
                "comentarios": comentarios,
                "conteudo_seqop": "",
                "comentarios_mpd": [c for c in comentarios if "MPD" in c.get("tipo_csd", "").upper()],
                "total_comentarios": len(comentarios),
                "total_mpd": sum(1 for c in comentarios if "MPD" in c.get("tipo_csd", "").upper()),
            }
            novos_dados.append(dados)

            log.info(f"    → {len(comentarios)} comentários ({dados['total_mpd']} MPD)")

        except Exception as e:
            log.error(f"    Erro: {e}")
            novos_dados.append({
                "url": url, "poco": info["poco"], "titulo": info["titulo"],
                "seq_id": info["seq_id"], "comentarios": [], "conteudo_seqop": "",
                "comentarios_mpd": [], "total_comentarios": 0, "total_mpd": 0,
                "erro": str(e),
            })

        # Salvar progresso a cada 10
        if idx % 10 == 0:
            _salvar_json(novos_dados, COMENTARIOS_JSON)
            log.info(f"  Progresso salvo ({idx}/{len(urls_pendentes)})")

    _salvar_json(novos_dados, COMENTARIOS_JSON)
    log.info(f"FASE 2 completa: {len(novos_dados)} SEQOPs enriquecidas")
    return novos_dados


# ==========================================================================
#  FASE 3: CRUZAMENTO IA – Comentários × Checklist Normativo
# ==========================================================================

def _agregar_comentarios(dados_comentarios: list[dict]) -> list[dict]:
    """Agrega todos os comentários relevantes (MPD + menções a MPD) em lista plana."""
    comentarios = []

    for seq in dados_comentarios:
        # Comentários MPD diretos
        for c in seq.get("comentarios_mpd", []):
            if c.get("texto"):
                texto_completo = c["texto"]
                if c.get("respostas"):
                    for r in c["respostas"]:
                        texto_completo += f"\n[Resposta de {r.get('autor', '?')}]: {r['texto']}"
                comentarios.append({
                    "poco": seq.get("poco", ""),
                    "titulo": seq.get("titulo", ""),
                    "versao": c.get("versao", ""),
                    "autor": c.get("autor", ""),
                    "tipo_csd": c.get("tipo_csd", "CSD-MPD"),
                    "texto": texto_completo,
                })

        # Comentários de outros CSDs que mencionam MPD/SBP
        for c in seq.get("comentarios", []):
            if "MPD" not in c.get("tipo_csd", "").upper():
                texto_upper = (c.get("texto", "") or "").upper()
                if any(kw in texto_upper for kw in ["MPD", "SBP", "BEARING", "FMCD", "PMCD"]):
                    texto_completo = c.get("texto", "")
                    if c.get("respostas"):
                        for r in c["respostas"]:
                            texto_completo += f"\n[Resposta de {r.get('autor', '?')}]: {r['texto']}"
                    comentarios.append({
                        "poco": seq.get("poco", ""),
                        "titulo": seq.get("titulo", ""),
                        "versao": c.get("versao", ""),
                        "autor": c.get("autor", ""),
                        "tipo_csd": c.get("tipo_csd", ""),
                        "texto": texto_completo,
                    })

    return comentarios


def fase3_cruzamento(checklist_norm: dict, dados_comentarios: list[dict], dados_licoes: dict = None) -> dict:
    """Usa IA para cruzar comentários com itens normativos, enriquecido com lições aprendidas.

    Executado em 3 passos:
      A. Mapear comentários → categorias/itens normativos
      B. Enriquecer cada grupo de categorias com evidência dos comentários + lições
      C. Identificar itens novos (gaps) exclusivos dos comentários
    """
    log.info("=" * 70)
    log.info("FASE 3: Cruzamento IA – Comentários × Normativo × Lições Aprendidas")
    log.info("=" * 70)

    # ── Agregar comentários ──────────────────────────────────────────
    comentarios = _agregar_comentarios(dados_comentarios)
    log.info(f"  Comentários agregados: {len(comentarios)}")

    if not comentarios:
        log.warning("  Nenhum comentário disponível! Retornando checklist normativo puro.")
        # Retornar normativo com flag de sem-comentários
        for cat in checklist_norm.get("categorias", []):
            for item in cat.get("itens", []):
                item["evidencia_comentarios"] = "Sem comentários coletados"
                item["frequencia_real"] = 0
                item["score_combinado"] = _calcular_score(item.get("criticidade", ""), 0)
                item["origem"] = "NORMATIVO"
        checklist_norm["_combinado"] = True
        checklist_norm["_total_comentarios"] = 0
        return checklist_norm

    # ── Preparar texto dos comentários (limitado) ────────────────────
    blocos_comentarios = []
    chars_total = 0
    for i, c in enumerate(comentarios, 1):
        bloco = (
            f"[C{i}] Poço: {c['poco']} | SEQOP: {c['titulo'][:60]} | "
            f"Autor: {c['autor']} | CSD: {c['tipo_csd']} | v{c['versao']}\n"
            f"{c['texto'][:500]}\n---"
        )
        if chars_total + len(bloco) > 60000:
            break
        blocos_comentarios.append(bloco)
        chars_total += len(bloco)

    log.info(f"  Usando {len(blocos_comentarios)} comentários ({chars_total:,} chars)")

    # ── Preparar resumo do checklist normativo ───────────────────────
    resumo_normativo = []
    for cat in checklist_norm.get("categorias", []):
        cat_id = cat.get("id", "")
        cat_nome = cat.get("nome", "")
        itens_resumo = []
        for item in cat.get("itens", []):
            itens_resumo.append(
                f"  {item.get('id', '')}: {item.get('descricao', '')[:150]}"
            )
        resumo_normativo.append(f"\n{cat_id} – {cat_nome}:\n" + "\n".join(itens_resumo))

    texto_normativo = "\n".join(resumo_normativo)
    texto_comentarios = "\n".join(blocos_comentarios)

    # ── Agregar lições aprendidas por categoria ──────────────────────
    licoes_por_cat = {}  # cat_id → [list of lesson summaries]
    total_licoes_usadas = 0
    if dados_licoes and dados_licoes.get("categorias"):
        for cat_lic in dados_licoes["categorias"]:
            cat_id = cat_lic.get("cat_id", "")
            licoes_cat = cat_lic.get("licoes", [])
            if not licoes_cat:
                continue
            # Selecionar as top lições por prioridade (max 15 por categoria para limitar tokens)
            resumos = []
            for lic in licoes_cat[:15]:
                tipo_sigla = {"Lição Aprendida": "LA", "Alerta Técnico": "AT",
                              "Boas Práticas": "BP", "Observação Poço": "OP"}.get(lic.get("tipo", ""), "??")
                titulo = lic.get("titulo", "")[:100]
                recomendacao = lic.get("recomendacao", "")[:300]
                evidencia = lic.get("evidencia", "")[:200]
                poco = lic.get("poco", "")
                resumo = f"[{lic.get('id','')}/{tipo_sigla}] {titulo}"
                if recomendacao:
                    resumo += f"\n  Recomendação: {recomendacao}"
                if evidencia:
                    resumo += f"\n  Evidência: {evidencia}"
                if poco:
                    resumo += f" (Poço: {poco})"
                resumos.append(resumo)
            licoes_por_cat[cat_id] = resumos
            total_licoes_usadas += len(resumos)
        log.info(f"  Lições aprendidas: {total_licoes_usadas} lições em {len(licoes_por_cat)} categorias")
    else:
        log.info("  Lições aprendidas: nenhuma disponível")

    # ── Passo A: Mapeamento comentários → itens normativos ───────────
    log.info("  Passo A: Mapeando comentários → itens normativos...")

    prompt_mapeamento = dedent(f"""\
    Você é um especialista em perfuração MPD na Petrobras.

    Abaixo está um CHECKLIST NORMATIVO com {sum(len(c.get('itens',[])) for c in checklist_norm.get('categorias',[]))} itens de verificação
    organizados em {len(checklist_norm.get('categorias',[]))} categorias (tipos de SEQOP), seguido de
    {len(blocos_comentarios)} COMENTÁRIOS REAIS de revisores CSD-MPD.

    TAREFA: Para cada comentário, identifique:
    1. Qual(is) item(ns) normativo(s) o comentário se relaciona (pelo ID, ex: CAT-01-003)
    2. Se o comentário traz um ponto que NÃO está no checklist normativo (LACUNA)

    CHECKLIST NORMATIVO:
    {texto_normativo}

    COMENTÁRIOS DOS REVISORES:
    {texto_comentarios}

    Responda em JSON:
    {{
      "mapeamentos": [
        {{
          "comentario_id": "C1",
          "itens_relacionados": ["CAT-01-003", "CAT-05-002"],
          "resumo": "Resumo do ponto cobrado pelo revisor",
          "eh_lacuna": false,
          "lacuna_descricao": ""
        }}
      ],
      "lacunas": [
        {{
          "descricao": "Item cobrado nos comentários mas ausente no checklist normativo",
          "comentarios_relacionados": ["C3", "C7"],
          "categoria_sugerida": "CAT-01",
          "criticidade_sugerida": "IMPORTANTE",
          "frequencia": 2
        }}
      ],
      "estatisticas": {{
        "itens_com_evidencia": 15,
        "itens_sem_evidencia": 53,
        "total_lacunas": 3,
        "comentarios_mapeados": {len(blocos_comentarios)},
        "comentarios_sem_match": 0
      }}
    }}
    """)

    resp_mapeamento = _chamar_ia([
        {"role": "system", "content": (
            "Você é um especialista em perfuração MPD da Petrobras que revisa SEQOPs. "
            "Analise com precisão. Responda em JSON válido, em português brasileiro."
        )},
        {"role": "user", "content": prompt_mapeamento},
    ], max_tokens=12000, temperature=0.15)

    mapeamento = _extrair_json(resp_mapeamento)
    if not mapeamento:
        log.warning("  Passo A: Falha ao parsear mapeamento. Usando fallback.")
        mapeamento = {"mapeamentos": [], "lacunas": [], "estatisticas": {}}
    else:
        n_map = len(mapeamento.get("mapeamentos", []))
        n_lac = len(mapeamento.get("lacunas", []))
        log.info(f"    ✓ {n_map} mapeamentos, {n_lac} lacunas identificadas")

    # ── Passo B: Enriquecer checklist com evidências ─────────────────
    log.info("  Passo B: Enriquecendo checklist com evidências dos comentários...")

    # Construir mapa de evidências: item_id → lista de comentário IDs + resumos
    evidencia_por_item = {}
    for m in mapeamento.get("mapeamentos", []):
        resumo = m.get("resumo", "")
        for item_id in m.get("itens_relacionados", []):
            if item_id not in evidencia_por_item:
                evidencia_por_item[item_id] = []
            evidencia_por_item[item_id].append({
                "comentario_id": m.get("comentario_id", ""),
                "resumo": resumo,
            })

    log.info(f"    Itens com evidências: {len(evidencia_por_item)}")

    # Agora, para cada grupo de categorias, pedir à IA para gerar a versão enriquecida
    # dividimos em 3 passos (como na Abordagem 2) para garantir completude

    categorias_norm = checklist_norm.get("categorias", [])

    # Distribuir categorias nos 3 passos
    n_cats = len(categorias_norm)
    split1 = min(4, n_cats)
    split2 = min(8, n_cats)
    grupos_cat = [
        ("A (Perfuração + Junta)", categorias_norm[:split1]),
        ("B (FP + Testes + Influxo)", categorias_norm[split1:split2]),
        ("C (Descidas de Cauda)", categorias_norm[split2:]),
    ]

    todas_categorias_enriquecidas = []

    for passo_label, cats_grupo in grupos_cat:
        if not cats_grupo:
            continue

        # Montar texto das categorias com suas evidências
        cats_texto = []
        for cat in cats_grupo:
            cat_id = cat.get("id", "")
            cat_nome = cat.get("nome", "")
            itens_texto = []
            for item in cat.get("itens", []):
                item_id = item.get("id", "")
                evidencias = evidencia_por_item.get(item_id, [])
                ev_str = ""
                if evidencias:
                    ev_resumos = [e["resumo"] for e in evidencias]
                    ev_str = f"\n      EVIDÊNCIA ({len(evidencias)} comentários): " + "; ".join(ev_resumos[:3])
                itens_texto.append(
                    f"    {item_id} ({item.get('criticidade', '?')}): {item.get('descricao', '')}"
                    f"\n      Detalhes: {item.get('detalhes', '')[:200]}"
                    f"\n      Ref: {item.get('referencia_normativa', '')}"
                    f"{ev_str}"
                )

            # Adicionar lições aprendidas relevantes para esta categoria
            licoes_cat_texto = ""
            licoes_desta_cat = licoes_por_cat.get(cat_id, [])
            if licoes_desta_cat:
                licoes_cat_texto = f"\n    LIÇÕES APRENDIDAS ({len(licoes_desta_cat)} registros do sistema Lessons):\n"
                for lic_resumo in licoes_desta_cat[:10]:
                    licoes_cat_texto += f"      {lic_resumo[:400]}\n"

            cats_texto.append(
                f"\n  {cat_id} – {cat_nome} ({cat.get('prioridade', '')})"
                f"\n  {cat.get('descricao', '')}\n"
                + "\n".join(itens_texto)
                + licoes_cat_texto
            )

        # Lacunas candidatas para este grupo de categorias
        lacunas_grupo = []
        for lac in mapeamento.get("lacunas", []):
            cat_sug = lac.get("categoria_sugerida", "")
            if any(cat.get("id", "") in cat_sug for cat in cats_grupo):
                lacunas_grupo.append(lac)

        lacunas_texto = ""
        if lacunas_grupo:
            lacunas_texto = "\n\nLACUNAS identificadas (itens cobrados nos comentários mas ausentes nas normas):\n"
            for lac in lacunas_grupo:
                lacunas_texto += (
                    f"  - {lac.get('descricao', '')} "
                    f"(criticidade: {lac.get('criticidade_sugerida', '?')}, "
                    f"freq: {lac.get('frequencia', 0)})\n"
                )

        prompt_enriquecer = dedent(f"""\
        Você é um especialista em perfuração MPD na Petrobras.

        Abaixo estão categorias de um checklist normativo com:
        - EVIDÊNCIAS DE COMENTÁRIOS REAIS de revisores CSD-MPD mapeadas a cada item
        - LIÇÕES APRENDIDAS do sistema Lessons Petrobras (alertas técnicos, boas práticas, observações de poço)

        TAREFA: Gere a versão ENRIQUECIDA dessas categorias:
        1. MANTENHA todos os itens normativos existentes, mas ENRIQUEÇA com:
           - "evidencia_comentarios": resumo da evidência real (comentários + lições aprendidas relevantes, ou "Sem evidência direta")
           - "frequencia_real": quantas vezes apareceu nos comentários de revisores (0 se não apareceu)
           - "score_combinado": ALTO (normativo CRITICA + freq > 0 OU tem lições/alertas técnicos), MEDIO (normativo IMPORTANTE ou freq > 0), BAIXO (RECOMENDADA + sem freq + sem lições)
           - "origem": "NORMATIVO" para itens existentes
           - "licoes_aprendidas": resumo curto das lições aprendidas relevantes ao item (ou "" se não houver)

        2. ADICIONE itens novos das lacunas E das lições aprendidas que cobrem aspectos não presentes no normativo:
           - "origem": "COMENTARIOS" para itens de revisores, "LICOES" para itens exclusivos das lições aprendidas
           - Coloque-os na categoria mais adequada
           - Cada lição aprendida relevante que cubra um aspecto NÃO presente no checklist normativo deve gerar item novo

        CATEGORIAS PARA ENRIQUECER:
        {"".join(cats_texto)}
        {lacunas_texto}

        Responda em JSON:
        {{
          "categorias": [
            {{
              "id": "CAT-XX",
              "nome": "...",
              "descricao": "...",
              "prioridade": "ALTA/MEDIA/BAIXA",
              "itens": [
                {{
                  "id": "CAT-XX-001",
                  "descricao": "Verificação...",
                  "detalhes": "...",
                  "referencia_normativa": "PE-XXXX §X.Y",
                  "criticidade": "CRITICA/IMPORTANTE/RECOMENDADA",
                  "aplicabilidade": "...",
                  "evidencia_comentarios": "Resumo da evidência real ou 'Sem evidência direta'",
                  "licoes_aprendidas": "Resumo das lições aprendidas relevantes ou ''",
                  "frequencia_real": 0,
                  "score_combinado": "ALTO/MEDIO/BAIXO",
                  "origem": "NORMATIVO ou COMENTARIOS ou LICOES"
                }}
              ]
            }}
          ]
        }}

        REGRAS OBRIGATÓRIAS:
        - Retorne EXATAMENTE {len(cats_grupo)} categorias (uma para cada categoria acima)
        - NÃO remova, mescle, ou omita NENHUMA categoria
        - NÃO remova, mescle, condense, ou omita NENHUM item normativo existente. Retorne TODOS os itens originais.
        - Se uma categoria tem N itens no normativo, a versão enriquecida deve ter NO MÍNIMO N itens (pode ter N+novos)
        - Mantenha IDs originais dos itens normativos
        - OBRIGATÓRIO: Cada lacuna listada DEVE ser adicionada como item novo com origem "COMENTARIOS"
        - Itens novos das lições aprendidas devem ter origem "LICOES"
        - Score ALTO = prioridade máxima na revisão (atenção redobrada)
        - Preserve referências normativas e detalhes técnicos
        - Mesmo categorias SEM evidência de comentários devem ser retornadas com todos os itens

        CONTAGEM MÍNIMA DE ITENS POR CATEGORIA (NÃO pode retornar menos):
        {chr(10).join(f"  - {c.get('id','?')} {c.get('nome','')}: mínimo {len(c.get('itens',[]))} itens" for c in cats_grupo)}
        """)

        n_itens_entrada = sum(len(c.get("itens", [])) for c in cats_grupo)
        log.info(f"    Passo {passo_label}: enriquecendo {len(cats_grupo)} categorias ({n_itens_entrada} itens mínimos)...")
        resp = _chamar_ia([
            {"role": "system", "content": (
                "Você é um especialista em perfuração MPD da Petrobras que revisa SEQOPs. "
                "Gere o checklist enriquecido em JSON válido, em português brasileiro. "
                "REGRA CRÍTICA: MANTENHA ABSOLUTAMENTE TODOS os itens normativos SEM EXCEÇÃO – "
                "não condense, mescle ou omita nenhum. "
                "ADICIONE os itens de lacuna (COMENTARIOS) e os novos de lições (LICOES). "
                "O número de itens na saída deve ser MAIOR OU IGUAL ao da entrada."
            )},
            {"role": "user", "content": prompt_enriquecer},
        ], max_tokens=16000, temperature=0.15)

        resultado = _extrair_json(resp)
        if resultado and "categorias" in resultado:
            cats = resultado["categorias"]
            n = sum(len(c.get("itens", [])) for c in cats)
            log.info(f"    ✓ {len(cats)} categorias, {n} itens (enriquecidas)")

            # ── Safeguard: recuperar categorias que a IA omitiu ──────
            nomes_retornados = {c.get("nome", "").lower().strip() for c in cats}
            ids_retornados = {c.get("id", "").upper().strip() for c in cats}
            for cat_orig in cats_grupo:
                nome_orig = cat_orig.get("nome", "").lower().strip()
                id_orig = cat_orig.get("id", "").upper().strip()
                # Verificar se esta categoria foi retornada (por nome ou ID)
                encontrada = (
                    nome_orig in nomes_retornados
                    or id_orig in ids_retornados
                    or any(nome_orig in n for n in nomes_retornados)
                )
                if not encontrada:
                    log.warning(f"    ⚠ Categoria '{cat_orig.get('nome','')}' omitida pela IA – recuperando do normativo")
                    cat_recuperada = dict(cat_orig)  # cópia
                    cat_recuperada["itens"] = []
                    for item in cat_orig.get("itens", []):
                        item_enriquecido = dict(item)
                        evs = evidencia_por_item.get(item.get("id", ""), [])
                        if evs:
                            item_enriquecido["evidencia_comentarios"] = "; ".join(e["resumo"] for e in evs[:3])
                            item_enriquecido["frequencia_real"] = len(evs)
                        else:
                            item_enriquecido["evidencia_comentarios"] = "Sem evidência direta nos comentários coletados"
                            item_enriquecido["frequencia_real"] = 0
                        item_enriquecido["score_combinado"] = _calcular_score(
                            item.get("criticidade", ""), item_enriquecido["frequencia_real"]
                        )
                        item_enriquecido["origem"] = "NORMATIVO"
                        cat_recuperada["itens"].append(item_enriquecido)
                    cats.append(cat_recuperada)
                    log.info(f"      → Recuperada com {len(cat_recuperada['itens'])} itens")

            # ── Safeguard item-level: recuperar itens normativos omitidos dentro de categorias ──
            for cat_retornada in cats:
                cat_id_ret = cat_retornada.get("id", "").upper().strip()
                cat_nome_ret = cat_retornada.get("nome", "").lower().strip()
                # Encontrar a categoria original correspondente
                cat_orig_match = None
                for cat_orig in cats_grupo:
                    if (cat_orig.get("id", "").upper().strip() == cat_id_ret
                            or cat_orig.get("nome", "").lower().strip() == cat_nome_ret
                            or cat_nome_ret in cat_orig.get("nome", "").lower()):
                        cat_orig_match = cat_orig
                        break
                if not cat_orig_match:
                    continue

                # Coletar descrições dos itens retornados (normalizado) para comparação
                itens_ret_descs = set()
                for it in cat_retornada.get("itens", []):
                    desc = (it.get("descricao", "") or "").lower().strip()[:80]
                    itens_ret_descs.add(desc)

                n_orig = len(cat_orig_match.get("itens", []))
                n_ret = len(cat_retornada.get("itens", []))
                itens_faltantes = []

                for item_orig in cat_orig_match.get("itens", []):
                    desc_orig = (item_orig.get("descricao", "") or "").lower().strip()[:80]
                    # Procurar match exato ou parcial nos retornados
                    encontrado = (
                        desc_orig in itens_ret_descs
                        or any(desc_orig[:40] in d for d in itens_ret_descs)
                        or any(d[:40] in desc_orig for d in itens_ret_descs if len(d) > 10)
                    )
                    if not encontrado:
                        itens_faltantes.append(item_orig)

                if itens_faltantes:
                    log.warning(f"    ⚠ {cat_retornada.get('id','?')}: {len(itens_faltantes)} itens normativos omitidos pela IA – recuperando")
                    for item in itens_faltantes:
                        item_enriquecido = dict(item)
                        evs = evidencia_por_item.get(item.get("id", ""), [])
                        if evs:
                            item_enriquecido["evidencia_comentarios"] = "; ".join(e["resumo"] for e in evs[:3])
                            item_enriquecido["frequencia_real"] = len(evs)
                        else:
                            item_enriquecido["evidencia_comentarios"] = "Sem evidência direta nos comentários coletados"
                            item_enriquecido["frequencia_real"] = 0
                        # Verificar se há lições para este item (licoes_por_cat contém strings formatadas)
                        licoes_cat = licoes_por_cat.get(cat_id_ret, [])
                        if licoes_cat:
                            desc_item = (item.get("descricao", "") or "").lower()
                            keywords = [kw for kw in desc_item.split() if len(kw) > 3][:4]
                            licoes_rel = [l for l in licoes_cat[:5] if any(
                                kw in l.lower() for kw in keywords
                            )]
                            if licoes_rel:
                                # Extrair recomendação das strings formatadas
                                resumos = []
                                for l in licoes_rel[:2]:
                                    linhas = l.split("\n")
                                    rec_line = next((ln.strip() for ln in linhas if ln.strip().startswith("Recomendação:")), "")
                                    rec = rec_line.replace("Recomendação:", "").strip() if rec_line else linhas[0][:100]
                                    resumos.append(rec[:100])
                                item_enriquecido["licoes_aprendidas"] = "; ".join(resumos)
                        item_enriquecido.setdefault("licoes_aprendidas", "")
                        item_enriquecido["score_combinado"] = _calcular_score(
                            item.get("criticidade", ""), item_enriquecido["frequencia_real"]
                        )
                        item_enriquecido["origem"] = "NORMATIVO"
                        cat_retornada["itens"].append(item_enriquecido)
                    log.info(f"      → Recuperados {len(itens_faltantes)} itens em {cat_retornada.get('id','?')}")

            todas_categorias_enriquecidas.extend(cats)
        else:
            log.warning(f"    ✗ Falha no passo {passo_label}. Usando normativo puro.")
            for cat in cats_grupo:
                cat_recuperada = dict(cat)
                cat_recuperada["itens"] = []
                for item in cat.get("itens", []):
                    item_enriquecido = dict(item)
                    evs = evidencia_por_item.get(item.get("id", ""), [])
                    if evs:
                        item_enriquecido["evidencia_comentarios"] = "; ".join(e["resumo"] for e in evs[:3])
                        item_enriquecido["frequencia_real"] = len(evs)
                    else:
                        item_enriquecido["evidencia_comentarios"] = "Sem evidência (falha na análise)"
                        item_enriquecido["frequencia_real"] = 0
                    item_enriquecido["score_combinado"] = _calcular_score(item.get("criticidade", ""), 0)
                    item_enriquecido["origem"] = "NORMATIVO"
                    cat_recuperada["itens"].append(item_enriquecido)
                todas_categorias_enriquecidas.append(cat_recuperada)

    # ── Safeguard lacunas: garantir que itens de COMENTÁRIOS não sejam perdidos ──
    lacunas_passo_a = mapeamento.get("lacunas", [])
    if lacunas_passo_a:
        # Verificar quais lacunas foram adicionadas como itens COMENTARIOS
        descs_existentes = set()
        for cat in todas_categorias_enriquecidas:
            for item in cat.get("itens", []):
                if "COMENT" in (item.get("origem", "") or "").upper():
                    descs_existentes.add((item.get("descricao", "") or "").lower().strip()[:60])

        lacunas_faltantes = []
        for lac in lacunas_passo_a:
            desc_lac = (lac.get("descricao", "") or "").lower().strip()[:60]
            # Verificar se já existe item COMENTARIOS com descrição similar
            encontrada = (
                desc_lac in descs_existentes
                or any(desc_lac[:30] in d for d in descs_existentes)
                or any(d[:30] in desc_lac for d in descs_existentes if len(d) > 10)
            )
            if not encontrada:
                lacunas_faltantes.append(lac)

        if lacunas_faltantes:
            log.warning(f"  ⚠ {len(lacunas_faltantes)} lacunas de comentários não foram adicionadas pela IA – injetando")
            for lac in lacunas_faltantes:
                cat_sug = lac.get("categoria_sugerida", "CAT-01")
                crit = lac.get("criticidade_sugerida", "IMPORTANTE")
                freq = lac.get("frequencia", 1)
                desc = lac.get("descricao", "")
                coments = lac.get("comentarios_relacionados", [])

                # Encontrar a categoria mais adequada
                cat_target = None
                for cat in todas_categorias_enriquecidas:
                    if cat_sug in (cat.get("id", "") or ""):
                        cat_target = cat
                        break
                if not cat_target and todas_categorias_enriquecidas:
                    cat_target = todas_categorias_enriquecidas[0]

                if cat_target:
                    # Buscar evidência nos comentários referenciados
                    ev_resumos = []
                    for cid in coments:
                        idx = int(cid.replace("C", "")) - 1 if cid.startswith("C") and cid[1:].isdigit() else -1
                        if 0 <= idx < len(comentarios):
                            c = comentarios[idx]
                            ev_resumos.append(f"{c.get('autor','?')}: {c['texto'][:150]}")

                    item_lacuna = {
                        "descricao": desc,
                        "detalhes": f"Identificado em comentários de revisores CSD-MPD ({freq}x mencionado)",
                        "referencia_normativa": "Identificado via revisão de pares",
                        "criticidade": crit,
                        "aplicabilidade": "Todas as SEQOPs MPD",
                        "evidencia_comentarios": "; ".join(ev_resumos[:3]) if ev_resumos else f"Mencionado {freq}x em comentários de revisores",
                        "licoes_aprendidas": "",
                        "frequencia_real": freq,
                        "score_combinado": _calcular_score(crit, freq),
                        "origem": "COMENTARIOS",
                    }
                    cat_target["itens"].append(item_lacuna)
                    log.info(f"    + Lacuna injetada em {cat_target.get('id', '?')}: {desc[:80]}")
        else:
            log.info(f"  ✓ Todas as {len(lacunas_passo_a)} lacunas de comentários foram preservadas pela IA")

    # ── Re-numerar IDs ───────────────────────────────────────────────
    for ci, cat in enumerate(todas_categorias_enriquecidas, 1):
        cat["id"] = f"CAT-{ci:02d}"
        for ii, item in enumerate(cat.get("itens", []), 1):
            item["id"] = f"CAT-{ci:02d}-{ii:03d}"

    # ── Montar checklist combinado final ─────────────────────────────
    # Estatísticas
    total_itens = sum(len(c.get("itens", [])) for c in todas_categorias_enriquecidas)
    itens_normativo = sum(
        1 for c in todas_categorias_enriquecidas
        for i in c.get("itens", [])
        if i.get("origem", "NORMATIVO") == "NORMATIVO"
    )
    itens_licoes = sum(
        1 for c in todas_categorias_enriquecidas
        for i in c.get("itens", [])
        if "LICOES" in (i.get("origem", "") or "").upper()
    )
    itens_comentarios = total_itens - itens_normativo - itens_licoes
    itens_com_evidencia = sum(
        1 for c in todas_categorias_enriquecidas
        for i in c.get("itens", [])
        if i.get("frequencia_real", 0) > 0
    )
    itens_com_licoes = sum(
        1 for c in todas_categorias_enriquecidas
        for i in c.get("itens", [])
        if (i.get("licoes_aprendidas") or "").strip()
    )
    pocos_set = set(d.get("poco", "") for d in dados_comentarios if d.get("poco"))

    resultado_final = {
        "titulo": "Checklist Combinado de Revisão de SEQOPs – CSD-MPD",
        "subtitulo": "Normas Petrobras + Comentários Reais + Lições Aprendidas",
        "versao": "2.0",
        "data_geracao": datetime.now().strftime('%d/%m/%Y'),
        "baseado_em": (
            f"{len(checklist_norm.get('categorias',[]))} categorias normativas + "
            f"{len(comentarios)} comentários de {len(pocos_set)} poços + "
            f"{total_licoes_usadas} lições aprendidas"
        ),
        "categorias": todas_categorias_enriquecidas,
        "observacoes_gerais": [
            "Este checklist combina requisitos normativos Petrobras com evidências de comentários reais de revisores CSD-MPD e lições aprendidas operacionais.",
            "Itens com score ALTO devem receber atenção prioritária na revisão.",
            "Itens marcados como 'COMENTARIOS' foram identificados exclusivamente em revisões reais.",
            "Itens marcados como 'LICOES' foram identificados a partir de lições aprendidas, alertas técnicos e boas práticas.",
            f"Baseado em {len(comentarios)} comentários de {len(pocos_set)} poços, {len(checklist_norm.get('documentos_fonte',[]))} documentos normativos e {total_licoes_usadas} lições aprendidas.",
            "Todos os itens devem ser validados por especialista antes de uso em produção.",
        ],
        "documentos_fonte": checklist_norm.get("documentos_fonte", []),
        "mapeamento_ia": mapeamento.get("estatisticas", {}),
        "lacunas_identificadas": mapeamento.get("lacunas", []),
        "_estatisticas": {
            "total_itens": total_itens,
            "itens_normativo": itens_normativo,
            "itens_comentarios": itens_comentarios,
            "itens_licoes": itens_licoes,
            "itens_com_evidencia": itens_com_evidencia,
            "itens_com_licoes": itens_com_licoes,
            "total_comentarios": len(comentarios),
            "total_pocos": len(pocos_set),
            "total_licoes_usadas": total_licoes_usadas,
            "modelo": MODELO_IA,
        },
    }

    _salvar_json(resultado_final, COMBINADO_JSON)

    log.info(f"\nFASE 3 completa:")
    log.info(f"  Categorias: {len(todas_categorias_enriquecidas)}")
    log.info(f"  Itens total: {total_itens}")
    log.info(f"    - Normativo: {itens_normativo}")
    log.info(f"    - Comentários (novos): {itens_comentarios}")
    log.info(f"    - Lições (novos): {itens_licoes}")
    log.info(f"    - Com evidência real: {itens_com_evidencia}")
    log.info(f"    - Com lições aprendidas: {itens_com_licoes}")
    log.info(f"  Lições usadas: {total_licoes_usadas}")

    return resultado_final


def _calcular_score(criticidade: str, frequencia: int) -> str:
    """Calcula score combinado a partir de criticidade normativa e frequência real."""
    crit = criticidade.upper() if criticidade else ""
    if ("CRIT" in crit and frequencia > 0) or frequencia >= 3:
        return "ALTO"
    elif "CRIT" in crit or "IMPORT" in crit or frequencia > 0:
        return "MEDIO"
    else:
        return "BAIXO"


# ==========================================================================
#  FASE 4: GERAR EXCEL PROFISSIONAL
# ==========================================================================

def fase4_gerar_excel(checklist: dict):
    """Gera o checklist combinado em formato Excel profissional."""
    log.info("=" * 70)
    log.info("FASE 4a: Gerando Excel combinado")

    if not checklist.get("categorias"):
        log.error("  Checklist vazio!")
        return

    wb = Workbook()

    # ── Aba 1: Checklist Combinado ───────────────────────────────────
    ws = wb.active
    ws.title = "Checklist Combinado MPD"

    # Estilos
    hdr_fill = PatternFill(start_color="0D2137", end_color="0D2137", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    cat_fill = PatternFill(start_color="B8D4E8", end_color="B8D4E8", fill_type="solid")
    cat_font = Font(bold=True, size=11, color="0D2137")
    crit_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    imp_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    rec_fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
    score_alto_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    score_medio_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    score_baixo_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    novo_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")  # itens de comentários
    licao_fill = PatternFill(start_color="E8F8E8", end_color="E8F8E8", fill_type="solid")  # itens de lições
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Título
    ws.merge_cells("A1:J1")
    title_cell = ws.cell(
        row=1, column=1,
        value=checklist.get("titulo", "Checklist Combinado de Revisão de SEQOPs – CSD-MPD")
    )
    title_cell.font = Font(bold=True, size=14, color="0D2137")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtítulo
    ws.merge_cells("A2:J2")
    ws.cell(row=2, column=1,
            value=checklist.get("subtitulo", "Normas Petrobras + Comentários Reais de Revisores")
    ).font = Font(italic=True, size=11, color="2E86AB")

    # Metadados
    ws.merge_cells("A3:J3")
    estat = checklist.get("_estatisticas", {})
    meta = (
        f"Gerado em: {checklist.get('data_geracao', '')} | "
        f"{checklist.get('baseado_em', '')} | "
        f"Versão: {checklist.get('versao', '1.0')} | Modelo: {MODELO_IA}"
    )
    ws.cell(row=3, column=1, value=meta).font = Font(italic=True, size=9, color="666666")

    ws.merge_cells("A4:J4")
    stats_text = (
        f"Itens: {estat.get('total_itens', '?')} total | "
        f"{estat.get('itens_normativo', '?')} normativos | "
        f"{estat.get('itens_comentarios', '?')} novos (comentários) | "
        f"{estat.get('itens_licoes', '?')} novos (lições) | "
        f"{estat.get('itens_com_evidencia', '?')} com evidência real | "
        f"{estat.get('itens_com_licoes', '?')} com lições aprendidas"
    )
    ws.cell(row=4, column=1, value=stats_text).font = Font(italic=True, size=9, color="888888")

    # Cabeçalhos
    colunas = [
        ("ID", 14), ("✓", 5), ("Score", 8), ("Verificação", 50),
        ("Detalhes / Limites", 42), ("Ref. Normativa", 20), ("Criticidade", 13),
        ("Evidência Real", 40), ("Origem", 12), ("Observações", 30),
    ]
    header_row = 6
    for ci, (nome, larg) in enumerate(colunas, 1):
        c = ws.cell(row=header_row, column=ci, value=nome)
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[c.column_letter].width = larg

    row_num = header_row + 1
    categorias = checklist.get("categorias", [])

    for cat in categorias:
        # Linha de categoria
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
        cat_cell = ws.cell(
            row=row_num, column=1,
            value=f"{cat.get('id', '')} – {cat.get('nome', '')} [{cat.get('prioridade', '')}]"
        )
        cat_cell.fill = cat_fill
        cat_cell.font = cat_font
        cat_cell.alignment = Alignment(vertical="center")
        for ci in range(1, 11):
            ws.cell(row=row_num, column=ci).border = border
        ws.row_dimensions[row_num].height = 24
        row_num += 1

        # Descrição
        if cat.get("descricao"):
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
            desc_cell = ws.cell(row=row_num, column=1, value=f"  {cat['descricao']}")
            desc_cell.font = Font(italic=True, size=9, color="444444")
            desc_cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # Itens
        for item in cat.get("itens", []):
            criticidade = (item.get("criticidade", "") or "").upper()
            score = (item.get("score_combinado", "") or "").upper()
            origem = (item.get("origem", "NORMATIVO") or "").upper()
            eh_novo = "COMENT" in origem
            eh_licao = "LICOES" in origem or "LICO" in origem

            # ID
            c_id = ws.cell(row=row_num, column=1, value=item.get("id", ""))
            c_id.border = border
            c_id.alignment = Alignment(horizontal="center", vertical="center")
            c_id.font = Font(size=9)

            # Checkbox
            check = ws.cell(row=row_num, column=2, value="☐")
            check.border = border
            check.alignment = center
            check.font = Font(size=14)

            # Score combinado
            score_cell = ws.cell(row=row_num, column=3, value=score)
            score_cell.border = border
            score_cell.alignment = center
            score_cell.font = Font(bold=True, size=9, color="FFFFFF")
            if "ALTO" in score:
                score_cell.fill = score_alto_fill
            elif "MEDIO" in score:
                score_cell.fill = score_medio_fill
            elif "BAIXO" in score:
                score_cell.fill = score_baixo_fill

            # Verificação
            verif = ws.cell(row=row_num, column=4, value=item.get("descricao", ""))
            verif.border = border
            verif.alignment = wrap
            verif.font = Font(size=10)

            # Detalhes
            det = ws.cell(row=row_num, column=5, value=item.get("detalhes", ""))
            det.border = border
            det.alignment = wrap
            det.font = Font(size=9)

            # Ref normativa
            ref = ws.cell(row=row_num, column=6, value=item.get("referencia_normativa", ""))
            ref.border = border
            ref.alignment = Alignment(wrap_text=True, vertical="center")
            ref.font = Font(size=9)

            # Criticidade
            crit = ws.cell(row=row_num, column=7, value=criticidade)
            crit.border = border
            crit.alignment = center
            crit.font = Font(size=9)
            if "CRIT" in criticidade:
                crit.fill = crit_fill
                crit.font = Font(bold=True, color="FFFFFF", size=9)
            elif "IMPORT" in criticidade:
                crit.fill = imp_fill
            elif "RECOM" in criticidade:
                crit.fill = rec_fill

            # Evidência de comentários reais + lições aprendidas
            evidencia = item.get("evidencia_comentarios", "")
            freq = item.get("frequencia_real", 0)
            licoes_apr = item.get("licoes_aprendidas", "")
            ev_parts = []
            if evidencia:
                ev_prefix = f"[{freq}x] " if freq > 0 else ""
                ev_parts.append(f"{ev_prefix}{evidencia}")
            if licoes_apr:
                ev_parts.append(f"[LIÇÕES] {licoes_apr}")
            ev_text = "\n".join(ev_parts) if ev_parts else ""
            ev_cell = ws.cell(row=row_num, column=8, value=ev_text)
            ev_cell.border = border
            ev_cell.alignment = wrap
            ev_cell.font = Font(size=9, italic=True)

            # Origem
            if eh_licao:
                origem_label = "LIÇÃO"
            elif eh_novo:
                origem_label = "NOVO"
            else:
                origem_label = "NORMA"
            origem_cell = ws.cell(row=row_num, column=9, value=origem_label)
            origem_cell.border = border
            origem_cell.alignment = center
            origem_cell.font = Font(size=8, bold=(eh_novo or eh_licao))
            if eh_licao:
                origem_cell.fill = licao_fill
                for ci_lic in [1, 2, 3, 4, 5]:
                    ws.cell(row=row_num, column=ci_lic).fill = licao_fill
            elif eh_novo:
                origem_cell.fill = novo_fill
                for ci_novo in [1, 2, 3, 4, 5]:
                    ws.cell(row=row_num, column=ci_novo).fill = novo_fill

            # Observações (vazia)
            obs = ws.cell(row=row_num, column=10, value="")
            obs.border = border

            ws.row_dimensions[row_num].height = 50
            row_num += 1

    # Observações gerais
    row_num += 1
    obs_gerais = checklist.get("observacoes_gerais", [])
    if obs_gerais:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
        ws.cell(row=row_num, column=1, value="OBSERVAÇÕES GERAIS").font = Font(bold=True, size=11)
        row_num += 1
        for o in obs_gerais:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
            ws.cell(row=row_num, column=1, value=f"• {o}").alignment = Alignment(wrap_text=True)
            row_num += 1

    ws.auto_filter.ref = f"A{header_row}:J{row_num}"
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Aba 2: Lacunas (itens exclusivos dos comentários) ────────────
    ws2 = wb.create_sheet("Lacunas Identificadas")
    ws2.cell(row=1, column=1, value="Lacunas – Itens cobrados pelos revisores sem base normativa explícita").font = Font(bold=True, size=12)

    lac_headers = [("Descrição", 60), ("Comentários", 15), ("Categoria Sugerida", 20),
                   ("Criticidade", 15), ("Frequência", 12)]
    for ci, (nome, larg) in enumerate(lac_headers, 1):
        c = ws2.cell(row=3, column=ci, value=nome)
        c.fill = hdr_fill
        c.font = hdr_font
        c.border = border
        ws2.column_dimensions[c.column_letter].width = larg

    row2 = 4
    for lac in checklist.get("lacunas_identificadas", []):
        ws2.cell(row=row2, column=1, value=lac.get("descricao", "")).border = border
        ws2.cell(row=row2, column=1).alignment = wrap
        ws2.cell(row=row2, column=2,
                 value=", ".join(lac.get("comentarios_relacionados", []))).border = border
        ws2.cell(row=row2, column=3, value=lac.get("categoria_sugerida", "")).border = border
        ws2.cell(row=row2, column=4, value=lac.get("criticidade_sugerida", "")).border = border
        ws2.cell(row=row2, column=5, value=lac.get("frequencia", 0)).border = border
        ws2.row_dimensions[row2].height = 30
        row2 += 1

    # ── Aba 3: Estatísticas ──────────────────────────────────────────
    ws3 = wb.create_sheet("Estatísticas")
    ws3.cell(row=1, column=1, value="Estatísticas do Checklist Combinado").font = Font(bold=True, size=12)

    estat = checklist.get("_estatisticas", {})
    stats = [
        ("Total de itens", estat.get("total_itens", 0)),
        ("Itens normativos", estat.get("itens_normativo", 0)),
        ("Itens novos (comentários)", estat.get("itens_comentarios", 0)),
        ("Itens novos (lições)", estat.get("itens_licoes", 0)),
        ("Itens com evidência real", estat.get("itens_com_evidencia", 0)),
        ("Itens com lições aprendidas", estat.get("itens_com_licoes", 0)),
        ("Total de comentários analisados", estat.get("total_comentarios", 0)),
        ("Total de lições usadas", estat.get("total_licoes_usadas", 0)),
        ("Poços analisados", estat.get("total_pocos", 0)),
        ("Modelo IA", estat.get("modelo", MODELO_IA)),
    ]
    for ri, (nome, valor) in enumerate(stats, 3):
        ws3.cell(row=ri, column=1, value=nome).font = Font(bold=True)
        ws3.cell(row=ri, column=2, value=str(valor))
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 20

    # Contagem por score
    ri = len(stats) + 5
    ws3.cell(row=ri, column=1, value="Distribuição por Score Combinado").font = Font(bold=True, size=11)
    ri += 1
    score_count = {"ALTO": 0, "MEDIO": 0, "BAIXO": 0}
    for cat in categorias:
        for item in cat.get("itens", []):
            sc = (item.get("score_combinado", "") or "").upper()
            if "ALTO" in sc:
                score_count["ALTO"] += 1
            elif "MEDIO" in sc:
                score_count["MEDIO"] += 1
            else:
                score_count["BAIXO"] += 1
    for label, count in score_count.items():
        ws3.cell(row=ri, column=1, value=f"Score {label}")
        ws3.cell(row=ri, column=2, value=count)
        ri += 1

    try:
        wb.save(COMBINADO_XLSX)
        log.info(f"  Excel salvo: {COMBINADO_XLSX}")
    except PermissionError:
        # Arquivo bloqueado (aberto no Excel?) → salvar com nome alternativo
        backup = COMBINADO_XLSX.with_stem(COMBINADO_XLSX.stem + "_novo")
        log.warning(f"  Arquivo bloqueado! Salvando como: {backup}")
        wb.save(backup)
        log.info(f"  Excel salvo: {backup}")


# ==========================================================================
#  FASE 4b: GERAR HTML INTERATIVO
# ==========================================================================

def fase4_gerar_html(checklist: dict):
    """Gera o checklist combinado em formato HTML interativo."""
    log.info("  Gerando HTML interativo...")

    categorias = checklist.get("categorias", [])
    if not categorias:
        log.error("  Checklist vazio!")
        return

    total_itens = sum(len(cat.get("itens", [])) for cat in categorias)
    criticos = sum(1 for cat in categorias for i in cat.get("itens", [])
                   if "CRIT" in (i.get("criticidade", "") or "").upper())
    importantes = sum(1 for cat in categorias for i in cat.get("itens", [])
                      if "IMPORT" in (i.get("criticidade", "") or "").upper())
    recomendados = total_itens - criticos - importantes

    score_alto = sum(1 for cat in categorias for i in cat.get("itens", [])
                     if "ALTO" in (i.get("score_combinado", "") or "").upper())
    score_medio = sum(1 for cat in categorias for i in cat.get("itens", [])
                      if "MEDIO" in (i.get("score_combinado", "") or "").upper())
    score_baixo = total_itens - score_alto - score_medio

    itens_novos = sum(1 for cat in categorias for i in cat.get("itens", [])
                      if "COMENT" in (i.get("origem", "") or "").upper())
    itens_licoes = sum(1 for cat in categorias for i in cat.get("itens", [])
                       if "LICOES" in (i.get("origem", "") or "").upper())
    itens_com_licoes = sum(1 for cat in categorias for i in cat.get("itens", [])
                          if (i.get("licoes_aprendidas") or "").strip())

    estat = checklist.get("_estatisticas", {})

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{checklist.get('titulo', 'Checklist Combinado MPD')}</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: 'Segoe UI', Tahoma, sans-serif; background: #f0f2f5; color: #333; padding: 20px; }}
        .container {{ max-width: 1300px; margin: 0 auto; }}

        .header {{
            background: linear-gradient(135deg, #0D2137, #2E86AB);
            color: white; padding: 30px; border-radius: 12px; margin-bottom: 20px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        }}
        .header h1 {{ font-size: 1.7em; margin-bottom: 5px; }}
        .header h2 {{ font-size: 1em; font-weight: 400; opacity: 0.85; margin-bottom: 10px; }}
        .header .meta {{ opacity: 0.75; font-size: 0.85em; line-height: 1.5; }}

        .stats {{ display: flex; gap: 10px; margin: 15px 0; flex-wrap: wrap; }}
        .stat {{
            background: white; padding: 12px 18px; border-radius: 10px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08); text-align: center; flex: 1; min-width: 100px;
        }}
        .stat .number {{ font-size: 1.8em; font-weight: bold; color: #0D2137; }}
        .stat .label {{ font-size: 0.78em; color: #888; }}

        .legend {{
            display: flex; gap: 15px; margin: 12px 0; padding: 12px 18px;
            background: white; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.05);
            flex-wrap: wrap; align-items: center;
        }}
        .legend-title {{ font-weight: 600; font-size: 0.85em; color: #555; }}
        .legend-item {{ display: flex; align-items: center; gap: 5px; font-size: 0.8em; }}
        .legend-dot {{ width: 12px; height: 12px; border-radius: 3px; }}

        .filters {{ margin: 12px 0; display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }}
        .filters label {{ font-size: 0.85em; color: #666; margin-right: 5px; }}
        .filter-btn {{
            padding: 6px 14px; border: 1px solid #ccc; background: white; border-radius: 20px;
            cursor: pointer; font-size: 0.82em; transition: all 0.2s;
        }}
        .filter-btn:hover {{ border-color: #0D2137; color: #0D2137; }}
        .filter-btn.active {{ background: #0D2137; color: white; border-color: #0D2137; }}

        .progress-bar {{ height: 8px; background: #e0e0e0; border-radius: 4px; margin: 12px 0; }}
        .progress-fill {{
            height: 100%; background: linear-gradient(90deg, #6bcb77, #2E86AB, #0D2137);
            border-radius: 4px; transition: width 0.4s ease;
        }}
        .progress-text {{ text-align: center; font-size: 0.85em; color: #666; }}

        .category {{
            background: white; border-radius: 10px; margin-bottom: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08); overflow: hidden;
        }}
        .cat-header {{
            background: #B8D4E8; padding: 14px 20px; cursor: pointer;
            display: flex; justify-content: space-between; align-items: center;
            transition: background 0.2s;
        }}
        .cat-header:hover {{ background: #a3c5db; }}
        .cat-header h2 {{ font-size: 1.05em; color: #0D2137; }}
        .cat-header .cat-info {{ display: flex; gap: 8px; align-items: center; }}
        .cat-header .cat-count {{
            font-size: 0.8em; color: #666; background: white; padding: 2px 10px; border-radius: 12px;
        }}

        .badge {{ padding: 3px 10px; border-radius: 12px; font-size: 0.72em; font-weight: 600; }}
        .badge-alta {{ background: #ff6b6b; color: white; }}
        .badge-media {{ background: #ffd93d; color: #333; }}
        .badge-baixa {{ background: #6bcb77; color: white; }}

        .cat-body {{ padding: 0; }}

        .item {{
            padding: 12px 20px; border-bottom: 1px solid #f0f0f0;
            display: flex; gap: 12px; align-items: flex-start; transition: background 0.2s;
        }}
        .item:hover {{ background: #f8f9fa; }}
        .item:last-child {{ border-bottom: none; }}
        .item.novo {{ border-left: 4px solid #2E86AB; background: #f8fcff; }}
        .item.licao {{ border-left: 4px solid #2E7D32; background: #f5faf5; }}
        .item input[type=checkbox] {{
            width: 20px; height: 20px; margin-top: 3px; cursor: pointer;
            accent-color: #0D2137;
        }}
        .item-content {{ flex: 1; }}
        .item-desc {{ font-size: 0.93em; font-weight: 500; line-height: 1.4; }}
        .item-details {{ font-size: 0.82em; color: #666; margin-top: 4px; line-height: 1.3; }}
        .item-ref {{ font-size: 0.78em; color: #0D2137; margin-top: 3px; font-weight: 500; }}
        .item-evidencia {{
            font-size: 0.8em; margin-top: 5px; padding: 6px 10px;
            background: #FFF8E1; border-left: 3px solid #FFB300; border-radius: 4px;
            color: #6D4C00;
        }}
        .item-evidencia.sem {{ background: #f5f5f5; border-left-color: #ddd; color: #999; }}
        .item .score {{
            padding: 3px 8px; border-radius: 6px; font-size: 0.65em;
            font-weight: 700; white-space: nowrap; flex-shrink: 0;
        }}
        .score-ALTO {{ background: #C0392B; color: white; }}
        .score-MEDIO {{ background: #E67E22; color: white; }}
        .score-BAIXO {{ background: #27AE60; color: white; }}
        .item .crit {{
            padding: 3px 8px; border-radius: 10px; font-size: 0.65em;
            font-weight: 600; white-space: nowrap; flex-shrink: 0;
        }}
        .crit-CRITICA {{ background: #ff6b6b; color: white; }}
        .crit-IMPORTANTE {{ background: #ffd93d; color: #333; }}
        .crit-RECOMENDADA {{ background: #6bcb77; color: white; }}
        .item .tags {{ display: flex; flex-direction: column; gap: 4px; flex-shrink: 0; }}
        .item .tag-novo {{
            font-size: 0.65em; background: #2E86AB; color: white;
            padding: 2px 8px; border-radius: 10px; font-weight: 600;
        }}
        .checked .item-desc {{ text-decoration: line-through; color: #999; }}
        .checked .item-details {{ color: #ccc; }}
        .hidden {{ display: none !important; }}

        .footer {{ text-align: center; color: #aaa; font-size: 0.78em; margin-top: 30px; padding: 20px; }}

        .obs-gerais {{
            background: white; border-radius: 10px; padding: 20px; margin-top: 15px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }}
        .obs-gerais h3 {{ color: #0D2137; margin-bottom: 10px; font-size: 1em; }}
        .obs-gerais li {{ margin: 6px 0; padding-left: 10px; font-size: 0.88em; color: #555; }}

        @media print {{
            body {{ background: white; padding: 0; }}
            .header {{ background: #0D2137 !important; -webkit-print-color-adjust: exact; }}
            .filters, .progress-bar, .progress-text, .legend {{ display: none; }}
            .category {{ page-break-inside: avoid; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📋 {checklist.get('titulo', 'Checklist Combinado MPD')}</h1>
            <h2>{checklist.get('subtitulo', '')}</h2>
            <div class="meta">
                Gerado em: {checklist.get('data_geracao', '')} |
                {checklist.get('baseado_em', '')} |
                Versão: {checklist.get('versao', '1.0')} | Modelo: {MODELO_IA}
            </div>
        </div>

        <div class="stats">
            <div class="stat"><div class="number">{total_itens}</div><div class="label">Total Itens</div></div>
            <div class="stat"><div class="number">{len(categorias)}</div><div class="label">Categorias</div></div>
            <div class="stat"><div class="number" style="color:#C0392B">{score_alto}</div><div class="label">Score ALTO</div></div>
            <div class="stat"><div class="number" style="color:#E67E22">{score_medio}</div><div class="label">Score MEDIO</div></div>
            <div class="stat"><div class="number" style="color:#27AE60">{score_baixo}</div><div class="label">Score BAIXO</div></div>
            <div class="stat"><div class="number" style="color:#2E86AB">{itens_novos}</div><div class="label">Novos (Coment.)</div></div>
            <div class="stat"><div class="number" style="color:#2E7D32">{itens_licoes}</div><div class="label">Novos (Lições)</div></div>
            <div class="stat"><div class="number" style="color:#6D4C41">{itens_com_licoes}</div><div class="label">Com Lições</div></div>
        </div>

        <div class="legend">
            <span class="legend-title">Legenda:</span>
            <span class="legend-item"><span class="legend-dot" style="background:#C0392B"></span> Score ALTO = normativo critico + evidencia real</span>
            <span class="legend-item"><span class="legend-dot" style="background:#E67E22"></span> Score MEDIO = importante ou com evidencia</span>
            <span class="legend-item"><span class="legend-dot" style="background:#27AE60"></span> Score BAIXO = recomendado, sem evidencia</span>
            <span class="legend-item"><span class="legend-dot" style="background:#2E86AB"></span> NOVO = item exclusivo dos comentarios</span>
            <span class="legend-item"><span class="legend-dot" style="background:#2E7D32"></span> LIÇÃO = item de lições aprendidas</span>
        </div>

        <div class="filters">
            <label>Filtrar:</label>
            <button class="filter-btn active" onclick="filterItems('ALL')">Todos</button>
            <button class="filter-btn" onclick="filterItems('SCORE-ALTO')">🔴 Score ALTO</button>
            <button class="filter-btn" onclick="filterItems('SCORE-MEDIO')">🟠 Score MEDIO</button>
            <button class="filter-btn" onclick="filterItems('NOVO')">🔵 Novos (Coment.)</button>
            <button class="filter-btn" onclick="filterItems('LICAO')">🟢 Lições</button>
            <button class="filter-btn" onclick="filterItems('CRITICA')">⚠️ Críticos</button>
            <button class="filter-btn" onclick="filterItems('UNCHECKED')">⬜ Pendentes</button>
        </div>

        <div class="progress-bar"><div class="progress-fill" id="progress" style="width: 0%"></div></div>
        <div class="progress-text" id="progressText">0 / {total_itens} verificados (0%)</div>
"""

    for cat in categorias:
        prioridade = (cat.get("prioridade", "MEDIA") or "MEDIA").upper()
        badge_cls = "badge-alta" if "ALT" in prioridade else ("badge-media" if "MED" in prioridade else "badge-baixa")
        n_itens_cat = len(cat.get("itens", []))

        html += f"""
        <div class="category">
            <div class="cat-header" onclick="toggleCategory(this)">
                <h2>{cat.get('id', '')} – {cat.get('nome', '')}</h2>
                <div class="cat-info">
                    <span class="cat-count">{n_itens_cat} itens</span>
                    <span class="badge {badge_cls}">{prioridade}</span>
                </div>
            </div>
            <div class="cat-body">
"""
        if cat.get("descricao"):
            html += f'                <div style="padding:8px 20px;font-size:0.85em;color:#666;font-style:italic;background:#f8f9fa">{cat["descricao"]}</div>\n'

        for item in cat.get("itens", []):
            crit = (item.get("criticidade", "RECOMENDADA") or "RECOMENDADA").upper()
            score = (item.get("score_combinado", "BAIXO") or "BAIXO").upper()
            ref = item.get("referencia_normativa", "") or ""
            detalhes = item.get("detalhes", "") or ""
            evidencia = item.get("evidencia_comentarios", "") or ""
            freq = item.get("frequencia_real", 0) or 0
            origem = (item.get("origem", "NORMATIVO") or "").upper()
            eh_novo = "COMENT" in origem
            eh_licao = "LICOES" in origem or "LICO" in origem
            licoes_apr = item.get("licoes_aprendidas", "") or ""

            item_class = "item licao" if eh_licao else ("item novo" if eh_novo else "item")

            html += f"""
                <div class="{item_class}" data-crit="{crit}" data-score="{score}" data-novo="{str(eh_novo).lower()}" data-licao="{str(eh_licao).lower()}" data-tem-licoes="{str(bool(licoes_apr.strip())).lower()}">
                    <input type="checkbox" onchange="toggleItem(this)">
                    <div class="item-content">
                        <div class="item-desc">{item.get('descricao', '')}</div>
"""
            if detalhes:
                html += f'                        <div class="item-details">{detalhes}</div>\n'
            if ref:
                html += f'                        <div class="item-ref">📋 {ref}</div>\n'

            # Evidência de comentários
            if evidencia and "sem evidência" not in evidencia.lower() and "sem evidencia" not in evidencia.lower():
                freq_tag = f"[{freq}x] " if freq > 0 else ""
                html += f'                        <div class="item-evidencia">💬 {freq_tag}{evidencia}</div>\n'
            else:
                html += f'                        <div class="item-evidencia sem">💬 Sem evidência direta nos comentários coletados</div>\n'

            # Lições aprendidas
            if licoes_apr and licoes_apr.strip():
                html += f'                        <div class="item-evidencia" style="background:#E8F5E9;border-left-color:#2E7D32;color:#1B5E20">📚 {licoes_apr}</div>\n'

            html += f"""                    </div>
                    <div class="tags">
                        <span class="score score-{score}">{score}</span>
                        <span class="crit crit-{crit}">{crit}</span>
"""
            if eh_licao:
                html += f'                        <span class="tag-novo" style="background:#2E7D32">LIÇÃO</span>\n'
            elif eh_novo:
                html += f'                        <span class="tag-novo">NOVO</span>\n'

            html += "                    </div>\n                </div>\n"

        html += "            </div>\n        </div>\n"

    # Observações gerais
    obs = checklist.get("observacoes_gerais", [])
    if obs:
        html += '        <div class="obs-gerais"><h3>📝 Observações Gerais</h3><ul>\n'
        for o in obs:
            html += f"            <li>{o}</li>\n"
        html += "        </ul></div>\n"

    html += f"""
        <div class="footer">
            Checklist Combinado CSD-MPD – Normas Petrobras + Comentários Reais + Lições Aprendidas<br>
            Modelo: {MODELO_IA} | {datetime.now().strftime('%d/%m/%Y %H:%M')} |
            {checklist.get('baseado_em', '')}
        </div>
    </div>

    <script>
        const total = {total_itens};
        let checkedCount = 0;
        let currentFilter = 'ALL';

        function toggleItem(cb) {{
            const item = cb.closest('.item');
            if (cb.checked) {{ item.classList.add('checked'); checkedCount++; }}
            else {{ item.classList.remove('checked'); checkedCount--; }}
            updateProgress();
            if (currentFilter === 'UNCHECKED') filterItems('UNCHECKED');
        }}

        function toggleCategory(header) {{
            const body = header.nextElementSibling;
            body.style.display = body.style.display === 'none' ? '' : 'none';
        }}

        function updateProgress() {{
            const pct = total > 0 ? (checkedCount / total * 100) : 0;
            document.getElementById('progress').style.width = pct.toFixed(1) + '%';
            document.getElementById('progressText').textContent =
                checkedCount + ' / ' + total + ' verificados (' + pct.toFixed(0) + '%)';
        }}

        function filterItems(filter) {{
            currentFilter = filter;
            document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
            event.target.classList.add('active');

            document.querySelectorAll('.item').forEach(item => {{
                const crit = item.dataset.crit;
                const score = item.dataset.score;
                const isNovo = item.dataset.novo === 'true';
                const isLicao = item.dataset.licao === 'true';
                const isChecked = item.querySelector('input').checked;

                if (filter === 'ALL') item.classList.remove('hidden');
                else if (filter === 'UNCHECKED') item.classList.toggle('hidden', isChecked);
                else if (filter === 'NOVO') item.classList.toggle('hidden', !isNovo);
                else if (filter === 'LICAO') item.classList.toggle('hidden', !isLicao);
                else if (filter.startsWith('SCORE-')) item.classList.toggle('hidden', score !== filter.replace('SCORE-',''));
                else item.classList.toggle('hidden', crit !== filter);
            }});

            document.querySelectorAll('.category').forEach(cat => {{
                const visible = cat.querySelectorAll('.item:not(.hidden)');
                cat.style.display = visible.length > 0 ? '' : 'none';
            }});
        }}
    </script>
</body>
</html>"""

    COMBINADO_HTML.write_text(html, encoding="utf-8")
    log.info(f"  HTML salvo: {COMBINADO_HTML}")


# ==========================================================================
#  PIPELINE PRINCIPAL
# ==========================================================================

def main():
    global MODELO_IA, MODELOS_FALLBACK

    parser = argparse.ArgumentParser(
        description="Gerador de Checklist COMBINADO para Revisão de SEQOPs (CSD-MPD) – Abordagem 3"
    )
    parser.add_argument("--enrich", action="store_true",
                        help="Enriquecer: coletar mais comentários via Selenium")
    parser.add_argument("--max-urls", type=int, default=0,
                        help="Limite de URLs para enriquecer (0 = todas pendentes)")
    parser.add_argument("--skip-enrich", action="store_true",
                        help="Usar comentários já salvos sem Selenium")
    parser.add_argument("--modelo", type=str, default=MODELO_IA,
                        help=f"Modelo de IA (default: {MODELO_IA})")
    parser.add_argument("--normativo", type=str, default=str(NORMATIVO_JSON),
                        help="Caminho do checklist normativo JSON")
    parser.add_argument("--comentarios", type=str, default=str(COMENTARIOS_JSON),
                        help="Caminho dos comentários JSON")
    parser.add_argument("--licoes", type=str, default=str(LICOES_JSON),
                        help="Caminho do JSON de lições aprendidas")
    parser.add_argument("--sem-licoes", action="store_true",
                        help="Ignorar lições aprendidas (usar apenas normativo + comentários)")
    args = parser.parse_args()

    if args.modelo != MODELO_IA:
        MODELO_IA = args.modelo
        MODELOS_FALLBACK = [MODELO_IA] + [m for m in MODELOS_FALLBACK if m != MODELO_IA]

    normativo_path = Path(args.normativo)
    comentarios_path = Path(args.comentarios)
    licoes_path = None if args.sem_licoes else Path(args.licoes)

    log.info("=" * 70)
    log.info("GERADOR DE CHECKLIST COMBINADO – Abordagem 3 v2.0")
    log.info("Normas Petrobras + Comentários Reais + Lições Aprendidas MPD")
    log.info(f"Data/hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    log.info(f"Modelo IA: {MODELO_IA}")
    log.info(f"Lições aprendidas: {'desabilitado' if args.sem_licoes else licoes_path}")
    log.info("=" * 70)

    # ── Fase 1: Carregar dados ───────────────────────────────────────
    checklist_norm, dados_comentarios, dados_licoes = fase1_carregar_dados(
        normativo_path, comentarios_path, licoes_path
    )

    if not checklist_norm:
        log.error("Não foi possível carregar o checklist normativo. Abortando.")
        return

    # ── Fase 2: Enriquecer (opcional) ────────────────────────────────
    if args.enrich:
        dados_comentarios = fase2_enriquecer(dados_comentarios, max_urls=args.max_urls)
    elif not dados_comentarios:
        log.warning("Sem comentários. Use --enrich para coletar ou forneça --comentarios.")
        log.info("Continuando apenas com checklist normativo + lições...")

    # ── Fase 3: Cruzamento IA ────────────────────────────────────────
    checklist_combinado = fase3_cruzamento(checklist_norm, dados_comentarios, dados_licoes)

    if not checklist_combinado.get("categorias"):
        log.error("Checklist combinado vazio! Abortando.")
        return

    # ── Fase 4: Gerar saídas ─────────────────────────────────────────
    fase4_gerar_excel(checklist_combinado)
    fase4_gerar_html(checklist_combinado)

    # ── Resumo final ─────────────────────────────────────────────────
    estat = checklist_combinado.get("_estatisticas", {})
    n_cats = len(checklist_combinado.get("categorias", []))
    n_itens = estat.get("total_itens", 0)

    log.info("\n" + "=" * 70)
    log.info("CHECKLIST COMBINADO GERADO COM SUCESSO!")
    log.info(f"  Categorias:          {n_cats}")
    log.info(f"  Itens total:         {n_itens}")
    log.info(f"    - Normativos:      {estat.get('itens_normativo', 0)}")
    log.info(f"    - Novos (coment.): {estat.get('itens_comentarios', 0)}")
    log.info(f"    - Novos (lições):  {estat.get('itens_licoes', 0)}")
    log.info(f"    - Com evidência:   {estat.get('itens_com_evidencia', 0)}")
    log.info(f"    - Com lições:      {estat.get('itens_com_licoes', 0)}")
    log.info(f"  Comentários usados:  {estat.get('total_comentarios', 0)}")
    log.info(f"  Lições usadas:       {estat.get('total_licoes_usadas', 0)}")
    log.info(f"  Poços analisados:    {estat.get('total_pocos', 0)}")
    log.info(f"  Excel: {COMBINADO_XLSX}")
    log.info(f"  HTML:  {COMBINADO_HTML}")
    log.info(f"  JSON:  {COMBINADO_JSON}")
    log.info("=" * 70)


if __name__ == "__main__":
    main()
